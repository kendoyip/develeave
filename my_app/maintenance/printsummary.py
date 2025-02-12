import pandas as pd

import gridfs
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.styles import borders
from openpyxl.styles.borders import Border
from openpyxl.styles.alignment import Alignment

from io import BytesIO 
from bson.objectid import ObjectId

import calendar
from datetime import date, datetime, timedelta 

import pymongo
from pymongo import MongoClient

# ** MMGAPP CONFIG FILE ** #
import sys; sys.path.append('F:\\mmgapp\\config'); import configpy; config = configpy.config()


# Access environment variables FOR mongoDB
connection = config.mongoDBConfig['mongoDB']
client = MongoClient(connection)

db = client['eleave']
#########################################################################################################
## e-leave
#########################################################################################################

#db

eleaveDtl = db["eleave_dtl"]
holidays = db["holidays"]
leaveTypes = db["leave_types"]
leaveGroups = db["leave_groups"]
maintenance = db["eleave_maintenance"]
reportMap = db["fileDirectory"]


#Global Constant
status = list(maintenance.find({"table": { "$eq" : "globalConstant"}}))
df = pd.DataFrame(status)

## parameters
#leaveOffice = "HKG"
#leaveYear = 2022
#leaveRacf = 'NF1KWY'
#leaveType = "LVE02"
#leaveType = "Casual Leave"
#approver = "NF1VCC"
#leaveType = "Work From Home"
#leaveApplying = [{"startDate": "2022-07-19", "startTime": "AM", "endDate": "2022-07-19", "endTime": "PM"},
#                 {"startDate" : "2022-07-20", "startTime": "PM", "endDate": "2022-07-20", "endTime": "PM"},
#                 { "startDate" : "2022-08-01", "startTime": "AM", "endDate" : "2022-08-02", "endTime": "AM"}
#                ]
#leaveApplyingScreen =  [   {"startDate": "2022-07-19", "startTime": "Full Day", "endDate": "2022-07-19", "endTime": "Full Day"},
#                            {"startDate": "2022-07-20", "startTime": "Half Day - PM", "endDate": "2022-07-20", "endTime": "Half Day - PM"},
#                            {"startDate": "2022-08-01", "startTime": "Full Day", "endDate": "2022-08-02", "endTime": "Half Day - AM"}
#                        ]
#leaveApplying = [{"startDate": "2022-07-04", "startTime": "AM", "endDate": "2022-07-04", "endTime": "AM"}]
#leaveApplyingScreen = [{ "startDate": "2022-07-04", "startTime": "Half Day - AM", "endDate": "2022-07-15", "endTime": "Half Day - AM"}]


leaveTypeLst = []
leaveGroupLst = []



def getLeaveTypes():
    global leaveTypeLst
    leaveTypeLst = list(leaveTypes.find({}))

def getStaffRecord (psRacf):
    if len(psRacf) > 0 :
        staffRecord = eleaveDtl.find_one ( {"staff.racf" : { '$regex' : psRacf, '$options' : "i"} , "staff.status": { '$regex': "ACTIVE", '$options': "i"} } )
        return(staffRecord)
    else:
        return None
    
def getLeave(psYear, psLeaveType, psLeaveStatus, psRecord):
    return (list(filter(lambda r: (r["type"].upper() == psLeaveType.upper() and r["applicationStatus"].upper() == psLeaveStatus.upper() and r["year"] == psYear), psRecord["leave_record"])))


def countLeave (psYear, psLeaveType, psLeaveStatus, psRecord):
    leaveDays = 0
    for record in getLeave(psYear, psLeaveType, psLeaveStatus, psRecord):
        for leaveDetails in record["details"]:
            leaveDays += leaveDetails["no_of_workday"]

    return leaveDays

def str2Date (psDateStr):
    return datetime.strptime(psDateStr, "%Y-%m-%d")

def getLeaveEntitlement(psYear, psLeaveTypeAttr, psRecord):
    entitlementLst = [ ]
    entitlement = {
        "leaveEntitle": 0,
        "carryForward": 0,
        "forfeitDate": str2Date("2000-01-01")
    }
    for e in psRecord["entitlement"]:
        if e["year"] == psYear:
            if psLeaveTypeAttr.get("entitlement_field", "") != "":
                leaveEntitle = e.get(psLeaveTypeAttr.get("entitlement_field"), 0)
            else:
                leaveEntitle = 0
            if psLeaveTypeAttr.get("carry_forward_field", "") != "":
                carryForward = e.get(psLeaveTypeAttr.get("carry_forward_field"), 0)
            else:
                carryForward = 0
            if psLeaveTypeAttr.get("forfeit_date_field", "") != "":
                forfeitDate = str2Date(e.get(psLeaveTypeAttr.get("forfeit_date_field"), "2000-01-01"))
            else:
                forfeitDate = str2Date("2000-01-01")
            entitlement = {
                    "leaveEntitle": leaveEntitle,
                    "carryForward": carryForward,
                    "forfeitDate": forfeitDate
            }
    entitlementLst.append(entitlement)  

    return (entitlementLst)

def getLeaveHistory(psYearStart, psYearEnd, psRecord):
    leaveHistoryAllLst = [ ]
    yr = psYearStart
    while yr <= psYearEnd:
        for r in psRecord["leave_record"]:
            if r["year"] == yr:
                for d in r["details"]:
                    for p in d["period"]:
                        currRecord = {
                            "ref_no": r["ref_no"],
                            "office": psRecord["staff"]["hr_office"],
                            "racf": psRecord["staff"]["racf"],
                            "staffname": psRecord["staff"]["name"],
                            "empID": psRecord["staff"]["empID"],
                            "dept": psRecord["staff"]["dept"],
                            "position": psRecord["staff"]["position"],
                            "year" : r["year"],
                            "type": r["type"],
                            "sharePointId": r["sharePointId"],
                            "startDate": d["start_date"],
                            "startTime": d["start_time"],
                            "endDate": d["end_date"],
                            "endTime": d["end_time"],
                            "applicationStatus": r["applicationStatus"],
                            "approvalStatus": r["approvalStatus"],
                            "workDay": d["no_of_workday"],
                            "calendarDay": d["no_of_calendarday"],
                            "submitDate": r["submit_date"],
                            "ldate": str2Date(p["ldate"]),
                            "ltime": p["ltime"],
                            "approver1": psRecord["staff"]["approver1"],
                            "approver2": psRecord["staff"]["approver2"],
                            "approver3": psRecord["staff"]["approver3"],
                            "lastUpdate": r["lastUpdate"],
                            "updateDate": r["updateDate"]                            
                        }
                      
                        leaveHistoryAllLst.append (currRecord)
        yr += 1
    return (leaveHistoryAllLst)

def getMMDDYYYY(psDateString):
    return (datetime.strftime(str2Date(psDateString), "%m/%d/%Y"))


def checkBalance(psYear, psLeaveTypeAttr, psRecord, psApplyingLeaveSlotLst):
    leaveEntitleLst = getLeaveEntitlement(psYear, psLeaveTypeAttr, psRecord)
    leaveHistoryLst = getLeaveHistory(psYear, psYear, psRecord)
    leaveHistoryLst = list(filter(lambda r: (r["applicationStatus"].upper() != df['gcStatusCancel'][0] and r["applicationStatus"].upper() != df['gcStatusReject'][0]), leaveHistoryLst))
    taken = 0
    
    leaveHistoryTypeLst = list(filter(lambda r: (r["type"].upper() == psLeaveTypeAttr.get("leave_type_id").upper() and r["year"] == psYear), leaveHistoryLst))

    for lve in leaveHistoryTypeLst:
            taken += 0.5

    for apply in psApplyingLeaveSlotLst:
            taken += 0.5

    return (leaveEntitleLst[0]["leaveEntitle"] + leaveEntitleLst[0]["carryForward"] - taken)

def getDisplayLeaveYear(psYear):
    if calendar.isleap(psYear):
        return (df['gcYearStartDate'][0] + str(psYear) + " - " + df['gcYearEndDateLeap'][0] + str(psYear + 1))
    else:
        return(df['gcYearStartDate'][0] + str(psYear) + " - " + df['gcYearEndDate'][0] + str(psYear + 1))      

def getDisplayRefNo(psRefNo, psOffice, psRacf):
    return(psOffice + str(psRefNo) + psRacf[-3:])


def listLeave (psInput):
    getLeaveTypes()
    psRacf = psInput.get("racf", "")
    psYear = psInput.get("year", 0)

    if len(psRacf) == 0 or psYear == 0:
        return ({"pass": False, "error_message" : "Incorrect parameters", "result": None, "Status_code": 505})
    displayLeaveHistoryDtl = [ ]
    displayLeaveHistoryHdr = [ ]
    
    staffRecord = getStaffRecord(psRacf)
    if not isinstance(staffRecord, dict):
        return ({"pass": False, "error_message" : "Staff Record Not Exist", "result": None, "Status_code": 504}) 

    for lveType in leaveTypeLst:
        leaveTypeHdr = {
            "leaveType": lveType.get("leave_type"),
            "leaveTypeId": lveType.get("leave_type_id"),
            "taken" : countLeave(psYear, lveType.get("leave_type_id"), df['gcStatusApproved'][0], staffRecord),
            "pending": countLeave(psYear, lveType.get("leave_type_id"), df['gcStatusPending'][0], staffRecord),
            "balance": checkBalance(psYear, lveType, staffRecord, [])
        }
        displayLeaveHistoryHdr.append(leaveTypeHdr)    
        
    leaveHistoryLst = getLeaveHistory(psYear, psYear, staffRecord)
    currRefNo = 0
    currStartDate = ""
    currStartTime = ""
    for lve in leaveHistoryLst:
        if (currRefNo == lve["ref_no"] and currStartDate != lve["startDate"] and currStartTime != lve["startTime"]) or (currRefNo != lve["ref_no"]):
            displayLeaveRecord = {
                    "submitDate":  getMMDDYYYY(lve["submitDate"]),
                    "refNo": getDisplayRefNo(lve["ref_no"], lve["office"], lve["racf"]),
                    "office": lve["office"],
                    "staffname": lve["staffname"],
                    "empID": lve["empID"],
                    "dept": lve["dept"],
                    "position": lve["position"],
                    "type_id": lve["type"],
                    "sharePointId": lve["sharePointId"],
                    "type" : list(filter(lambda r: (r["leave_type_id"].upper() == lve["type"]), leaveTypeLst))[0].get("leave_type"),
                    "year": getDisplayLeaveYear(lve["year"]),
                    "leaveFrom": getMMDDYYYY(lve["startDate"]),
                    "startPeriod": lve["startTime"],
                    "leaveTo": getMMDDYYYY(lve["endDate"]),
                    "endPeriod": lve["endTime"],
                    "workday": lve["workDay"],
                    "calendarDay": lve["calendarDay"],
                    "applicationStatus": lve["applicationStatus"],
                    "approver1": lve["approver1"],
                    "approver2": lve["approver2"],
                    "approver3": lve["approver3"],
                    "approvalStatus" : lve["approvalStatus"],
                    "lastUpdate": lve["lastUpdate"],
                    "updateDate": getMMDDYYYY(lve["updateDate"])
            }
            displayLeaveHistoryDtl.append(displayLeaveRecord)
            currRefNo = lve["ref_no"]
            currStartDate = lve["startDate"]
            currStartTime = lve["endDate"]

    fullRecord = {
        "header":  displayLeaveHistoryHdr,
        "details": displayLeaveHistoryDtl
    }
    leaveRecordLst = [ ]
    leaveRecordLst.append (fullRecord) 

    if len(leaveRecordLst) == 0:
        return ({"pass": True, "error_message" : None, "result": [], "Status_code": 200}) 
    else:
        return ({"pass": True, "error_message" : None, "result": leaveRecordLst, "Status_code": 200}) 

def genReport(psWS, psRptDict, psRptFormat):
    for lstKey, lstValue in psRptDict.items():
        try:
            if lstKey in psRptFormat["cell"]:
                if (isinstance(lstValue, list)) == False:
                    psWS[(psRptFormat["cell"][lstKey])].value = lstValue
                else:
                    row = 0
                    col = 0
                    for lveRow in lstValue:
                        for key, v in lveRow.items():
                            if psRptFormat["cell"][lstKey]["next_record"] == "Row":
                                mcell = True
                                while mcell == True:
                                    colId = column_index_from_string(coordinate_from_string(psRptFormat["cell"][lstKey]["start_cell"])[0]) + col
                                    rowId = coordinate_from_string(psRptFormat["cell"][lstKey]["start_cell"])[1] + row
                                    if not isinstance(psWS.cell(row=rowId, column=colId), MergedCell):
                                        mcell = False
                                    else:
                                        mcell = True
                                        col += 1                                                                                       
                                psWS.cell(row=rowId, column=colId, value=v)
                                psWS.cell(row=rowId, column=colId).border = Border(left=borders.Side(border_style='thin', color="FF000000", style=None), 
                                right=borders.Side(border_style='thin', color="FF000000", style=None), 
                                top=borders.Side(border_style='thin', color="FF000000", style=None),
                                bottom=borders.Side(border_style='thin', color="FF000000", style=None))
                                col += 1 
                        row += 1
                        col = 0
        except:
            print ("Error found")
            return "Error", 701
    return "OK", 0    



# psInput =  {'year': para['year'], 'racf': para['racf']}    
def apiPrintSummary(psInput):
                   
    result = listLeave(psInput)
 
    rpt = reportMap.find_one ( { "report": "Leave Summary"} )
    
    if (result.get("pass")): 
        hdr = result.get("result")[0]["header"]
        alTaken = 0
        alPending = 0
        alBalance = 0
        clTaken = 0
        clPending = 0
        clBalance = 0
        slTaken = 0
        slPending = 0
        slBalance = 0
        for lve in leaveTypeLst:
            hdrData =  list(filter(lambda r: (r["leaveTypeId"].upper() == lve["leave_type_id"]), hdr))[0]
            if lve["leave_type_id"] == "LVE01":
                alTaken = hdrData["taken"]
                alPending = hdrData["pending"]
                alBalance = hdrData["balance"]
            elif lve["leave_type_id"] == "LVE02":
                clTaken = hdrData["taken"]
                clPending = hdrData["pending"]
                clBalance = hdrData["balance"]
            elif lve["leave_type_id"] == "LVE04":
                slTaken = hdrData["taken"] + slTaken
                slPending = hdrData["pending"] + slPending
                slBalance = hdrData["balance"] + slBalance
            elif lve["leave_type_id"] == "LVE05":
                slTaken = hdrData["taken"] + slTaken
                slPending = hdrData["pending"] + slPending
                slBalance = hdrData["balance"] + slBalance
        rptDtlLst = [ ]
        for record in result.get("result"):
            for dtl in record["details"]:
                rptDtl = {
                    "submitDate": dtl.get("submitDate"),
                    "refNo": dtl.get("refNo"),
                    "office": dtl.get("office"),
                    "staffname": dtl.get("staffname"),
                    "dept": dtl.get("dept"),
                    "type": dtl.get("type"),
                    "year": dtl.get("year"),
                    "leaveFrom": dtl.get("leaveFrom"),
                    "startPeriod": dtl.get("startPeriod"),
                    "leaveTo": dtl.get("leaveTo"),
                    "endPeriod": dtl.get("endPeriod"),
                    "workday": dtl.get("workday"),
                    "calendarDay": dtl.get("calendarDay"),
                    "applicationStatus": dtl.get("applicationStatus"),
                    "lastUpdate": dtl.get("lastUpdate"),
                    "updateDate": dtl.get("updateDate")        
                }
                rptDtlLst.append(rptDtl)

            # Sort by leave start requested by Sandy 1/26/2023
            rptDtlLst.sort(key=lambda x: x.get('leaveFrom'))
            # added by Vincent Cheng on 11/23 
            try:
                if record["details"][0].get("year"):
                    pass
            except:            
                return 501   
    
            report = {
                "hdrCalendarYear": record["details"][0].get("year"),
                "hdrUser": record["details"][0].get("staffname") + "\nLeave Application Summary",
                "hdrALTaken": alTaken,
                "hdrALPending": alPending,
                "hdrALBalance": alBalance,
                "hdrSLTaken": slTaken,
                "hdrSLPending": slPending,
                "hdrCLTaken": clTaken,
                "hdrCLPending": clPending,
                "hdrCLBalance": clBalance,
                "dtl": rptDtlLst
                
            }

        #filename when using in Heroku:
        fs = gridfs.GridFS(db)
        wb = load_workbook(filename=BytesIO(fs.get(ObjectId(rpt["file"]["fileObj"])).read()))

        # filename in development:
        #wb = load_workbook(filename=rpt["file"]["fileName"])
        ws = wb[rpt["file"]["wsName"]]
        
        result = genReport(ws, report, rpt)
        if result[1] == 0:
            out = BytesIO()
            wb.save(out)
            out.seek(0)        
 
            wb.save(filename="C:/Work/eleave/my_app/maintenance/testing.xlsx")
            wb.close()            
            #print('sending file...')
            #return send_file(out,  attachment_filename='a_file.xls', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')      
            return 200
        else:            
            return 501    
      
psInput =  {'year': 2023, 'racf': "NF1BHC"}    
apiPrintSummary(psInput)