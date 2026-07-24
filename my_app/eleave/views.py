from datetime import date, datetime, timedelta 
from dateutil.relativedelta import relativedelta
from flask import jsonify, request, current_app, send_file, Blueprint
from flask import session, request, jsonify
from openpyxl import load_workbook
from openpyxl.cell import MergedCell
from io import BytesIO 
from bson.objectid import ObjectId
from datetime import datetime
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.styles import borders
from openpyxl.styles.borders import Border
from openpyxl.styles.alignment import Alignment
from email.mime.multipart import MIMEMultipart
from email.mime.application import MIMEApplication
from email.mime.text import MIMEText
from copy import copy

import pandas as pd
import gridfs
import mimetypes
import calendar
import smtplib
import json
import requests
import os
import base64
import pytz
from dotenv import load_dotenv
load_dotenv()

import checkLogged
from my_app import db,  client
from dateutil import parser
import time

#ical
import pytz
from icalendar import Calendar, Event,  vText




eleave = Blueprint('eleave', __name__)

#########################################################################################################
## e-leave
#########################################################################################################

#db
eleaveDtl = db["eleave_dtl"]
holidays = db["holidays"]
leaveTypes = db["leave_types"]
leaveGroups = db["leave_groups"]
maintenance = db["eleave_maintenance"]
reportMap = db["fileDirectory"]
otherLeaves = db["other_leaves"]
summer_hours = db["summer_hours"]

#Global Constant
status = list(maintenance.find({"table": { "$eq" : "globalConstant"}}))
df = pd.DataFrame(status)


## parameters
#leaveOffice = "HKG"
#leaveYear = 2022
#leaveRacf = 'NF1KWY'
#leaveType = "LVE02"
#leaveType = "Casual Leave"
#approver = "NF1VCC"
#leaveType = "Work From Home"
#leaveApplying = [{"startDate": "2022-07-19", "startTime": "AM", "endDate": "2022-07-19", "endTime": "PM"},
#                 {"startDate" : "2022-07-20", "startTime": "PM", "endDate": "2022-07-20", "endTime": "PM"},
#                 { "startDate" : "2022-08-01", "startTime": "AM", "endDate" : "2022-08-02", "endTime": "AM"}
#                ]
#leaveApplyingScreen =  [   {"startDate": "2022-07-19", "startTime": "Full Day", "endDate": "2022-07-19", "endTime": "Full Day"},
#                            {"startDate": "2022-07-20", "startTime": "Half Day - PM", "endDate": "2022-07-20", "endTime": "Half Day - PM"},
#                            {"startDate": "2022-08-01", "startTime": "Full Day", "endDate": "2022-08-02", "endTime": "Half Day - AM"}
#                        ]
#leaveApplying = [{"startDate": "2022-07-04", "startTime": "AM", "endDate": "2022-07-04", "endTime": "AM"}]
#leaveApplyingScreen = [{ "startDate": "2022-07-04", "startTime": "Half Day - AM", "endDate": "2022-07-15", "endTime": "Half Day - AM"}]


leaveTypeLst = []
leaveGroupLst = []


@eleave.route('/api/getUserList',methods=['POST'])
@checkLogged.check_logged
def getUserList():            
       
    ## loading staff list  
    query = { "staff":1}        
    query_filter =  {"staff.status": { "$eq" : "ACTIVE"}}                 
    userList = []     
    col = eleaveDtl 
    results = col.find(query_filter, query)    ## eleave_dtl
    for result in results:   
        userList.append({ 'racf': result['staff']['racf'], 'name' : result['staff']['name'], 'office': result['staff']['office']})
    

    try:
        if len(userList) > 0:                             
            return  jsonify({'userList' : userList}), 200   
        else:
            return  jsonify({'error_message' : 'Error to get user list.  Please contact regional PBT !'}), 501
    except:
        return jsonify({'error_message' : 'Error to get user list.  Please contact regional PBT !'}), 501         

## Functions 

def genReport(psWS, psRptDict, psRptFormat):
    for lstKey, lstValue in psRptDict.items():
        try:
            if lstKey in psRptFormat["cell"]:
                if (isinstance(lstValue, list)) == False:
                    psWS[(psRptFormat["cell"][lstKey])].value = lstValue
                else:
                    row = 0
                    col = 0
                    for lveRow in lstValue:
                        for key, v in lveRow.items():
                            if psRptFormat["cell"][lstKey]["next_record"] == "Row":
                                mcell = True
                                while mcell == True:
                                    colId = column_index_from_string(coordinate_from_string(psRptFormat["cell"][lstKey]["start_cell"])[0]) + col
                                    rowId = coordinate_from_string(psRptFormat["cell"][lstKey]["start_cell"])[1] + row
                                    if not isinstance(psWS.cell(row=rowId, column=colId), MergedCell):
                                        mcell = False
                                    else:
                                        mcell = True
                                        col += 1                                                                                       
                                psWS.cell(row=rowId, column=colId, value=v)
                                psWS.cell(row=rowId, column=colId).border = Border(left=borders.Side(border_style='thin', color="FF000000", style=None), 
                                right=borders.Side(border_style='thin', color="FF000000", style=None), 
                                top=borders.Side(border_style='thin', color="FF000000", style=None),
                                bottom=borders.Side(border_style='thin', color="FF000000", style=None))
                                col += 1 
                        row += 1
                        col = 0
        except:
            print ("Error found")
            return "Error", 701
    return "OK", 0    


def genApplyForm(ws, approvalRecordLst, rpt):

    # Basic information
    for excel_range in rpt['cell']:
        # Static information not included the office display and leave entry
        if excel_range != "LeaveDetail"  and excel_range != "applicantName":
            col_index = column_index_from_string(coordinate_from_string(rpt['cell'][excel_range])[0])
            row_index = (coordinate_from_string(rpt['cell'][excel_range])[1])
            ws.cell(row=row_index, column=col_index, value=approvalRecordLst[0][excel_range])
        # Date Join format
        if excel_range == "date_joined":
            date_joinded = datetime.strptime(str(approvalRecordLst[0][excel_range]), '%Y-%m-%d')
            date_joinded = date_joinded.strftime('%m/%d/%Y')
            ws.cell(row=row_index, column=col_index, value=date_joinded)


    # Count number of record
    record_count = len(approvalRecordLst[0]['details'])

    # Handle excel_range = "LeaveDetail"
    for rows in range ((coordinate_from_string(rpt['cell']['LeaveDetail'])[1]), record_count*2 +(coordinate_from_string(rpt['cell']['LeaveDetail'])[1]), 2):
        if rows == 36:
            col_index = column_index_from_string(coordinate_from_string(rpt['cell']['LeaveDetail'])[0])
            row_index = (coordinate_from_string(rpt['cell']['LeaveDetail'])[1])
            # Leave Start Date
            ws.cell(row=row_index, column=col_index, value=getMMDDYYYY(approvalRecordLst[0]['details'][rows-36]['startDate']))
            # Workday Name of start date
            ws.cell(row=row_index, column=col_index+ 3, value= "(" + getWorkdayName(approvalRecordLst[0]['details'][rows-36]['startDate']) + ")")
            # Full Day / AM / PM
            ws.cell(row=row_index, column=col_index+ 6, value=(approvalRecordLst[0]['details'][rows-36]['startTime']))
            # Leave End Date
            ws.cell(row=row_index, column=col_index+ 12, value=getMMDDYYYY(approvalRecordLst[0]['details'][rows-36]['endDate']))
            # Workday Name of end date
            ws.cell(row=row_index, column=col_index+ 15, value= "(" + getWorkdayName(approvalRecordLst[0]['details'][rows-36]['endDate']) + ")")
            # Full Day / AM / PM
            ws.cell(row=row_index, column=col_index+ 18, value=(approvalRecordLst[0]['details'][rows-36]['endTime']))
            # No of Working Days
            ws.cell(row=row_index, column=col_index+ 22, value=(approvalRecordLst[0]['details'][rows-36]['workday']))
            # No of Calendar Day
            ws.cell(row=row_index, column=col_index+ 28, value=(approvalRecordLst[0]['details'][rows-36]['calendarDay']))

        elif rows > 36:
            ws.insert_rows(rows, 2)
            # row and column index to output file
            
            index = int((rows-36)/2)
            col_index = column_index_from_string(coordinate_from_string(rpt['cell']['LeaveDetail'])[0])

            # Formatting, format must be the same as the first row of leave detail
            for n in range (0, ws.max_column):
                ws.cell(row=rows, column= 1 + n).value = copy(ws.cell(row = rows - 2, column= 1 + n).value)
                ws.cell(row=rows + 1, column= 1 + n).value = copy(ws.cell(row = rows - 1, column= 1 + n).value)
                ws.cell(row=rows, column= 1 + n).fill = copy(ws.cell(row = rows - 2, column= 1 + n).fill)
                ws.cell(row=rows + 1, column= 1 + n).fill = copy(ws.cell(row = rows - 1, column= 1 + n).fill)                
                ws.cell(row=rows, column= 1 + n).font = copy(ws.cell(row = rows - 2, column= 1 + n).font)
                ws.cell(row=rows + 1, column= 1 + n).font = copy(ws.cell(row = rows - 1, column= 1 + n).font)    
                ws.cell(row=rows, column= 1 + n).number_format = copy(ws.cell(row = rows - 2, column= 1 + n).number_format)
                ws.cell(row=rows + 1, column= 1 + n).number_format = copy(ws.cell(row = rows - 1, column= 1 + n).number_format)    
                ws.cell(row=rows, column= 1 + n).border = copy(ws.cell(row = rows - 2, column= 1 + n).border)
                ws.cell(row=rows + 1, column= 1 + n).border = copy(ws.cell(row = rows - 1, column= 1 + n).border)
                ws.cell(row=rows, column= 1 + n).alignment = copy(ws.cell(row = rows - 2, column= 1 + n).alignment)
                ws.cell(row=rows + 1, column= 1 + n).alignment = copy(ws.cell(row = rows - 1, column= 1 + n).alignment)
            
            # Merge Cells
            ws.merge_cells(start_row=rows, start_column=col_index, end_row=rows, end_column=col_index + 2 )
            ws.merge_cells(start_row=rows, start_column=col_index + 6, end_row=rows, end_column=col_index + 7)
            ws.merge_cells(start_row=rows, start_column=col_index + 12, end_row=rows, end_column=col_index + 14)
            ws.merge_cells(start_row=rows, start_column=col_index + 15, end_row=rows, end_column=col_index + 17)
            ws.merge_cells(start_row=rows, start_column=col_index + 18, end_row=rows, end_column=col_index + 19)
            ws.merge_cells(start_row=rows, start_column=col_index + 22, end_row=rows, end_column=col_index + 25)
            ws.merge_cells(start_row=rows, start_column=col_index + 28, end_row=rows, end_column=col_index + 32)

            # Leave Start Date
            ws.cell(row=rows, column=col_index, value=getMMDDYYYY(approvalRecordLst[0]['details'][index]['startDate']))
            # Workday Name of start date
            ws.cell(row=rows, column=col_index+ 3, value= "(" + getWorkdayName(approvalRecordLst[0]['details'][index]['startDate']) + ")")
            # Full Day / AM / PM
            ws.cell(row=rows, column=col_index+ 6, value=(approvalRecordLst[0]['details'][index]['startTime']))
            # Leave End Date
            ws.cell(row=rows, column=col_index+ 12, value=getMMDDYYYY(approvalRecordLst[0]['details'][index]['endDate']))
            # Workday Name of end date
            ws.cell(row=rows, column=col_index+ 15, value= "(" + getWorkdayName(approvalRecordLst[0]['details'][index]['endDate']) + ")")
            # Full Day / AM / PM
            ws.cell(row=rows, column=col_index+ 18, value=(approvalRecordLst[0]['details'][index]['endTime']))
            # No of Working Days
            ws.cell(row=rows, column=col_index+ 22, value=(approvalRecordLst[0]['details'][index]['workday']))
            # No of Calendar Day
            ws.cell(row=rows, column=col_index+ 28, value=(approvalRecordLst[0]['details'][index]['calendarDay']))
    
    # Handle excel_range = "applicantName"
    row_applicant = record_count*2 +(coordinate_from_string(rpt['cell']['LeaveDetail'])[1]) + 1
    col_index = column_index_from_string(coordinate_from_string(rpt['cell']["applicantName"])[0])
    row_index = row_applicant
    ws.cell(row=row_index, column=col_index, value=(approvalRecordLst[0]['staff']))
    ws.cell(row=row_index+1, column=col_index, value=(approvalRecordLst[0]['position']))
    ws.cell(row=row_index, column=col_index+19, value=getMMDDYYYY(approvalRecordLst[0]['submit_date'])).alignment = Alignment(horizontal='center', vertical = 'center', wrap_text=True, wrapText=True)
    ws.merge_cells(start_row=row_index, start_column=col_index+19, end_row=row_index, end_column=col_index + 25)

    # Approver list below the Applicant Name in the form
    row_index = row_applicant + 2
    
    # Count number of record
    record_count = 1
    if len(approvalRecordLst[0]['approver2']) > 0: record_count = record_count + 1
    if len(approvalRecordLst[0]['approver3']) > 0: record_count = record_count + 1
    if record_count == 2: ws.insert_rows(row_index + 2, 2)
    if record_count == 3: ws.insert_rows(row_index + 2, 4)

    for n in range (0, ws.max_column):
        for i in range (1, record_count):
            ws.cell(row=row_index + i * 2, column= 1 + n).value = copy(ws.cell(row = (row_index + i * 2) - 2, column= 1 + n).value)
            ws.cell(row=(row_index + i * 2) + 1, column= 1 + n).value = copy(ws.cell(row = (row_index + i * 2) - 1, column= 1 + n).value)
            if n == 0 and i == 1:
                ws.cell(row=row_index + i * 2, column= 1 + n).value = str(ws.cell(row=row_index + i * 2, column= 1 + n).value).replace("1st","2nd")
            elif n == 0 and i == 2:
                ws.cell(row=row_index + i * 2, column= 1 + n).value = str(ws.cell(row=row_index + i * 2, column= 1 + n).value).replace("2nd","3rd")
            ws.cell(row=row_index + i * 2, column= 1 + n).fill = copy(ws.cell(row = (row_index + i * 2) - 2, column= 1 + n).fill)
            ws.cell(row=(row_index + i * 2) + 1, column= 1 + n).fill = copy(ws.cell(row = (row_index + i * 2) - 1, column= 1 + n).fill)                
            ws.cell(row=row_index + i * 2, column= 1 + n).font = copy(ws.cell(row = (row_index + i * 2) - 2, column= 1 + n).font)
            ws.cell(row=(row_index + i * 2) + 1, column= 1 + n).font = copy(ws.cell(row = (row_index + i * 2) - 1, column= 1 + n).font)     
            ws.cell(row=row_index + i * 2, column= 1 + n).number_format = copy(ws.cell(row = (row_index + i * 2) - 2, column= 1 + n).number_format)
            ws.cell(row=(row_index + i * 2) + 1, column= 1 + n).number_format = copy(ws.cell(row = (row_index + i * 2) - 1, column= 1 + n).number_format)  
            ws.cell(row=row_index + i * 2, column= 1 + n).border = copy(ws.cell(row = (row_index + i * 2) - 2, column= 1 + n).border)
            ws.cell(row=(row_index + i * 2) + 1, column= 1 + n).border = copy(ws.cell(row = (row_index + i * 2) - 1, column= 1 + n).border)
            ws.cell(row=row_index + i * 2, column= 1 + n).alignment = copy(ws.cell(row = (row_index + i * 2) - 2, column= 1 + n).alignment)
            ws.cell(row=(row_index + i * 2) + 1, column= 1 + n).alignment = copy(ws.cell(row = (row_index + i * 2) - 1, column= 1 + n).alignment)

    #approver list out
    for i in range (0, record_count):
        ws.cell(row=row_index + i * 2, column=col_index, value=(approvalRecordLst[0]['approver'+str(i+1)]))
        ws.cell(row=(row_index + i * 2) + 1, column=col_index, value=(approvalRecordLst[0]['approver_pos'+str(i+1)]))
        if len(str((approvalRecordLst[0]['approval_date'+str(i+1)]))) > 0:
            ws.cell(row=row_index + i * 2, column=col_index+19, value= "APPROVED \n" + getMMDDYYYY(approvalRecordLst[0]['approval_date'+str(i+1)])).alignment = Alignment(horizontal='center', vertical = 'center', wrap_text=True, wrapText=True)
        ws.merge_cells(start_row=row_index + i * 2, start_column=col_index + 19, end_row=row_index + i * 2, end_column=col_index + 25)   


#convert date from string (yyyy-mm-dd) to date format.
#parameter : must be in string (yyyy-mm-dd) format
#return :
#Date in datetime format.
def str2Date (psDateStr):
    return datetime.strptime(psDateStr, "%Y-%m-%d")

#covert date to string 
#parameter : must be in datetime format
#return :
#date in string format : "YYYY-MM-DD"
def date2Str(psDate):
    return datetime.strftime(psDate, "%Y-%m-%d")


# get staff record 
# parameter:
# psRacf - RACF of the user
# return:
# staff record in MongoDB.
def getStaffRecord (psRacf):
    if len(psRacf) > 0 :
        staffRecord = eleaveDtl.find_one ( {"staff.racf" : { '$regex' : psRacf, '$options' : "i"} , "staff.status": { '$regex': "ACTIVE", '$options': "i"} } )

        return(staffRecord)
    else:
        return None




def getLeaveTypes():
    global leaveTypeLst
    leaveTypeLst = list(leaveTypes.find({}))

def getLeaveGroups():
    global leaveGroupLst
    leaveGroupLst = list(leaveGroups.find({}))

def getAllOffice():
    # allOffice = ['HKG', 'REG', 'TPE', 'DEL', 'FLR', 'CHN']
    allOffice = eleaveDtl.distinct('staff.hr_office')

    # NY for approval only, drop NY in the list
    if 'NYO' in allOffice:
        allOffice.remove('NYO')

    # CHN drop as well
    if 'CHN' in allOffice:
        allOffice.remove('CHN')

    return allOffice

def getAllSpecialLeave():
    result = [ ]
    allSpeicalLeave = list(leaveTypes.find({"other_leave": True}))
    for rec in allSpeicalLeave:
        result.append({'leave_type_id': rec['leave_type_id'], 'leave_type': rec['leave_type'].title()})

    return result

def specialLeaveRefNo(ofc, year):
    maxRefNo = 1
    code = ""

    try:
        allSpecialLeave = list(otherLeaves.find({"office": ofc, "year": year}))
        maxRefNo += int (len(allSpecialLeave))
    except:
        pass

    if ofc == "HKG":
        code = "HK"

    elif ofc == "TPE":
        code = "TW"

    elif ofc == "DEL":
        code = "IN"

    elif ofc == "REG":
        code = "RG"

    # elif ofc == "CHN": # terminated
    #     code = "CN"

    elif ofc == "FLR":
        code = "IT"

    zerodigit = "0" * (3 - int(len(str(maxRefNo))))
    return str(f"{code}-{year}-{zerodigit}{maxRefNo}")


def getAllSpecialRef():

    # 1. Fetch the special leave types and create a lookup dictionary

    special_leaves = getAllSpecialLeave()
    leave_type_map = {item['leave_type_id']: item['leave_type'] for item in special_leaves}
    
    records = list(otherLeaves.find({}))
    
    for rec in records:
        if '_id' in rec:
            rec['_id'] = str(rec['_id']) 
        
        # 2. Get the leave_type_id from the current record
        # (Using .get() prevents KeyError if the field is missing)
        type_id = rec.get('leave_type')
        
        # 3. Look up the description and add it to the record
        # If it's not found in the map, it defaults to None (or you can set a default string)
        rec['leave_type_name'] = leave_type_map.get(type_id, "Unknown Leave Type")

    
    return records



# get long ref no. for displaying 
# parameter:
# psOffice - hr_office of the staff
# psRefNo - leave ref_no in database
# psRacf - Racf of staff
# return
# ref_no for display, format <office><ref_no><last 3 characters of RACF>
def getDisplayRefNo(psRefNo, psOffice, psRacf):
    return(psOffice + str(psRefNo) + psRacf[-3:])

def  getActualRefNo(psRefNo):
    return(int(psRefNo[3:10]))
    
# get date from string format to mm/dd/yyyy format
# parameter:
# psDateString - Date in String format , i.e. YYYY-MM-DD
# return:
# Date in string format as mm/dd/yyyy
def getMMDDYYYY(psDateString):
    return (datetime.strftime(str2Date(psDateString), "%m/%d/%Y"))

def getWorkdayName(psDateString):
    workdayName = (datetime.strftime(str2Date(psDateString), "%a"))
    return workdayName

# get display leave year
# parameter:
# psYear - leave year in int.
# return:
# leave year period in string, format : "Mar 1, year - Feb 28 (or 29), year"
def getDisplayLeaveYear(psYear):

    if calendar.isleap(psYear):
        return (df['gcYearStartDate'][0] + str(psYear) + " - " + df['gcYearEndDateLeap'][0] + str(psYear + 1))
    else:
        return(df['gcYearStartDate'][0] + str(psYear) + " - " + df['gcYearEndDate'][0] + str(psYear + 1))      

def getLeave(psYear, psLeaveType, psLeaveStatus, psRecord):
    return (list(filter(lambda r: (r["type"].upper() == psLeaveType.upper() and r["applicationStatus"].upper() == psLeaveStatus.upper() and r["year"] == psYear), psRecord["leave_record"])))


def countLeave (psYear, psLeaveType, psLeaveStatus, psRecord):
    leaveDays = 0
    for record in getLeave(psYear, psLeaveType, psLeaveStatus, psRecord):
        for leaveDetails in record["details"]:
            leaveDays += leaveDetails["no_of_workday"]

    return leaveDays

def getYearEntitlement(Year, StaffRecord, LeaveType):
    for rec in StaffRecord['entitlement']:
        if int(rec['year']) == int(Year):
            if LeaveType == 'LVE01':
                return float(rec['annual_entitlement'])
            if LeaveType == 'LVE02':
                return float(rec['casual_entitlement'])

def getYearCarryForward(Year, StaffRecord, Type):

    for rec in StaffRecord['entitlement']:
        if int(rec['year']) == int(Year) and Type == "LVE01":
            return float(rec['carry_forward'])
        
    return 0

# get leave year period
# parameter:
# psYear - eleave Year
# return:
# list with leave year start date and leave year end date

# Status_code 200: passed
# Status_code 801: Fail to get leave year period


def getLeaveYrPeriod(psYear):

    try:
        yrIndex = json.loads(os.getenv('YEARS')).get('year').index(psYear)
        eleavePeriod = json.loads(os.getenv('YEARS')).get('period')[yrIndex]
        leaveYrStart= datetime.strptime(eleavePeriod.split("-")[0].strip(), "%b %d, %Y")
        leaveYrEnd = datetime.strptime(eleavePeriod.split("-")[1].strip(), "%b %d, %Y")
        leaveYrPeriod = {
            'leaveYrStart': leaveYrStart,
            'leaveYrEnd': leaveYrEnd
        }
        leaveYrPeriodLst = []
        leaveYrPeriodLst.append(leaveYrPeriod)
        return ({"pass": True, "error_message": "", "result":leaveYrPeriodLst, "Status_code": 200})
    except Exception as e :
        print (e)
        return ({"pass": False, "error_message": "Fail to get leave year period", "result":[], "Status_code": 801})

def chkPeriod(psStartDate, psEndDate, psYear):
    result = getLeaveYrPeriod(psYear)
    if result.get('pass'):
        sDate =  str2Date(psStartDate) 
        eDate = str2Date(psEndDate)
        yrStart = result.get('result')[0].get('leaveYrStart')
        yrEnd = result.get('result')[0].get('leaveYrEnd')
        if sDate >= yrStart and sDate <= yrEnd and eDate >= yrStart and eDate <= yrEnd:
            return ({'pass': True, 'error_message': "", 'result': None})
        else:
            return({'pass': False, 'error_message': "Leave applying is not within the leave year", 'result': None, 'Status_code': 507})
        
    else:
        return ({'pass': False, 'error_message': result.get('error_message'), 'result': None, 'Status_code': 508})

# get leave history of a staff for a particular year
# paramters:
# psRacf - Racf of the user
# psYearStart - Starting Leave Year required
# psYearEnd - Ending Leave Year required
# return:
# list of leave history in the format :
# [{"ref_no": int, "year": int, "type": string, "startDate": string, "startTime": string, "endDate": string, "endTime": string, "status": string, "workDay": int, calendarDay: int, "ldate": datetime, "ltime": string]
def getLeaveHistory(psYearStart, psYearEnd, psRecord):

    try:
        period = getLeaveYrPeriod(psYearStart)
        periodStart = period.get('result')[0].get('leaveYrStart')
        periodEnd = period.get('result')[0].get('leaveYrEnd')
    except:
        pass
    
    leaveHistoryAllLst = [ ]
    yr = psYearStart
    while yr <= psYearEnd:
        for r in psRecord["leave_record"]:
            #if r["year"] == yr:
            for d in r["details"]:
                for p in d["period"]:
                    if  periodStart <= datetime.strptime(p["ldate"], "%Y-%m-%d") <= periodEnd:
                        currRecord = {
                            "ref_no": r["ref_no"] if r["otherRefNo"] == "" else r["otherRefNo"],
                            "other_leave": False if r["otherRefNo"] == "" else True,
                            "office": psRecord["staff"]["hr_office"],
                            "racf": psRecord["staff"]["racf"],
                            "staffname": psRecord["staff"]["name"],
                            "empID": psRecord["staff"]["empID"],
                            "dept": psRecord["staff"]["dept"],
                            "position": psRecord["staff"]["position"],
                            "year" : r["year"],
                            "type": r["type"],
                            "sharePointId": r["sharePointId"],
                            "startDate": d["start_date"],
                            "startTime": d["start_time"],
                            "endDate": d["end_date"],
                            "endTime": d["end_time"],
                            "applicationStatus": r["applicationStatus"],
                            "approvalStatus": r["approvalStatus"],
                            "workDay": d["no_of_workday"],
                            "calendarDay": d["no_of_calendarday"],
                            "submitDate": r["submit_date"],
                            "ldate": str2Date(p["ldate"]),
                            "ltime": p["ltime"],
                            "approver1": psRecord["staff"]["approver1"],
                            "approver2": psRecord["staff"]["approver2"],
                            "approver3": psRecord["staff"]["approver3"],
                            "lastUpdate": r["lastUpdate"],
                            "updateDate": r["updateDate"]                            
                        }
                        
                        leaveHistoryAllLst.append (currRecord)
        yr += 1
    return (leaveHistoryAllLst)



# get leave entitlement of a staff for a particular leave type in a particular year
# parameters:
# psRecord : staff record
# psYear : leave year
# psLeaveType : leave type 
# return:
# return leave entitlement, format ["{leaveEntitle": int, "carryForward": int, "forfeoitDate": datetime} ]
def getLeaveEntitlement(psYear, psLeaveTypeAttr, psRecord):

    entitlementLst = [ ]
    entitlement = {
        "leaveEntitle": 0,
        "carryForward": 0,
        "forfeitDate": str2Date("2000-01-01")
    }
    for e in psRecord["entitlement"]:
        if e["year"] == psYear:
            if psLeaveTypeAttr.get("entitlement_field", "") != "":
                leaveEntitle = e.get(psLeaveTypeAttr.get("entitlement_field"), 0)
            else:
                leaveEntitle = 0
            if psLeaveTypeAttr.get("carry_forward_field", "") != "":
                carryForward = e.get(psLeaveTypeAttr.get("carry_forward_field"), 0)
            else:
                carryForward = 0
            if psLeaveTypeAttr.get("forfeit_date_field", "") != "":
                forfeitDate = str2Date(e.get(psLeaveTypeAttr.get("forfeit_date_field"), "2000-01-01"))
            else:
                forfeitDate = str2Date("2000-01-01")
            entitlement = {
                    "leaveEntitle": leaveEntitle,
                    "carryForward": carryForward,
                    "forfeitDate": forfeitDate
            }
    entitlementLst.append(entitlement)  

    return (entitlementLst)


# get no. of working day in the time slot
# parameter:
# psPeriod : time slot list containing leave period in working days
# return:
# no. of work days 
def getWorkDay(psPeriod):
    return (len(psPeriod) / 2)

#check leave balance
#parameters:
#psLeaveEntitle : Leave entitlement for the leave type.
#psCarryForward : Annual leave carry foward from last year
#psForfeitDate:  Forfeit Date for carry forward
#psLeaveHistoryLst : List of time slot with leave history in the format [{"ref_no" int, "year" int, "type" string, "status" string, "ldate": datetime, "ltime": "AM" / "PM"}]
#psLeaveSlotLst : Leave applying
#psLeaveType : Leave type applying
#psYear : Leave year
#return:
#leave balanace after taking counting the leave applying.
def checkBalance(psYear, psLeaveTypeAttr, psRecord, psApplyingLeaveSlotLst):
    leaveEntitleLst = getLeaveEntitlement(psYear, psLeaveTypeAttr, psRecord)
    leaveHistoryLst = getLeaveHistory(psYear, psYear, psRecord)
    leaveHistoryLst = list(filter(lambda r: (r["applicationStatus"].upper() != df['gcStatusCancel'][0] and r["applicationStatus"].upper() != df['gcStatusReject'][0]), leaveHistoryLst))
    taken = 0
    
    leaveHistoryTypeLst = list(filter(lambda r: (r["type"].upper() == psLeaveTypeAttr.get("leave_type_id").upper() and r["year"] == psYear), leaveHistoryLst))
    # count no. of leave taken before and after the Carry Forward Forfeit Date
    for lve in leaveHistoryTypeLst:
            taken += 0.5
    #    else:
    #        afterForfeit += 0.5
    
    # count no. of leave already taken + applying before and after the Carry Forward Forfeit Date 
    for apply in psApplyingLeaveSlotLst:
            taken += 0.5
    #    else:
    #        afterForfeit += 0.5

    # if total leave taken + applying before the Carry Forward Forfeit Date > total leave entitlement + leave carry forward, return leave balance
    #if leaveEntitleLst[0]["leaveEntitle"] + leaveEntitleLst[0]["carryForward"] < beforeForfeit:
    #    return (leaveEntitleLst[0]["leaveEntitle"] + leaveEntitleLst[0]["carryForward"] - beforeForfeit)

    # if leave remaining after the Forefeit Date >= leave entitlement, set leave entitlement after the Forfeit Date = leave entitlement
    #if leaveEntitleLst[0]["leaveEntitle"] + leaveEntitleLst[0]["carryForward"] - beforeForfeit >= leaveEntitleLst[0]["leaveEntitle"]:
    #    entitleAfterForfeit = leaveEntitleLst[0]["leaveEntitle"]

    # if leave remaining after the Forefeit Date < leave entitlement, set leave entitlement after the Forfeit Date = leave remaining of that year
    #else:
    #    entitleAfterForfeit = leaveEntitleLst[0]["leaveEntitle"] + leaveEntitleLst[0]["carryForward"] - beforeForfeit
    
    # if leave entitlement after the Forfeit Date < leave taken + applying after the Foefeit Date, returm leave balance
    #if entitleAfterForfeit < afterForfeit:
    #    return (entitleAfterForfeit - afterForfeit)

    # if leave taken + applying not exceeding the leave entitlement, return leave balance
    #return (entitleAfterForfeit - afterForfeit)
    return (leaveEntitleLst[0]["leaveEntitle"] + leaveEntitleLst[0]["carryForward"] - taken)

#list all dates for a Week of Day in a specific year.
#parameters :
#psYear : year 
#psDay : Day of week needed, 1 - Monday, 2 - Tuesday, 3 - Wednesday, .... 7- Sunday
#return:
#dates of the day of week required of that year in datetime format.
def alldays(psYear, psDay):
    #d = date(psYear, 1, 1)
    d = str2Date(str(psYear) + "-01-01")
    d += timedelta(days = (psDay - d.isoweekday()) % 7)
    while d.year == psYear:
        yield d
        d += timedelta(days = 7)    

#Get all weekend from year submitted, previous year and next year
#parameter : 
#psYear : year required
#return:
#all sat. and sunday of a required year in the format : [{"ldate": datetime, "ltime": "AM" / "PM"}]
def getAllWeekend(psYear):
    weekendLst = [ ] 
    yr = psYear
    while yr <= psYear:
        dow = 6
        while dow <= 7:
            for d in alldays(yr, dow):
                weekend = {
                    "ldate": d,
                    "ltime": "AM",
                    "type" : "weekend"

                }
                weekendLst.append(weekend)
                weekend = {
                    "ldate": d,
                    "ltime": "PM",
                    "type" : "weekend"
                }
                weekendLst.append(weekend)

            dow += 1
        yr += 1
    return weekendLst

#get all holidays after the year input.  Exclude weekend.
#parameters :
#psYear : Year
#psOffice : Office for the holidays required. 
#return : 
#holidays list of the required year of that office in the format [{"ldate": datetime, "ltime": "AM" / "PM"}]
def getHolidays(psYear, psOffice):
    # convert Date in holiday from string to Date format and exclude weekend.
    holidayLst = list(holidays.find ({ "$and" : [
                                    { "Year":  { "$eq" : psYear } },
                                    { "Office": { "$eq" : psOffice} }
                                ] }
    ) )
    holidaySlotLst = [ ]
    for h in holidayLst:
        if (str2Date(h["Date"])).isoweekday() != 6 and (str2Date(h["Date"])).isoweekday() !=7:
            slot = {
                "ldate": str2Date(h["Date"]),
                "ltime": h["Time"],
                "type": "holiday"
            }
            holidaySlotLst.append (slot)
    return holidaySlotLst

# get new leave application ref no.
# parameters :
# psLeaveHistoryLst : List of time slot with leave history in the format [{"ref_no" int, "year" int, "type" string, "status" string, "ldate": datetime, "ltime": "AM" / "PM"}]
# psLeaveYear: year
# return:
# if this is the first leave application in the year, return as "leave year" + "001"
# else, return the max ref_no + 1
#def getNewRefNo(psYear, psRecord):
#    currYearRecordLst = list(filter(lambda r: (r["year"] == psYear), psRecord["leave_record"]))
    
#    if len(currYearRecordLst) == 0:
#        return (int(str(psYear) + "001"))
#    else:
#        maxRefNo = max(r["ref_no"] for r in currYearRecordLst)
#        return (maxRefNo + 1)

def getNewRefNo(year, racf):
    maxRefNo = 1
    try: 

        record = getStaffRecord(racf)
        leave_records = record.get('leave_record', False)
        if leave_records:
            for record in leave_records:
                if year == record.get('year'):
                    maxRefNo += 1
        
        zerodigit = "0" * (3 - int(len(str(maxRefNo))))

    except:
        zerodigit = "0" * (3 - int(len(str(maxRefNo))))

    return int(f"{year}{zerodigit}{maxRefNo}")

# update / save leave record to database
# parameters:
# psRecord : all records in the collection eleave_dtl in MongoDB
# psID : Object ID of the applicant
# psField : Field that we record to make changes
# psValue : New Value
# return:
# True if update successful
# False if update fail
def updateDB(psRecord, psID, psField, psValue):
    query = { "_id" : psID}
    entry = psRecord.find_one(query)
    entry[psField] = psValue
    try:
        psRecord.update_one(query, { '$set': {psField : entry[psField]}})
        return True
    except:
        return False

def updateDB2(psID, psUpdateLst):
    session = client.start_session(causal_consistency=True)
    session.start_transaction ()
    try: 
        for item in psUpdateLst:
            field = item.get("field")
            value = item.get("value")
            e = eleaveDtl.update_one (
                {"_id": psID },
                { "$set" : { field : value } },
                session=session
            )            
    except Exception as e:
        session.abort_transaction()
        return ({"pass": False, "error_message" : str(e), "result": [ ], "Status_code": 200})
    else:
        session.commit_transaction()
        return ({"pass": True, "error_message": "", "result" : [], "Status_code": 200})
    finally:
        session.end_session()

def applicationStatusForEmail(leaveContent, finalapprover, request, action):

    Approval_Status_str = ["", "", ""]

    #print (leaveContent[0]["applicationStatus"])
    #print (request)
    #print (action)

    # Situation 1 : approver 1 approved with date, and the status is not rejected = approver 1 approved
    if (len(leaveContent[0]["approval"].get("approval_date1")) > 0) and (action != "REJECT"):
        Approval_Status_str[0] = getStaffRecord(leaveContent[0]["approval"].get("approver1"))['staff']["name"] + " : " + "Approved" + "\n"
        if finalapprover == 2:
            Approval_Status_str[1] = getStaffRecord(leaveContent[0]["approval"].get("approver2"))['staff']["name"] + " : " + "Pending" + "\n"
        if finalapprover == 3:
            Approval_Status_str[1] = getStaffRecord(leaveContent[0]["approval"].get("approver2"))['staff']["name"] + " : " + "Pending" + "\n"
            Approval_Status_str[2] = getStaffRecord(leaveContent[0]["approval"].get("approver3"))['staff']["name"] + " : " + "Pending" + "\n"
    # Situation 2 : approver 2 approved with date, and the status is not rejected = approver 2 approved
    if (len(leaveContent[0]["approval"].get("approval_date2")) > 0) and (action != "REJECT"):
        Approval_Status_str[0] = getStaffRecord(leaveContent[0]["approval"].get("approver1"))['staff']["name"] + " : " + "Approved" + "\n"
        if finalapprover == 2:
            Approval_Status_str[1] = getStaffRecord(leaveContent[0]["approval"].get("approver2"))['staff']["name"] + " : " + "Approved" + "\n"
        if finalapprover == 3:
            Approval_Status_str[1] = getStaffRecord(leaveContent[0]["approval"].get("approver2"))['staff']["name"] + " : " + "Approved" + "\n"
            Approval_Status_str[2] = getStaffRecord(leaveContent[0]["approval"].get("approver3"))['staff']["name"] + " : " + "Pending" + "\n"
    # Situation 3 : approver 3 approved with date, and the status is not rejected = approver 3 approved
    if (len(leaveContent[0]["approval"].get("approval_date3")) > 0) and (action != "REJECT"):
        Approval_Status_str[0] = getStaffRecord(leaveContent[0]["approval"].get("approver1"))['staff']["name"] + " : " + "Approved" + "\n"
        if finalapprover == 2:
            Approval_Status_str[1] = getStaffRecord(leaveContent[0]["approval"].get("approver2"))['staff']["name"] + " : " + "Approved" + "\n"
        if finalapprover == 3: 
            Approval_Status_str[1] = getStaffRecord(leaveContent[0]["approval"].get("approver2"))['staff']["name"] + " : " + "Approved" + "\n"
            Approval_Status_str[2] = getStaffRecord(leaveContent[0]["approval"].get("approver3"))['staff']["name"] + " : " + "Approved" + "\n"
    # Situation 4 : approver 1 approved with date, and the status is rejected = approver 1 rejected
    if (len(leaveContent[0]["approval"].get("approval_date1")) > 0) and (action == "REJECT"):
        Approval_Status_str[0] = getStaffRecord(leaveContent[0]["approval"].get("approver1"))['staff']["name"] + " : " + "Rejected" + "\n"
        if finalapprover == 2:
            Approval_Status_str[1] = getStaffRecord(leaveContent[0]["approval"].get("approver2"))['staff']["name"] + " : " + "-" + "\n"
        if finalapprover == 3:
            Approval_Status_str[1] = getStaffRecord(leaveContent[0]["approval"].get("approver2"))['staff']["name"] + " : " + "-" + "\n"
            Approval_Status_str[2] = getStaffRecord(leaveContent[0]["approval"].get("approver3"))['staff']["name"] + " : " + "-" + "\n"  
    # Situation 5 : approver 2 approved with date, and the status is rejected = approver 2 rejected 
    if (len(leaveContent[0]["approval"].get("approval_date2")) > 0) and (action == "REJECT"):
        Approval_Status_str[0] = getStaffRecord(leaveContent[0]["approval"].get("approver1"))['staff']["name"] + " : " + "Approved" + "\n"
        if finalapprover == 2:
            Approval_Status_str[1] = getStaffRecord(leaveContent[0]["approval"].get("approver2"))['staff']["name"] + " : " + "Rejected" + "\n"
        if finalapprover == 3:
            Approval_Status_str[1] = getStaffRecord(leaveContent[0]["approval"].get("approver2"))['staff']["name"] + " : " + "Rejected" + "\n"
            Approval_Status_str[2] = getStaffRecord(leaveContent[0]["approval"].get("approver3"))['staff']["name"] + " : " + "-" + "\n"     
    # Situation 6 : approver 3 approved with date, and the status is rejected = approver 3 rejected 
    if (len(leaveContent[0]["approval"].get("approval_date3")) > 0) and (action == "REJECT"):
        Approval_Status_str[0] = getStaffRecord(leaveContent[0]["approval"].get("approver1"))['staff']["name"] + " : " + "Approved" + "\n"
        if finalapprover == 2:
            Approval_Status_str[1] = getStaffRecord(leaveContent[0]["approval"].get("approver2"))['staff']["name"] + " : " + "Approved" + "\n"
        if finalapprover == 3:
            Approval_Status_str[1] = getStaffRecord(leaveContent[0]["approval"].get("approver2"))['staff']["name"] + " : " + "Approved" + "\n"
            Approval_Status_str[2] = getStaffRecord(leaveContent[0]["approval"].get("approver3"))['staff']["name"] + " : " + "Rejected" + "\n"     

    Approval_Status = Approval_Status_str[0] + Approval_Status_str[1] + Approval_Status_str[2]

    return Approval_Status


def getSummaryForm(year, racf):                  
    psInput =  {'year': year, 'racf': racf}    

    result = listLeave(psInput)
 
    rpt = reportMap.find_one ( { "report": "Leave Summary"} )
    
    if (result.get("pass")): 
        hdr = result.get("result")[0]["header"]
        alTaken = 0
        alPending = 0
        alBalance = 0
        clTaken = 0
        clPending = 0
        clBalance = 0
        slTaken = 0
        slPending = 0
        slBalance = 0
        for lve in leaveTypeLst:
            hdrData =  list(filter(lambda r: (r["leaveTypeId"].upper() == lve["leave_type_id"]), hdr))[0]
            if lve["leave_type_id"] == "LVE01":
                alTaken = hdrData["taken"]
                alPending = hdrData["pending"]
                alBalance = hdrData["balance"]
            elif lve["leave_type_id"] == "LVE02":
                clTaken = hdrData["taken"]
                clPending = hdrData["pending"]
                clBalance = hdrData["balance"]
            elif lve["leave_type_id"] == "LVE04":
                slTaken = hdrData["taken"] + slTaken
                slPending = hdrData["pending"] + slPending
                slBalance = hdrData["balance"] + slBalance
            elif lve["leave_type_id"] == "LVE05":
                slTaken = hdrData["taken"] + slTaken
                slPending = hdrData["pending"] + slPending
                slBalance = hdrData["balance"] + slBalance
        rptDtlLst = [ ]
        for record in result.get("result"):
            for dtl in record["details"]:
                rptDtl = {
                    "submitDate": dtl.get("submitDate"),
                    "refNo": dtl.get("refNo"),
                    "office": dtl.get("office"),
                    "staffname": dtl.get("staffname"),
                    "dept": dtl.get("dept"),
                    "type": dtl.get("type"),
                    "year": dtl.get("year"),
                    "leaveFrom": dtl.get("leaveFrom"),
                    "startPeriod": dtl.get("startPeriod"),
                    "leaveTo": dtl.get("leaveTo"),
                    "endPeriod": dtl.get("endPeriod"),
                    "workday": dtl.get("workday"),
                    "calendarDay": dtl.get("calendarDay"),
                    "applicationStatus": dtl.get("applicationStatus"),
                    "lastUpdate": dtl.get("lastUpdate"),
                    "updateDate": dtl.get("updateDate")        
                }
                rptDtlLst.append(rptDtl)

            # Sort by leave start requested by Sandy 1/26/2023
            rptDtlLst.sort(key=lambda x: x.get('leaveFrom'))
            
            # added by Vincent Cheng on 11/23 
            try:
                if record["details"][0].get("year"):
                    pass
            except:            
                return jsonify({"error_message" : "Sorry, we failed to generate Leave Summary.  Perhaps no data for the year"}), 501    
                
    
            report = {
                "hdrCalendarYear": record["details"][0].get("year"),
                "hdrUser": record["details"][0].get("staffname") + "\nLeave Application Summary",
                "hdrALTaken": alTaken,
                "hdrALPending": alPending,
                "hdrALBalance": alBalance,
                "hdrSLTaken": slTaken,
                "hdrSLPending": slPending,
                "hdrCLTaken": clTaken,
                "hdrCLPending": clPending,
                "hdrCLBalance": clBalance,
                "dtl": rptDtlLst
                
            }
        #filename when using in Heroku:
        fs = gridfs.GridFS(db)
        wb = load_workbook(filename=BytesIO(fs.get(ObjectId(rpt["file"]["fileObj"])).read()))

        # filename in development:
        #wb = load_workbook(filename=rpt["file"]["fileName"])
        ws = wb[rpt["file"]["wsName"]]
        
        result = genReport(ws, report, rpt)
        if result[1] == 0:
            out = BytesIO()
            wb.save(out)
            out.seek(0)        
 
            wb.close()            
            return out
        else:            
            return jsonify({"error_message" : "Sorry, we failed to generate Leave Summary.  Perhaps no data for the year"}), 501    


def getApplicationForm(ref, racf):

    # Get Staff record for output
    StaffRecord = getStaffRecord(racf)

    # Find leave list by racf and ref
    ref_no = ref.replace(StaffRecord['staff']['hr_office'],"") # remove hr office in reference no
    ref_no = ref_no.replace(StaffRecord['staff']['racf'][-3:],"") # remove staff racf in reference no

    # Get leave balance 
    getLeaveTypes()

    approvalRecordLst = [ ]

    for rec in StaffRecord['leave_record']:

        #Select exact application by reference number
        if rec['ref_no'] == int(ref_no):
            leaveDetailsLst = [ ]
            for details in rec["details"]:
                # get rows for each leave application
                leaveDetails = {
                    "startDate": details["start_date"],
                    "startTime": details["start_time"],
                    "endDate": details["end_date"],
                    "endTime": details["end_time"],
                    "workday": details["no_of_workday"],
                    "calendarDay": details["no_of_calendarday"]
                    }
                leaveDetailsLst.append(leaveDetails)

                # Check the balance from Thomas function
                displayLeaveHistoryHdr = [ ]
                for lveType in leaveTypeLst:
                    leaveTypeHdr = {
                                    "leaveType": lveType.get("leave_type"),
                                    "leaveTypeId": lveType.get("leave_type_id"),
                                    "taken" : countLeave(rec['year'], lveType.get("leave_type_id"), df['gcStatusApproved'][0], StaffRecord),
                                    "pending": countLeave(rec['year'], lveType.get("leave_type_id"), df['gcStatusPending'][0], StaffRecord),
                                    "balance": checkBalance(rec['year'], lveType, StaffRecord, [])
                                    }
                    displayLeaveHistoryHdr.append(leaveTypeHdr)
                
            # Summarize the number of balance
            for leaveType in displayLeaveHistoryHdr:
                if leaveType['leaveTypeId'] == rec['type']:
                    DaysOfApproved = leaveType['taken']
                    DaysOfPending = leaveType['pending']
                    DaysOfleft = leaveType['balance']
                    if rec['type'] == 'LVE01':
                        DaysOfCarryForward = getYearCarryForward(rec['year'], StaffRecord, rec['type'])
                        DaysOfEntitlement = str(DaysOfCarryForward) + " (" + str(int(rec['year']-1)) + ") " + "+ " + str(getYearEntitlement(rec['year'], StaffRecord, rec['type'])) + " (" + str(int(rec['year'])) + ") "
                    elif rec['type'] == 'LVE02':
                        DaysOfCarryForward = 0
                        DaysOfEntitlement = str(getYearEntitlement(rec['year'], StaffRecord, rec['type'])) + " (" + str(int(rec['year'])) + ") "
                    elif rec['type'] == 'LVE03':
                        DaysOfCarryForward = 0
                        DaysOfEntitlement = "N/A"
                        DaysOfleft = "N/A"
                    elif rec['type'] == 'LVE04':
                        DaysOfCarryForward = 0
                        DaysOfEntitlement = "N/A"
                        DaysOfleft = "N/A"
                    elif rec['type'] == 'LVE05':
                        DaysOfCarryForward = 0
                        DaysOfEntitlement = "N/A"
                        DaysOfleft = "N/A"
                    else:
                        DaysOfCarryForward = 0
                        DaysOfEntitlement = "N/A"
                        DaysOfleft = "N/A"
                    

            get_approver1 = ""
            get_pos_approver1 = ""
            get_approver2 = ""
            get_pos_approver2 = ""
            get_approver3 = ""
            get_pos_approver3 = ""

            if len(str(rec['approval']['approver1'])) > 0:
                get_approver1 = getStaffRecord(rec['approval']['approver1'])['staff']['name']
                get_pos_approver1 = getStaffRecord(rec['approval']['approver1'])['staff']['position']

            if len(str(rec['approval']['approver2'])) > 0:
                get_approver2 = getStaffRecord(rec['approval']['approver2'])['staff']['name']
                get_pos_approver2 = getStaffRecord(rec['approval']['approver2'])['staff']['position']

            if len(str(rec['approval']['approver3'])) > 0:
                get_approver3 = getStaffRecord(rec['approval']['approver3'])['staff']['name']
                get_pos_approver3 = getStaffRecord(rec['approval']['approver3'])['staff']['position']

            if rec['sharePointId'] == "":
                DissharePointid = ""
            else:
                DissharePointid = "(" + str(rec['sharePointId']) + ")"

            try:
                TakenApproved = float(DaysOfApproved + DaysOfPending)
                NoDaysTakenApproved = str(float(TakenApproved)) + " (" + str(float(DaysOfApproved)) + " + "+ str(float(DaysOfPending)) + ") "
            except:
                TakenApproved = "NA"
                NoDaysTakenApproved = "N/A"

            # Go back to build the structure for excel output file
            # Array item label must be the same as MongoDB cell field in fileDrectory

            leaveRecord = {
                "staff": StaffRecord['staff']['name'],
                "racf": racf,
                "position": StaffRecord['staff']['position'],
                "dept": StaffRecord['staff']['dept'],
                "date_joined": StaffRecord['staff']['date_join'],
                "ref_no": ref if rec['otherRefNo'] == "" else rec['otherRefNo'],
                "sharePointid": DissharePointid,
                "approver1": get_approver1,
                "approver_pos1": get_pos_approver1,
                "approval_date1": rec['approval']['approval_date1'],
                "approver2": get_approver2,
                "approver_pos2": get_pos_approver2,
                "approval_date2": rec['approval']['approval_date2'],
                "approver3": get_approver3,
                "approver_pos3": get_pos_approver3,
                "approval_date3": rec['approval']['approval_date3'],
                "NoDaysEntitlement": DaysOfEntitlement ,
                "NoDaysTakenApproved": NoDaysTakenApproved,
                "NoDaysLeft": DaysOfleft,
                "type_id": rec["type"],
                "leaveTypeBalance": list(filter(lambda r: (r["leave_type_id"].upper() == rec["type"]), leaveTypeLst))[0].get("leave_type") + " BALANCE",
                "type": list(filter(lambda r: (r["leave_type_id"].upper() == rec["type"]), leaveTypeLst))[0].get("leave_type"),
                "calendarYear": getDisplayLeaveYear(rec["year"]),
                "submit_date": rec['submit_date'],
                "details": leaveDetailsLst
                }
            
            #Output to array to excel file
            approvalRecordLst.append(leaveRecord)

    # Get mapping from MongoDB
    rpt = reportMap.find_one ( { "report": "Application Form"} )

    #filename when using in Heroku:
    fs = gridfs.GridFS(db)
    wb = load_workbook(filename=BytesIO(fs.get(ObjectId(rpt["file"]["fileObj"])).read()))
    ws = wb[rpt["file"]["wsName"]]

    try:
        genApplyForm(ws, approvalRecordLst, rpt)
    except Exception as e:
        print (e)
        return jsonify({"error_message" : "Sorry, we failed to generate Application form"}), 501    

    # Output 
    out = BytesIO()
    wb.save(out)
    out.seek(0)

    wb.close()            
    print('sending file...')

    return out



def postmarker(message, title, sendTo, sendCC, attachment=None, attachmentname=None):

    # Convert bytesIO attachment list to encode varaible for restful API
    attachedfiles = [ ]

    if attachment is not None:
        for index, bytesio in enumerate(attachment):
            bytesio.seek(0)
            content = bytesio.read()

            encoded = base64.b64encode(content).decode('utf-8')

            attachedfiles.append({'Name': attachmentname[index],
                                'Content': encoded,
                                'ContentType' : 'application/octet-stream'
            })

    # convert html body to email
    body_plain = message
    # html body
    line_break = '\n' #used to replace line breaks with html breaks
    body_html = f'''
                <html>
                <head></head>
                <body>
                {'<br/>'.join(body_plain.split(line_break))}
                </body>
                </html>
                '''

    headers = { 'Content-Type': 'application/json', 
                'X-Postmark-Server-Token': os.environ.get('POSTMARKER_API_KEY'), 
                'Accept':'application/json'}

    parameters = {
                  'MessageStream': 'e-leave',
                  'From': os.environ.get('EMAIL_SENDER'), 
                  'To': sendTo, 
                  'Cc': sendCC,
                  'Bcc': os.environ.get('Send_BCc'),
                  'Subject': title, 
                  'HtmlBody': body_html,
                  'Attachments':  attachedfiles
                  }
    

    data = json.dumps(parameters)

    r = requests.post('https://api.postmarkapp.com/email', headers=headers, data=data)

    response = json.loads(r.text)
    if response['ErrorCode'] == 0:
        print('Message ID = %s' % response['MessageID'])
    else:
        print('Message not sent')

def localSend(message, title, sendTo, sendCC, attachment=None, attachmentname=None):

    sender = os.environ.get('EMAIL_SENDER')

    # subject, text body
    subject = title

    body_plain = message.decode('utf-8') if isinstance(message, bytes) else message

    # html body
    line_break = '\n' #used to replace line breaks with html breaks
    body_html = f'''
                <html>
                <head></head>
                <body>
                {'<br/>'.join(body_plain.split(line_break))}
                </body>
                </html>
                '''
    
    # create message container
    message = MIMEMultipart('alternative')
    message['Subject'] = subject
    message['From'] = sender
    message['To'] = sendTo
    message['Cc'] = sendCC

    # prepare plain and html message parts
    part1 = MIMEText(body_plain, 'plain')
    part2 = MIMEText(body_html, 'html')
    # attach parts to message
    message.attach(part1)
    message.attach(part2)
    # convert recipient as list
    sendTo = list(sendTo.split(";"))
    sendCC = list(sendCC.split(";"))

    # attachment
    if attachment is not None and attachment != "":
        for index, bytesIOfile in enumerate(attachment):
            try:
                part3 = MIMEApplication(bytesIOfile.getvalue())
                application_type = mimetypes.guess_type("a.xlsx")[0] or 'application/octet-stream' + " ;charset=UTF-8"
                part3.add_header('Content-Disposition', 'attachment', filename=attachmentname[index])
                part3.add_header('Content-Type', application_type)
                message.attach(part3)
            except:
                pass

    # get local host and port
    host = os.environ.get('SMTP_HOST')
    port = os.environ.get('SMTP_PORT')

    # send the message
    server = smtplib.SMTP(host, port)
    server.ehlo()
    server.sendmail(sender, (sendTo + sendCC), message.as_string())
    server.close()

def convertRACFToAttendees(attendees):
    attendees_list = []

    for item in attendees.split(';'):
        code = item.strip()
        
        if not code:
            continue
            
        attendee = {
            "emailAddress": {
                "address": getStaffRecord(code)['staff']["email"]
            },
            "type": "optional"
        }
        attendees_list.append(attendee)

    return attendees_list


def addEventToCalendar(psRecord, psRefNo, attendees):

    # Get full leave content
    leaveContent = list(filter(lambda r: (r["ref_no"] == int(psRefNo)), psRecord["leave_record"]))

    # Get access token from auth
    access_token = session.get('access_token')

    headers = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/json"   
    }

    # Leave information
    typename = list(leaveTypes.find({'leave_type_id': leaveContent[0]["type"]}))[0]['leave_type']
    applicant = psRecord["staff"]["email"]
    applicant_name = psRecord["staff"]["name"]

    # Format the display name
    if "SICK" in typename:
        typename = "Sick Leave"
    else:
        typename = typename.title()

    for leaveitem in leaveContent[0]["details"]:
        s_date = datetime.strptime(str(leaveitem.get("start_date")), '%Y-%m-%d').strftime('%Y-%m-%d')
        e_date = datetime.strptime(str(leaveitem.get("end_date")), '%Y-%m-%d').strftime('%Y-%m-%d')
        s_time = leaveitem.get("start_time")
        e_time = leaveitem.get("end_time")

        if s_time == "Full Day" and e_time == "Full Day":
            s_time, e_time = "AM", "PM"

    timeZone = leaveContent[0]['timeZone']

    # Calculate start time and end time
    if s_time == "AM" and e_time == "AM": 
        hour1 = "08"
        mintues1 = "00"
        hour2 = "13"
        mintues2 = "30"
    elif s_time == "AM" and e_time == "PM": 
        hour1 = "08"
        mintues1 = "00"
        hour2 = "17"
        mintues2 = "30"
    elif s_time == "PM" and e_time == "PM": 
        hour1 = "13"
        mintues1 = "30"
        hour2 = "17"
        mintues2 = "30"
    elif s_time == "PM" and e_time == "AM": 
        hour1 = "13"
        mintues1 = "30"
        hour2 = "08"
        mintues2 = "00"


    attendees_list = convertRACFToAttendees(attendees)
    
    attendees_list.insert(0, {
        "emailAddress": {
            "address": applicant
        },
        "type": "required"
    })

    event = {
        
            "subject": f"{typename} - {applicant_name}",

            "start": {
                "dateTime": f"{s_date}T{hour1}:{mintues1}:00",
                "timeZone": f"{timeZone}"
            },
            "end": {
                "dateTime": f"{e_date}T{hour2}:{mintues2}:00",
                "timeZone": f"{timeZone}"
            },
            "attendees": attendees_list
    }

    # Construct the upload URL    
    url = f"https://graph.microsoft.com/v1.0/me/events"
            
    print(f"Add event in the calendar...")
            
    response = requests.post(url, headers=headers, data=json.dumps(event))
        
    if response.status_code == 201:
        event_data = response.json()
        event_id = event_data["id"]
        leave_index = next((i for i, r in enumerate(psRecord["leave_record"]) if r["ref_no"] == int(psRefNo)), None)
        
        if leave_index is not None:
            eleaveDtl.update_one(
                {"_id": psRecord["_id"]},
                {"$set": {f"leave_record.{leave_index}.event_id": event_id}}
            )
    else:
        print(f"Getting the list failed with status code: {response.status_code}")
        print(response.json())     

        leave_index = next((i for i, r in enumerate(psRecord["leave_record"]) if r["ref_no"] == int(psRefNo)), None)
        
        eleaveDtl.update_one(
            {"_id": psRecord["_id"]},
            {"$set": {f"leave_record.{leave_index}.event_id": "error"}}
        )  
        return "error" 

def cancelCalendar(event_id):

    # Get access token from auth
    access_token = session.get('access_token')

    headers = {
        "Authorization": f"Bearer {access_token}",
        "Content-Type": "application/json"   
    }

    # content
    body = {
        "comment" : "Meeting Cancelled."
    }

                    
    # Construct the api URL    
    url = f"https://graph.microsoft.com/v1.0/me/events/{event_id}/cancel"
            
    print(f"Deleting the event in the calendar...")
           
    response = requests.post(url, headers=headers, data=json.dumps(body))
        
    if response.status_code in [200, 201, 202]:
        print("Calendar event deleted successfully !")
        if response.status_code == 201:
            event_data = response.json()
            event_id = event_data["id"] # Extract event id            
            
    else:
        print(f"Getting the list failed with status code: {response.status_code}")
        print(response.json())     

def sendEmail(psRecord, psRefNo, otherRefNo, psAction, psRequest, finalapprover = 1, currentapprover = 1):


    leaveContent = list(filter(lambda r: (r["ref_no"] == int(psRefNo)), psRecord["leave_record"]))
    # print (psRecord["leave_record"])
    leavePeriod = ""
    for leaveitem in leaveContent[0]["details"]:
        start_date = datetime.strptime(str(leaveitem.get("start_date")), '%Y-%m-%d').strftime('%m/%d/%Y')
        start_time = leaveitem.get("start_time")
        end_date = datetime.strptime(str(leaveitem.get("end_date")), '%Y-%m-%d').strftime('%m/%d/%Y')
        end_time = leaveitem.get("end_time")
        leavePeriod = leavePeriod + datetime.strptime(str(leaveitem.get("start_date")), '%Y-%m-%d').strftime('%m/%d/%Y') + " " + leaveitem.get("start_time") + " to " + datetime.strptime(str(leaveitem.get("end_date")), '%Y-%m-%d').strftime('%m/%d/%Y') + " " + leaveitem.get("end_time") + "\n"
    
    # Make Full to AM PM
    if (start_time == "Full Day" and end_time == "Full Day"): 
        start_time = "AM"
        end_time = "PM"

    # Avoid input error
    if start_time != "Full Day" and end_time == "Full Day":
        end_time = "PM"
    elif start_time == "Full Day" and end_time != "Full Day":
        start_time = end_time
    
    # Make email list for sending out to specific recipient by defined cc_general in MongoDB
    try:
        cc_general = str(psRecord["staff"]["cc_general"]).replace(",", ";")
        cc_general = cc_general.split(";")
        for index, recipient in enumerate(cc_general):
            cc_general[index] = getStaffRecord(recipient.strip())["staff"]["alteremail"] if getStaffRecord(recipient.strip())["staff"]["alteremail"] is not None else getStaffRecord(recipient.strip())['staff']["email"]
        cc_general_list = ';'.join(map(str, cc_general))
    except:
        cc_general_list = ""

    sickleave_count = 0

    for leaveitem in leaveContent[0]["details"]:
        try:
            if (leaveitem.get("no_of_consective") > sickleave_count) and (leaveContent[0]["type"] in ['LVE04', 'LVE05']): sickleave_count = leaveitem.get("no_of_consective")
        except:
            sickleave_count = 0

    cc_sl_limit_list = ""

    if (int(sickleave_count) > 2) :
        # Make email list for sending out to specific recipient by defined cc_sl_limit in MongoDB
        cc_sl_limit = str(psRecord["staff"]["cc_sl_limit"]).replace(",", ";")
        cc_sl_limit = cc_sl_limit.split(";")
        for index, recipient in enumerate(cc_sl_limit):
            try:
                cc_sl_limit[index] = getStaffRecord(recipient.strip())['staff']["alteremail"] if getStaffRecord(recipient.strip())['staff']["alteremail"] is not None else getStaffRecord(recipient.strip())["staff"]["email"]
            except:
                pass
        
        cc_sl_limit_list = ';'.join(cc_sl_limit)
    
    cc_sl_limit_list = str(cc_sl_limit_list)

    typename = list(leaveTypes.find({'leave_type_id': leaveContent[0]["type"]}))[0]['leave_type']

    tz = leaveContent[0]["timeZone"]

    cc_no_pay = ""
    if leaveContent[0]["type"] == "LVE06":
        # Make email list for sending out to specific recipient by defined cc_sl_limit in MongoDB
        cc_no_pay = str(psRecord["staff"]["cc_no_pay"]).replace(",", ";")
        cc_no_pay = cc_no_pay.split(";")
        for index, recipient in enumerate(cc_no_pay):
            try:
                cc_no_pay[index] = getStaffRecord(recipient.strip())['staff']["alteremail"] if getStaffRecord(recipient.strip())['staff']["alteremail"] is not None else getStaffRecord(recipient.strip())["staff"]["email"]
            except:
                pass
        
        cc_no_pay = ';'.join(cc_no_pay)
    
    cc_no_pay = str(cc_no_pay)

    # Sharepoint document
    sickleavewithcertmsg = ""
    site = "https://macysinc.sharepoint.com/sites/MMGOverseas/eleave" + str(leaveContent[0]["year"]) + "/Forms/AllItems.aspx?id=%2Fsites%2FMMGOverseas%2Feleave" + str(leaveContent[0]["year"]) +"%2F" + str(psRecord["staff"]["racf"]) + "%2F&FilterType1=Text&viewid=720379f5-eb23-410f-81d1-f4e43a1a1cab&FilterField1=SharePointID&FilterValue1=" + str(leaveContent[0]["sharePointId"])
    if leaveContent[0]["type"] == "LVE04": sickleavewithcertmsg = "\n" + "Please download the supporting document(s) via the SharePoint link: " + "\n" + str(site)
    else: sickleavewithcertmsg = ""

    # ical file
    icsmessage = "(Optional) You may double-click the Outlook Calendar.ics to add this event to your calendar.  If you cancel it later, you will need to manually remove this event from the calendar. "
    cancelicsmessage = "If this calendar event has been previously added, please remember to manually delete this event."

    # Manual make current approver
    next_approver = ""
    if currentapprover == 1 and finalapprover != 1:
        if (psRequest == df['gcActionCancel'][0]) and (psAction == df['gcActionCancel'][0]):
            next_approver = getStaffRecord(psRecord["staff"]["approver2"])["staff"]["alteremail"] if getStaffRecord(psRecord["staff"]["approver2"])["staff"]["alteremail"] is not None else getStaffRecord(psRecord["staff"]["approver2"])["staff"]["email"]
        else:
            next_approver = getStaffRecord(leaveContent[0]["approval"].get("approver2"))["staff"]["alteremail"] if getStaffRecord(leaveContent[0]["approval"].get("approver2"))["staff"]["alteremail"] is not None else getStaffRecord(leaveContent[0]["approval"].get("approver2"))['staff']["email"]
    elif currentapprover == 1 and finalapprover == 1: 
        if ((psRequest == df['gcActionCancel'][0]) and (psAction == df['gcActionCancel'][0])) == False:
            next_approver = ""
    elif currentapprover == 2 and finalapprover == 3:
        if ((psRequest == df['gcActionCancel'][0]) and (psAction == df['gcActionCancel'][0])) == False:
            next_approver = getStaffRecord(leaveContent[0]["approval"].get("approver3"))["staff"]["alteremail"] if getStaffRecord(leaveContent[0]["approval"].get("approver3"))["staff"]["alteremail"] is not None else getStaffRecord(leaveContent[0]["approval"].get("approver3"))['staff']["email"]
    elif currentapprover == 2 and finalapprover == 2: 
        if ((psRequest == df['gcActionCancel'][0]) and (psAction == df['gcActionCancel'][0])) == False:
            next_approver = ""

    # Manual make reference number in HR format i.e. REG2022001KWY
    ref_no = getDisplayRefNo(psRefNo=int(psRefNo), psOffice=psRecord["staff"]["hr_office"], psRacf=psRecord["staff"]["racf"])

    if otherRefNo == "":
        otherRefNo = ref_no

    # Manual make application status per pending situation
    Approval_Status = applicationStatusForEmail(leaveContent, finalapprover, psRequest, psAction)

    # Apply/Cancel leave must send to approver 1
    if ((psRequest == df['gcActionApply'][0] and psAction == df['gcActionApply'][0])):
        # Send to approver 1 first 
        sendTo = getStaffRecord(psRecord["staff"]["approver1"])["staff"]["alteremail"] if getStaffRecord(psRecord["staff"]["approver1"])["staff"]["alteremail"] is not None else getStaffRecord(psRecord["staff"]["approver1"])['staff']["email"]
        sendCc = ""
        title = "<E-LEAVE> " + str(psRecord["staff"]["name"]) + " (" + str(psRecord["staff"]["dept"]) + ") " + " - " + "Apply " + str(typename) + " #PENDING"
        message = "Dear People Leader," + "\n" + "\n"  + "Please click the following link to approve:  https://eleave.mmgoverseas.app/#/ApprovalCenter" + "\n" + "\n" + "Leave Period" + "\n" + leavePeriod + "\n" + "Thanks," + "\n" + "e-Leave"
        try:
            postmarker(message, title, sendTo, sendCc, None, None)
        except:
            localSend(message, title, sendTo, sendCc)
        # Send confirmation email to application
        sendTo = psRecord["staff"]["alteremail"] if psRecord["staff"]["alteremail"] is not None else psRecord["staff"]["email"]
        sendCc = ""
        title = "<E-LEAVE> " + str(psRecord["staff"]["name"]) + " (" + str(psRecord["staff"]["dept"]) + ") " + " - " + "Apply " + str(typename)
        message = "Dear Applicant," + "\n" + "\n" + "Your application has been sent to your people leader for approval successfully. "  + "\n" + "\n" + "Leave Period" + "\n" + leavePeriod + "\n"  + "\n"  + "Thanks," + "\n" + "e-Leave"
        try:
            postmarker(message, title, sendTo, sendCc, None, None)
        except:
            localSend(message, title, sendTo, sendCc)
    elif (psRequest == df['gcActionCancel'][0]) and (psAction == df['gcActionCancel'][0]):
        sendTo = getStaffRecord(psRecord["staff"]["approver1"])["staff"]["alteremail"] if getStaffRecord(psRecord["staff"]["approver1"])["staff"]["alteremail"] is not None else getStaffRecord(psRecord["staff"]["approver1"])['staff']["email"]
        sendCc = ""
        title = "<E-LEAVE> " + str(psRecord["staff"]["name"]) + " (" + str(psRecord["staff"]["dept"]) + ") " + " - " + "Cancel " + str(typename) + " #PENDING"
        message = "Dear People Leader," + "\n" + "\n"  + "Please click the following link to approve:  https://eleave.mmgoverseas.app/#/ApprovalCenter" + "\n" + "\n" + "Leave Period" + "\n" + leavePeriod + "\n" + "Thanks," + "\n" + "e-Leave"
        try:
            postmarker(message, title, sendTo, sendCc, None, None)
        except:
            localSend(message, title, sendTo, sendCc)   
        # Send confirmation email to application
        sendTo = psRecord["staff"]["alteremail"] if psRecord["staff"]["alteremail"] is not None else psRecord["staff"]["email"]
        sendCc = ""
        title = "<E-LEAVE> " + str(psRecord["staff"]["name"]) + " (" + str(psRecord["staff"]["dept"]) + ") " + " - " + "Cancel " + str(typename)
        message = "Dear Applicant," + "\n" + "\n" + "Your application has been sent to your people leader for cancel approval successfully. "  + "\n" + "\n" + "Leave Period" + "\n" + leavePeriod + "\n"  + "\n"  + "Thanks," + "\n" + "e-Leave"
        try:
            postmarker(message, title, sendTo, sendCc, None, None)
        except:
            localSend(message, title, sendTo, sendCc)
    # Approved by approver must send to applicant and next approver , if it is final approver, it will send out the sick leave limit list to HR
    if psAction == df['gcActionApprove'][0] and psRequest == df['gcActionApply'][0]:
        sendTo = psRecord["staff"]["alteremail"] if psRecord["staff"]["alteremail"] is not None else psRecord["staff"]["email"]
        if finalapprover == currentapprover:
            sendCc = ";".join([p for p in [cc_general_list, cc_sl_limit_list, cc_no_pay] if p and p.strip()])
            title = "<E-LEAVE> " + str(psRecord["staff"]["name"]) + " (" + str(psRecord["staff"]["dept"]) + ") " + " - " + "Apply " + str(typename) + " #APPROVED"
            if "SICK" in str(typename).upper() and int(sickleave_count) > 2:
                message = "Dear Applicant, " + "\n" + "\n"  + "Approval Status :" + "\n" + Approval_Status  + "\n" + "Leave Period" + "\n" + leavePeriod + sickleavewithcertmsg + "\n" + "*Reminder: The total number of sick leave taken is " + str(sickleave_count) + " consecutive days" + "\n\n" + icsmessage + "\n" + "\n"  + "Thanks," + "\n" + "e-Leave"
            else:
                message = "Dear Applicant, " + "\n" + "\n"  + "Approval Status :" + "\n" + Approval_Status  + "\n" + "Leave Period" + "\n" + leavePeriod + sickleavewithcertmsg + "\n" + icsmessage + "\n"  + "\n"  + "Thanks," + "\n" + "e-Leave"
            attached_list = [getApplicationForm(ref_no, psRecord["staff"]["racf"]),
                             getSummaryForm(leaveContent[0]["year"], psRecord["staff"]["racf"]),
                             geticalFile(sendTo,str(psRecord["staff"]["name"])+" - "+str(typename).title(),str(typename).title() + " Period : " + leavePeriod, start_date, start_time, end_date, end_time, tz)
                            ]
            filename_list = ["Leave Record for "+ otherRefNo + " ("+ psRecord["staff"]["racf"] + ")" + ".xlsx",
                             "Leave Summary for " + str((leaveContent[0]["year"])) + " (" + str(psRecord["staff"]["racf"]) + ")" + ".xlsx",
                             "Outlook Calendar.ics"
                            ]
            try:
                postmarker(message, title, sendTo, sendCc, attached_list, filename_list)
            except:
                localSend(message, title, sendTo, sendCc, attached_list, filename_list)

            addCalendar = leaveContent[0]['addCalendar']

            if addCalendar:
                attendees = psRecord["staff"]["calendarAttendees"]
                addEventToCalendar(psRecord, psRefNo, attendees)

    # Reject end instantly (Apply)
    if psAction == df['gcActionReject'][0] and (psRequest == df['gcActionApply'][0]):
        sendTo = psRecord["staff"]["alteremail"] if psRecord["staff"]["alteremail"] is not None else psRecord["staff"]["email"]
        sendCc = ";".join([p for p in [cc_general_list, cc_sl_limit_list] if p and p.strip()])
        title = "<E-LEAVE> " + str(psRecord["staff"]["name"]) + " (" + str(psRecord["staff"]["dept"]) + ") " + " - " + "Apply " + str(typename) + " #REJECTED"
        message = "Dear Applicant, " + "\n" + "\n"  + "Approval Status :" + "\n" + Approval_Status + "\n" + "Leave Period" + "\n" + leavePeriod + "\n" + "Thanks," + "\n" + "e-Leave"
        try:
            postmarker(message, title, sendTo, sendCc)
        except:
            localSend(message, title, sendTo, sendCc)
    # Reject end instantly (Cancel)
    if psAction == df['gcActionReject'][0] and (psRequest == df['gcActionCancel'][0]):
        sendTo = psRecord["staff"]["alteremail"] if psRecord["staff"]["alteremail"] is not None else psRecord["staff"]["email"]
        sendCc = ";".join([p for p in [cc_general_list, cc_sl_limit_list] if p and p.strip()])
        title = "<E-LEAVE> " + str(psRecord["staff"]["name"]) + " (" + str(psRecord["staff"]["dept"]) + ") " + " - " + "Cancel " + str(typename) + " #REJECTED"
        message = "Dear Applicant, " + "\n" + "\n"  + "Cancel Approval Status :" + "\n" + Approval_Status + "\n" + "Leave Period" + "\n" + leavePeriod + "\n" + "Thanks," + "\n" + "e-Leave"
        try:
            postmarker(message, title, sendTo, sendCc, None, None)
        except:
            localSend(message, title, sendTo, sendCc)

    # Approved by approver must send to applicant and next approver , if it is final approver, it will send out the sick leave limit list to HR
    if psAction == df['gcActionApprove'][0] and psRequest == df['gcActionCancel'][0]:
        sendTo = psRecord["staff"]["alteremail"] if psRecord["staff"]["alteremail"] is not None else psRecord["staff"]["email"]
        if finalapprover == currentapprover:
            sendCc = ";".join([p for p in [cc_general_list, cc_sl_limit_list, cc_no_pay] if p and p.strip()])
            title = "<E-LEAVE> " + str(psRecord["staff"]["name"]) + " (" + str(psRecord["staff"]["dept"]) + ") " + " - " + "Cancel " + str(typename) + " #APPROVED"
            message = "Dear Applicant, " + "\n" + "\n"  + "Cancel Approval Status :" + "\n" + Approval_Status + "\n" + "Leave Period" + "\n" + leavePeriod + "\n"+ cancelicsmessage + "\n" + "\n" + "Thanks," + "\n" + "e-Leave"
            try:
                postmarker(message, title, sendTo, sendCc, None, None)
            except:
                localSend(message, title, sendTo, sendCc)

            addCalendar = leaveContent[0]['addCalendar']

            if addCalendar:
                event_id = leaveContent[0]['event_id']
                cancelCalendar(event_id)

    # Send pending leave to next approver (Cancel)
    if (finalapprover > currentapprover) and (psAction == df['gcActionApprove'][0] and psRequest == df['gcActionCancel'][0]):
        sendTo = next_approver
        sendCc = ""
        title = "<E-LEAVE> " + str(psRecord["staff"]["name"]) + " (" + str(psRecord["staff"]["dept"]) + ") " + " - " + "Cancel " + str(typename) + " #PENDING"
        message = "Dear People Leader," + "\n" + "\n"  + "Please click the following link to approve:  https://eleave.mmgoverseas.app/#/ApprovalCenter" + "\n" + "\n" + "Leave Period" + "\n" + leavePeriod + "\n" + "Thanks," + "\n" + "e-Leave"
        try:
            postmarker(message, title, sendTo, sendCc, None, None)
        except:
            localSend(message, title, sendTo, sendCc)

    # Send pending leave to next approver (Apply)
    if (finalapprover > currentapprover) and (psAction == df['gcActionApprove'][0] and psRequest == df['gcActionApply'][0]):
        sendTo = next_approver
        sendCc = ""
        title = "<E-LEAVE> " + str(psRecord["staff"]["name"]) + " (" + str(psRecord["staff"]["dept"]) + ") " + " - " + "Apply " + str(typename) + " #PENDING"
        message = "Dear People Leader," + "\n" + "\n"  + "Please click the following link to approve:  https://eleave.mmgoverseas.app/#/ApprovalCenter" + "\n" + "\n" + "Leave Period" + "\n" + leavePeriod + "\n" + "Thanks," + "\n" + "e-Leave"
        try:
            postmarker(message, title, sendTo, sendCc, None, None)
        except:
            localSend(message, title, sendTo, sendCc)

def getSummerHours(office, year):

    result = summer_hours.find_one(
        {"Office": office, "Year": year}
    )
    return result

def getPublicHolidays(office, start_date, end_date):

    # Getting list for public holiday including weekend and public holidays stored at MongoDB
    calendar_d = [ ]

    holiday_d = list(holidays.find({'Office': office}))

    #rcs = start_date # record checking start
    #rce = end_date # record checking end

    rcs = (start_date - relativedelta(days=365)).strftime('%Y-%m-%d') # record checking start
    rce = (end_date + relativedelta(days=365)).strftime('%Y-%m-%d') # record checking end    

    check_range =  pd.date_range(rcs, rce).tolist()

    for date in check_range:
        calendar_d.append({'Date': f"{date.strftime('%m')}/{date.strftime('%d')}/{date.strftime('%Y')}", 'Time': "AM", 'Day of Week': date.strftime('%A')})
        calendar_d.append({'Date': f"{date.strftime('%m')}/{date.strftime('%d')}/{date.strftime('%Y')}", 'Time': "PM", 'Day of Week': date.strftime('%A')})

    holiday_lst = [ ]


    for cal in calendar_d:
        
        # Date 
        date = datetime.strptime(str(cal['Date']), '%m/%d/%Y').strftime('%Y-%m-%d')
        # Time
        time = cal['Time']
        # Day of week
        dw = datetime.strptime(str(cal['Date']), '%m/%d/%Y').strftime('%A')
        # Remark
        remark = next((ph['Remark'] for ph in holiday_d if str(ph['Date']) == str(date) and str(ph['Time']) == time), "")

        if remark == "":
            if dw == "Saturday" or dw == "Sunday":
                remark = "Weekend"

        if remark != "":
            data = {
                    "Date": date,
                    "Time": time,
                    "Remark": remark
            }

            holiday_lst.append(data)
        

    return holiday_lst

def getAllApply(start_date, start_time, end_date, end_time, type, office, otherLeaveRef):

    continuous_dates = []

    current_date = start_date
    current_time = start_time

    holiday_d = getPublicHolidays(office, start_date, end_date)

    # Handle Inclusive Weekend checking
    if otherLeaveRef == "":
        special_leave = False
        exclusive = True
    else:
        special_leave = True
        if (list(leaveTypes.find({'leave_type_id': type}))[0]['other_leave'] and otherLeaveRef[0]['excluded_holidays']):
            exclusive = True
        else:
            exclusive = False

    # Normal leave and speical leave with exclusive weekend :
    if (special_leave == False) or (special_leave and exclusive):

        while current_date < end_date or (current_date == end_date and current_time <= end_time):

            date = current_date.strftime('%Y-%m-%d')

            if next((ph['Remark'] for ph in holiday_d if str(ph['Date']) == str(date) and str(ph['Time']) == current_time), "") == "":

                continuous_dates.append({
                    'applied_date': date,
                    'applied_time': current_time,
                    'applied_type': type
                })
                
            if current_time == 'AM':
                current_time = 'PM'
            else:
                current_time = 'AM'
                current_date += relativedelta(days=1)

    elif special_leave and exclusive == False:

        while current_date < end_date or (current_date == end_date and current_time <= end_time):

            date = current_date.strftime('%Y-%m-%d')

            # if next((ph['Remark'] for ph in holiday_d if str(ph['Date']) == str(date) and str(ph['Time']) == current_time), "") == "":

            continuous_dates.append({
                'applied_date': date,
                'applied_time': current_time,
                'applied_type': type
            })
                
            if current_time == 'AM':
                current_time = 'PM'
            else:
                current_time = 'AM'
                current_date += relativedelta(days=1)
                
    return continuous_dates

def getAllLeave(racf, year, leavetype, consecutive_search = False, otherRefNo = ""):

    leave_d = list(eleaveDtl.find({'staff.racf': racf}))

    if consecutive_search:
        # get same type group
        leave_t = list(leaveTypes.find({'leave_type_id': leavetype}))
        leave_g = list(leaveTypes.find({'consecutive_days_group': leave_t[0]['consecutive_days_group']}))
        same_leave = [leave['leave_type_id'] for leave in leave_g]
    elif type(leavetype) is list:
        leave_t = list(leaveTypes.find({'leave_type_id': {'$in': leavetype}}))
        same_leave = [leave['leave_type_id'] for leave in leave_t]
    else:
        leave_t = list(leaveTypes.find({}))
        same_leave = [leave['leave_type_id'] for leave in leave_t]

    history = [ ]

    for staff in leave_d:
        for leave_record in staff['leave_record']:
            for details in leave_record['details']:
                for period in details['period']:
                    if leave_record['type'] in same_leave and otherRefNo != "":
                        if leave_record['applicationStatus'] != df['gcStatusReject'][0] and leave_record['applicationStatus'] != df['gcStatusCancel'][0] and leave_record['otherRefNo'] == otherRefNo:
                            data = {
                                "applied_date": period['ldate'],
                                "applied_time": period['ltime'],
                                "applied_type": leave_record['type']
                            }

                            history.append(data)
                    elif leave_record['type'] in same_leave and otherRefNo == "":
                        if consecutive_search:
                            if (leave_record['year'] == year or leave_record['year'] == year - 1) and leave_record['applicationStatus'] != df['gcStatusReject'][0] and leave_record['applicationStatus'] != df['gcStatusCancel'][0]:
                                data = {
                                    "applied_date": period['ldate'],
                                    "applied_time": period['ltime'],
                                    "applied_type": leave_record['type']
                                }

                                history.append(data)
                        else:
                            # Normal leave within year or other leave without year limitation check
                            if ((leave_record['year'] == year) and leave_record['applicationStatus'] != df['gcStatusReject'][0] and leave_record['applicationStatus'] != df['gcStatusCancel'][0]) or (leave_record['otherRefNo'] != "" and leave_record['applicationStatus'] != df['gcStatusReject'][0] and leave_record['applicationStatus'] != df['gcStatusCancel'][0]):
                                data = {
                                    "applied_date": period['ldate'],
                                    "applied_time": period['ltime'],
                                    "applied_type": leave_record['type']
                                }

                                history.append(data)
    
    return history

def checkSummerHoursPeriod(office, year, psInput):

    # Get summer hours list
    summer = getSummerHours(office, year)

    if summer:

        # Parse DB strings to datetime objects for accurate comparison
        summer_start = datetime.strptime(summer["Date_Start"], "%Y-%m-%d")
        summer_end = datetime.strptime(summer["Date_End"], "%Y-%m-%d")

        # Loop through each item using index so we can modify the original dictionary
        for i in range(len(psInput["applying"])):
            app = psInput["applying"][i]

            # Parse application start and end dates
            app_start = datetime.strptime(app["startDate"], "%Y-%m-%d")
            app_end = datetime.strptime(app["endDate"], "%Y-%m-%d")    

            # Parse application start and end time
            start_time = app["startTime"]
            end_time = app["endTime"]

            # Check if the date is a Friday (weekday == 4)
            is_start_friday = app_start.weekday() == 4
            is_end_friday = app_end.weekday() == 4

            # Check if dates fall completely within the summer hours window
            is_start_in_summer = summer_start <= app_start <= summer_end
            is_end_in_summer = summer_start <= app_end <= summer_end

            # 3. If it's a Friday inside the summer period
            # 3.1 check start date in summer hours first
            if is_start_in_summer and is_start_friday and start_time == "PM":
                return True
            if is_end_in_summer and is_end_friday and end_time == "AM":
                return True


def checkConsecutive(racf, year, apply_h, type, office):

    spt = datetime.now()

    leave_h = getAllLeave(racf, year, type, True)

    min_date = min(apply_h, key=lambda x: x['applied_date'])['applied_date']
    max_date = max(apply_h, key=lambda x: x['applied_date'])['applied_date']

    min_date = datetime.strptime(min_date, '%Y-%m-%d')
    max_date = datetime.strptime(max_date, '%Y-%m-%d')

    #########################################
    # Step 1 : Make the date array before start date 30 days, and after end date 30 days
    #########################################
    calendar_d = [ ]

    rcs = (min_date - relativedelta(days=30)).strftime('%Y-%m-%d') # record checking start
    rce = (max_date + relativedelta(days=30)).strftime('%Y-%m-%d') # record checking end

    check_range =  pd.date_range(rcs, rce).tolist()

    # holiday list
    holiday_h = getPublicHolidays(office, datetime.strptime(rcs, '%Y-%m-%d'), datetime.strptime(rce, '%Y-%m-%d'))

    # Make the range df for (given_date before 60 days) to (given_date after 60 days)
    for date in check_range:
        calendar_d.append({'Date': f"{date.strftime('%m')}/{date.strftime('%d')}/{date.strftime('%Y')}", 'Time': "AM", 'Day of Week': date.strftime('%A')})
        calendar_d.append({'Date': f"{date.strftime('%m')}/{date.strftime('%d')}/{date.strftime('%Y')}", 'Time': "PM", 'Day of Week': date.strftime('%A')})

    cdf = pd.DataFrame(calendar_d)

    #########################################
    # Step 2 : Add Public Holiday & Weekend & Applied Leave in date array - column : Applied 
    #          Add Applying Leave in date array  - column : Apply
    #########################################

    result_d = [ ]

    for cindex, ckrow in cdf.iterrows():
        # Date
        check_date = datetime.strptime(str(ckrow['Date']), '%m/%d/%Y').strftime('%Y-%m-%d')
        # Time
        check_time = str(ckrow['Time'])
        # Mon, Tue, Wed, ...
        check_dow = str(ckrow['Day of Week'])

        # Applied Checking
        # 1. public holiday first
        # 2. weekend
        # 3. applied leave type
        # 4. Emtpy
        applied = ""
        applied = next((ph['Remark'] for ph in holiday_h if str(ph['Date']) == str(check_date) and str(ph['Time']) == check_time), "")

        if applied == "":
            if check_dow == "Saturday" or check_dow == "Sunday":
                applied = "Weekend"
            else:
                applied = next((his['applied_type'] for his in leave_h if str(his['applied_date']) == str(check_date) and str(his['applied_time']) == check_time), "")

        # Apply
        apply = next((cur['applied_type'] for cur in apply_h if str(cur['applied_date']) == str(check_date) and str(cur['applied_time']) == check_time), "")

        data = {
            "Date": check_date,
            "Time": check_time,
            "Day of Week": check_dow,
            "Applied": applied,
            "Apply": apply
            }
        
        result_d.append(data)

    rdf = pd.DataFrame(result_d)


    #########################################
    # Step 3 : Count the numbere of consecutive day based on the leave type
    #########################################

    # Initial Variable for counting
    checking_list = [ ]
    count = 0
    sl_count = 0 # need to return the day count requested by HR for displaying warning email and message in client
    no_pay_count = 0
    casual_count = 0

    # Get information from leave_groups e.g. maximum applied days for sick leave/ maximum consective days for annual
    try:
        max_al_days = list(leaveGroups.find({'groupID': list(leaveTypes.find({'leave_type_id': type}))[0]['leave_group']}))[0]['max_consecutive_days']
    except:
        max_al_days = 0

    # Get information from leave_groups number
    leave_group_no = (list(leaveTypes.find({'leave_type_id': type}))[0]['leave_group'])

    for i, r in rdf.iterrows():


        # General consective count
        if r['Applied'] != "" or r['Apply'] != "":
            count += 0.5
            checking_list.append(r['Apply'])


        # Annual Leave, Causal Leave 

        if type == "LVE01" or type == "LVE02" or leave_group_no == 1:

            # Return error if the 14 days consective is applying, not the past applied (No including No-pay because No-pay cannot exceed 5 working days)
            if count > max_al_days and any(value != '' for value in checking_list) and type != "LVE06":
                return ({"consecutive": True, "error_message" : "Reminder: Maximum vacation taken at any one time is 2 WEEKS including Public Holidays, Saturdays and Sundays", "result": None,  "Status_code": 506, "no_of_consective": sl_count})

            # Special Checking for India: Cannot be consecutive 4
            if office == "DEL": 

                if r['Applied'] == "" and r['Apply'] == "":
                    casual_count = 0
                elif r['Applied'] == "LVE02" or r['Apply'] == "LVE02":
                    casual_count += 0.5

                if casual_count > 3:
                    return ({"consecutive": True, "error_message" : "Reminder: Casual leave cannot be applied for more than 3 consecutive days", "result": None,  "Status_code": 506, "no_of_consective": sl_count})

            # No pay leave cannot exceed 5 working days
            if type == "LVE06":

                if r['Applied'] == "" and r['Apply'] == "":
                    no_pay_count = 0
                elif r['Applied'] == "LVE06" or r['Apply'] == "LVE06":
                    no_pay_count += 0.5

                if no_pay_count > 5:
                    return ({"consecutive": True, "error_message" : "No pay Leave cannot exceed 5 working days. Please contact local HR for assistance if required", "result": None,  "Status_code": 506, "no_of_consective": sl_count})

            # Consider consecutive weekends only if a leave has been applied before
            if r['Day of Week'] == "Saturday" and rdf.iloc[i-1]['Applied'] == "" and rdf.iloc[i-1]['Apply'] == "":
                count = 0
                casual_count = 0
                no_pay_count = 0
            if (r['Day of Week'] == "Saturday" or r['Day of Week'] == "Sunday") and count <= 0.5:
                count = 0
                casual_count = 0
                no_pay_count = 0
            # Consider public holidays only if a leave has been applied before
            if r['Applied'] != "" and 'LVE' not in r['Applied'] and rdf.iloc[i-1]['Applied'] == "" and rdf.iloc[i-1]['Apply'] == "":
                count = 0
                casual_count = 0
                no_pay_count = 0
            if r['Applied'] != "" and 'LVE' not in r['Applied'] and count <= 0.5:
                count = 0
                casual_count = 0
                no_pay_count = 0

        # Sick Leave with medical cert
        elif type == "LVE04":

            # Always ignore consecutive day for weekends/public holiday
            if (r['Day of Week'] == "Saturday" or r['Day of Week'] == "Sunday"):
                count -= 0.5
            elif r['Applied'] != "" and 'LVE' not in r['Applied']:
                count -= 0.5
            
            # If the consective date included 2 no cert, then error (Might not need it)
            #if r['Applied'] == "LVE05":
            #    no_cert_count += 0.5
            #if no_cert_count > 1:
            #    return ({"consecutive": True, "error_message" : "Reminder: For any sick leave periods that exceed 2 contiguous days, sick leave certificate is required", "result": None, "Status_code": 506})

            if any(value != '' for value in checking_list) and count > 0:
                sl_count = count
        
        
        # Sick Leave without medical cert
        elif type == "LVE05":

            # Always ignore consecutive day for weekends/public holiday
            if (r['Day of Week'] == "Saturday" or r['Day of Week'] == "Sunday"):
                count -= 0.5
            elif r['Applied'] != "" and 'LVE' not in r['Applied']:
                count -= 0.5

            # No cert cannot more than 1 days for applying
            if checking_list.count(type) > 2:
                return ({"consecutive": True, "error_message" : "Reminder: For any sick leave periods that exceed 2 contiguous days, sick leave certificate is required", "result": None, "Status_code": 506, "no_of_consective": sl_count})

            # Return error if the 2 days consective is applying, not the past applied
            if count > 1 and any(value != '' for value in checking_list):
                return ({"consecutive": True, "error_message" : "Reminder: For any sick leave periods that exceed 2 contiguous days, sick leave certificate is required", "result": None, "Status_code": 506, "no_of_consective": sl_count})

            # Return sick leave total consective day for display
            if any(value != '' for value in checking_list) and count > 0:
                sl_count = checking_list.count(type) * 0.5

        # Break counter if applied & apply both blank
        if r['Applied'] == "" and r['Apply'] == "":
            count = 0
            
        # reset checking list if the chain is broken
        if count == 0:
            checking_list = [ ]
        
        # For developing checking
        # print (f"{r['Date']}/{r['Time']}/{r['Applied']}/{r['Apply']}, count : {count}, no pay count : {no_pay_count}, causal count: {casual_count}")

    # ept = datetime.now()
    # execution_time = ept - spt
    # print(f"Finsih Consecutive Checking Execution time: {execution_time.total_seconds()} seconds")
    
    return ({"consecutive": False, "error_message" : "Passed", "no_of_consective": sl_count})

def getOOOdays(racf, year, apply_h, type, office):

    leave_d = list(eleaveDtl.find({'staff.racf': racf}))

    # get same type group
    leave_t = list(leaveTypes.find({'leave_type_id': type}))
    leave_g = list(leaveTypes.find({'calendar_days_group': leave_t[0]['calendar_days_group']}))
    same_leave = [leave['leave_type_id'] for leave in leave_g]

    # start/ End date
    if len(apply_h) < 2:
        start_date =  datetime.strptime(apply_h[0]['applied_date'], '%Y-%m-%d')
        end_date = datetime.strptime(apply_h[0]['applied_date'], '%Y-%m-%d')
    else:
        start_date =  datetime.strptime(apply_h[0]['applied_date'], '%Y-%m-%d')
        end_date = datetime.strptime(apply_h[-1]['applied_date'], '%Y-%m-%d')

    # Applied
    leave_h = getAllLeave(racf, year, same_leave, False)

    # public holiday
    holiday_h = getPublicHolidays(office, start_date, end_date)

    rcs = (start_date - relativedelta(days=365)).strftime('%Y-%m-%d') # record checking start
    rce = (end_date + relativedelta(days=365)).strftime('%Y-%m-%d') # record checking end

    check_range =  pd.date_range(rcs, rce).tolist()
    calendar_d = [ ]

    # Make the range df for (given_date before 60 days) to (given_date after 300 days)
    for date in check_range:
        calendar_d.append({'Date': f"{date.strftime('%m')}/{date.strftime('%d')}/{date.strftime('%Y')}", 'Time': "AM", 'Day of Week': date.strftime('%A')})
        calendar_d.append({'Date': f"{date.strftime('%m')}/{date.strftime('%d')}/{date.strftime('%Y')}", 'Time': "PM", 'Day of Week': date.strftime('%A')})

    cdf = pd.DataFrame(calendar_d)

    result_d = [ ]

    for cindex, ckrow in cdf.iterrows():
        # Date
        check_date = datetime.strptime(str(ckrow['Date']), '%m/%d/%Y').strftime('%Y-%m-%d')
        # Time
        check_time = str(ckrow['Time'])
        # Mon, Tue, Wed, ...
        check_dow = str(ckrow['Day of Week'])

        # Applied Checking
        # 1. public holiday first
        # 2. weekend
        # 3. applied leave type
        # 4. Emtpy
        applied = ""
        applied = next((ph['Remark'] for ph in holiday_h if str(ph['Date']) == str(check_date) and str(ph['Time']) == check_time), "")

        if applied == "":
            applied = next((his['applied_type'] for his in leave_h if str(his['applied_date']) == str(check_date) and str(his['applied_time']) == check_time), "")

        # Apply
        apply = next((cur['applied_type'] for cur in apply_h if str(cur['applied_date']) == str(check_date) and str(cur['applied_time']) == check_time), "")

        data = {
            "Date": check_date,
            "Time": check_time,
            "Day of Week": check_dow,
            "Applied": applied,
            "Apply": apply
            }
        
        result_d.append(data)

    rdf = pd.DataFrame(result_d)

    # checking list
    checking_list = [ ]
    count = 0
    result_ooo = 0
    for i, r in rdf.iterrows():

        #print (f"Date : {r['Date']}/{r['Time']} , Applied : {r['Applied']}, Applying : {r['Apply']}, count = {count}, result_ooo = {result_ooo}")

        # General consective count
        if r['Applied'] != "" or r['Apply'] != "":
            count += 0.5
            checking_list.append(r['Apply'])
        # Break counter if applied & apply both blank
        if r['Applied'] == "" and r['Apply'] == "":
                count = 0
        if any(value != '' for value in checking_list) and count > 0:
            result_ooo = count

        # reset checking list if the chain is broken
        if count == 0:
            checking_list = [ ]

    return float(result_ooo)

def getPhInclusiveWorkDays(racf, year, apply_h, type, office):

    # leave_d = list(eleaveDtl.find({'staff.racf': racf}))

    # get same type group
    # leave_t = list(leaveTypes.find({'leave_type_id': type}))
    # leave_g = list(leaveTypes.find({'calendar_days_group': leave_t[0]['calendar_days_group']}))
    # same_leave = [leave['leave_type_id'] for leave in leave_g]

    # start/ End date
    if len(apply_h) < 2:
        start_date =  datetime.strptime(apply_h[0]['applied_date'], '%Y-%m-%d')
        end_date = datetime.strptime(apply_h[0]['applied_date'], '%Y-%m-%d')
    else:
        start_date =  datetime.strptime(apply_h[0]['applied_date'], '%Y-%m-%d')
        end_date = datetime.strptime(apply_h[-1]['applied_date'], '%Y-%m-%d')

    # public holiday
    holiday_h = getPublicHolidays(office, start_date, end_date)

    rcs = (start_date - relativedelta(days=365)).strftime('%Y-%m-%d') # record checking start
    rce = (end_date + relativedelta(days=365)).strftime('%Y-%m-%d') # record checking end

    check_range =  pd.date_range(rcs, rce).tolist()
    calendar_d = [ ]

    # Make the range df for (given_date before 60 days) to (given_date after 300 days)
    for date in check_range:
        calendar_d.append({'Date': f"{date.strftime('%m')}/{date.strftime('%d')}/{date.strftime('%Y')}", 'Time': "AM", 'Day of Week': date.strftime('%A')})
        calendar_d.append({'Date': f"{date.strftime('%m')}/{date.strftime('%d')}/{date.strftime('%Y')}", 'Time': "PM", 'Day of Week': date.strftime('%A')})

    cdf = pd.DataFrame(calendar_d)

    result_d = [ ]

    for cindex, ckrow in cdf.iterrows():
        # Date
        check_date = datetime.strptime(str(ckrow['Date']), '%m/%d/%Y').strftime('%Y-%m-%d')
        # Time
        check_time = str(ckrow['Time'])
        # Mon, Tue, Wed, ...
        check_dow = str(ckrow['Day of Week'])

        # Applied Checking
        # 1. public holiday first
        # 2. weekend
        # 3. applied leave type
        # 4. Emtpy
        applied = ""
        applied = next((ph['Remark'] for ph in holiday_h if str(ph['Date']) == str(check_date) and str(ph['Time']) == check_time), "")

        # Apply
        apply = next((cur['applied_type'] for cur in apply_h if str(cur['applied_date']) == str(check_date) and str(cur['applied_time']) == check_time), "")

        data = {
            "Date": check_date,
            "Time": check_time,
            "Day of Week": check_dow,
            "Applied": applied,
            "Apply": apply
            }
        
        result_d.append(data)

    rdf = pd.DataFrame(result_d)

    # checking list
    checking_list = [ ]
    count = 0
    result = 0
    for i, r in rdf.iterrows():

        #print (f"Date : {r['Date']}/{r['Time']} , Applied : {r['Applied']}, Applying : {r['Apply']}, count = {count}, result = {result}")

        # General consective count
        if r['Applied'] != "" or r['Apply'] != "":
            count += 0.5
            checking_list.append(r['Apply'])
        # Consider consecutive weekends only if a leave has been applied before
        if r['Day of Week'] == "Saturday" and rdf.iloc[i-1]['Applied'] == "" and rdf.iloc[i-1]['Apply'] == "":
            count = 0
        if (r['Day of Week'] == "Saturday" or r['Day of Week'] == "Sunday") and count <= 0.5:
            count = 0
        # Break counter if applied & apply both blank
        if r['Applied'] == "" and r['Apply'] == "":
                count = 0
        if any(value != '' for value in checking_list) and count > 0:
            result = count

        # reset checking list if the chain is broken
        if count == 0:
            checking_list = [ ]

    return float(result)




def applyLeave (psInput):

    # print (psInput)

    # Special handling other leave
    if psInput['otherLeaveRef'] == "":
        type = psInput['type']
        year = psInput['year']
        otherRefNo = ""
        otherLeaveRef = ""
        
    else:
        # Get otherLeaveRef
        otherLeaveRef = psInput['otherLeaveRef'].copy()
        type = otherLeaveRef[0]['leave_type']
        year =  otherLeaveRef[0]['year']
        otherRefNo = otherLeaveRef[0]['ref_no']

        del psInput['otherLeaveRef']


    # Read Input from client input
    xdf = pd.DataFrame(psInput)

    # Read other general parameter from client input
    racf = psInput['racf']
    office = psInput['office']
    submit = psInput['updateDB']
    #submit = False
    timeZone = psInput['timeZone']
    spid = psInput['sharePointId']
    super = psInput['superUser']
    addCalendar = psInput['addCalendar']

    # Determine Value
    is_other_leave = leaveTypes.find_one({'leave_type_id': type}).get('other_leave', False) if leaveTypes.find_one({'leave_type_id': type}) else False
    excluded_holidays = otherLeaveRef[0].get('excluded_holidays', False) if otherLeaveRef else False

    # warning output
    warnings = ""

    # List storing
    allApplying = [ ]
    allDetails = [ ]

    # Accumulated workdays for other leave
    acc_workdays = 0

    for index, row in xdf.iterrows():

        # Date
        start_date = datetime.strptime(row['applying']['startDate'], '%Y-%m-%d')
        end_date = datetime.strptime(row['applying']['endDate'], '%Y-%m-%d')
        # Time
        start_time = row['applying']['startTime']
        end_time = row['applying']['endTime']
        # Time on screen
        start_time_os = row['applyingScreen']['startTime']
        end_time_os = row['applyingScreen']['endTime']

        # Get current applying (applying date list for skip weekend, public holiday)
        applying = getAllApply(start_date, start_time, end_date, end_time, type, office, otherLeaveRef)
        # Get applied leave history
        applied = getAllLeave(racf, year, type, False)
        # Get public holiday list
        holiday = getPublicHolidays(office, start_date, end_date)

        # inital for other leave calculation, if it is not other leave, then return 0
        inclusive_workdays = 0

        ################################################### Basic checking ###################################################

        # check start date and end date cannot be weekends/ holidays (except speical leave with inclusive weekend)
        # Start
        if (not is_other_leave) or (is_other_leave and excluded_holidays):
            find = next((ph['Remark'] for ph in holiday if str(ph['Date']) == str(start_date.strftime('%Y-%m-%d')) and str(ph['Time']) == str(start_time)), "")
            if find != "":
                return ({"pass": False, "error_message" : "Leave applying start in Weekends / Holidays", "result": None, "Status_code": 502})
            # End
            find = next((ph['Remark'] for ph in holiday if str(ph['Date']) == str(end_date.strftime('%Y-%m-%d')) and str(ph['Time']) == str(end_time)), "")
            if find != "":
                return ({"pass": False, "error_message" : "Leave applying end in Weekends / Holidays", "result": None, "Status_code": 502})


        # Check applying date cannot be overlap
        # Applied
        for rec in applying:
            find = next((hist['applied_type'] for hist in applied if str(hist['applied_date']) == str(rec['applied_date']) and str(hist['applied_time']) == str(rec['applied_time'])), "")
            if find != "":
                return ({"pass": False, "error_message" : "Leave applying is overlapping", "result": None, "Status_code": 502})
        # Applying
        for rec in applying:
            find = next((app['applied_type'] for app in allApplying if str(app['applied_date']) == str(rec['applied_date']) and str(app['applied_time']) == str(rec['applied_time'])), "")
            if find != "":
                return ({"pass": False, "error_message" : "Leave applying is overlapping", "result": None, "Status_code": 502})
            
        # Check approver is active
        for i in range(1, 4):
            if getStaffRecord(racf)['staff'][f"approver{i}"] != "":
                approver = getStaffRecord(racf)['staff'][f"approver{i}"]
                approver_status = getStaffRecord(approver)['staff']["status"]
                if approver_status != "ACTIVE":
                    return ({"pass": False, "error_message" : "Invalid approver status, please contact HR for confirmation", "result": None,  "Status_code": 509})

        # Check summer hours period
        summer_hrs_failed = checkSummerHoursPeriod(office, year, psInput)
        if summer_hrs_failed and not super:
            return ({"pass": False, "error_message" : "During Summer Hours, any leave on Friday must be applied as ONE full day. Half-day (AM/PM) is not allowed", "result": None,  "Status_code": 514})


        ################################################### Entitlement Leave checking ###################################################

        if list(leaveTypes.find({'leave_type_id': type}))[0]['other_leave'] is False:

            # Check within period
            find = chkPeriod(datetime.strftime(start_date, "%Y-%m-%d"), datetime.strftime(end_date, "%Y-%m-%d"), year)
            if find['pass'] is False:
                return ({"pass": False, "error_message" : find['error_message'], "result": None, "Status_code": 502})

            # Get all applying
            allApplying += applying

            # Check entitlement is enough or not
            # Annual/ Casual leave
            if type == "LVE01" or type == "LVE02":
                entitled = (getYearEntitlement(year, getStaffRecord(racf), type) + getYearCarryForward(year, getStaffRecord(racf), type))
                allworkday = float(len(getAllLeave(racf, year, [type], False))) * 0.5 + float(len(allApplying)) * 0.5

                # print (year)
                # print (type)
                # print (getYearCarryForward(year, getStaffRecord(racf), type))

                if allworkday > entitled:
                    return ({"pass": False, "error_message" : "Not enough days left for the leave", "result": None,  "Status_code": 501})
            
            # Sick leave
            if type == "LVE04" or type == "LVE05":
                max_sl_days = list(leaveGroups.find({'groupID': list(leaveTypes.find({'leave_type_id': type}))[0]['leave_group']}))[0]['max_applied_days']
                #allslworkday = float(len(getAllLeave(racf, year, type, True))) * 0.5 + float(len(allApplying)) * 0.5
                allslworkday = float(len(getAllLeave(racf, year, ["LVE04", "LVE05"], False))) * 0.5 + float(len(allApplying)) * 0.5
                
                if allslworkday > max_sl_days:
                    # For India office, cannot be passed
                    if office == "DEL":
                        return ({"pass": False, "error_message" : "Your sick leave entitlement has been fully used. Please apply for Annual Leave / Casual Leave / No Pay Leave for further processing.", "result": None,  "Status_code": 513})
                    # For other office, just display the warning
                    # else:
                    #     warnings = "Reminder:  Total Full Paid Sick Leave taken has already reached 7 days which is the maximum cap of current leave calendar year (included below leave application)"
            
            # No pay leave
            if type == "LVE06":
                entitled = (getYearEntitlement(year, getStaffRecord(racf), "LVE01") + getYearCarryForward(year, getStaffRecord(racf), "LVE01"))
                allworkday = float(len(getAllLeave(racf, year, ["LVE01"], False))) * 0.5
                if (entitled - allworkday) > 0:
                    warnings = "Reminder: You have remaining annual leave entitlement for the current leave calendar year"

            # Check consecutive
            cons = checkConsecutive(racf, year, allApplying, type, office)
            if cons['consecutive'] and not super:
                return ({"pass": False, "error_message" : cons['error_message'], "result": None,  "Status_code": cons['Status_code']})
            


        ################################################### Other Leave checking ###################################################
        elif list(leaveTypes.find({'leave_type_id': type}))[0]['other_leave']:

            # parameter reading from user input
            o_start = otherLeaveRef[0]['period_start']
            o_end = otherLeaveRef[0]['period_end']

            accumulated = otherLeaveRef[0]['accumulated']
            excluded_holidays = otherLeaveRef[0]['excluded_holidays']

            # Get all applying
            allApplying += applying

            # Within allowed period
            for applyingrecord in allApplying:
                if (o_start <= applyingrecord['applied_date'] <= o_end) is False:
                    return ({"pass": False, "error_message" : "Leave applying is not within the allowed period", "result": None, "Status_code": 511})

            # Entitled Days checking
            entitled = otherLeaveRef[0]['entitled_days']
            allworkday = float(len(getAllLeave(racf, year, [type], False, otherRefNo))) * 0.5 + float(len(allApplying)) * 0.5
            if allworkday > entitled:
                return ({"pass": False, "error_message" : "Not enough days left for the leave", "result": None,  "Status_code": 501})

            # Forced consecutive leave
            if accumulated is False:
                if any(item.get('otherRefNo') == otherRefNo for item in getStaffRecord(racf)['leave_record']):
                    return ({"pass": False, "error_message" : "Not allow to apply multiple times", "result": None,  "Status_code": 512})
                if applying != allApplying:
                    return ({"pass": False, "error_message" : "Not allow to apply multiple times", "result": None,  "Status_code": 512})

            # Counting leave include holiday/ weekend
            if excluded_holidays is False:
                inclusive_workdays = getPhInclusiveWorkDays(racf, year, applying, type, office)
                acc_workdays += inclusive_workdays

                # check entitled days
                if acc_workdays > entitled:
                    return ({"pass": False, "error_message" : "Not enough days left for the leave", "result": None,  "Status_code": 501})

            # Check consecutive
            cons = checkConsecutive(racf, year, allApplying, type, office)
            if cons['consecutive'] and not super:
                return ({"pass": False, "error_message" : cons['error_message'], "result": None,  "Status_code": cons['Status_code']})


        # Get work Days/ out of office Days for display
        workday = float(len(applying)) * 0.5 if inclusive_workdays == 0 else inclusive_workdays
        oooday = getOOOdays(racf, year, applying, type, office)
        no_of_consective = oooday if type != "LVE04" and type != "LVE05" else float(cons['no_of_consective'])

        # Convert update list for MongoDB update
        if submit:
            # Approver List
            approvers = {}
            for i in range(1, 4):
                approvers[f"approver{i}"] = getStaffRecord(racf)['staff'][f"approver{i}"]
                approvers[f"approval_date{i}"] = ""

            periods = [ ]
            for rec in applying:
                period = {
                        'ldate': rec['applied_date'],
                        'ltime': rec['applied_time'],
                        }
                periods.append(period)
        
            detail = {
                        'start_date': datetime.strftime(start_date, "%Y-%m-%d"),
                        'start_time': start_time_os,
                        'end_date': datetime.strftime(end_date, "%Y-%m-%d"),
                        'end_time': end_time_os,
                        'no_of_workday': workday,
                        'no_of_calendarday': oooday,
                        'no_of_consective': no_of_consective,
                        'period': periods
                        }
            allDetails.append(detail)

    # Submit to update
    if allDetails != [ ]:

        ref_no = getNewRefNo(year, racf)

        # Update List
        updateLst = [{
                    'ref_no': ref_no,
                    'sharePointId': spid,
                    'otherRefNo': otherRefNo,
                    'year': year,
                    'type': type,
                    'applicationStatus': df['gcStatusPending'][0],
                    'approvalStatus': df['gcStatusPending1'][0], 
                    'submit_date': datetime.strftime(date.today(), "%Y-%m-%d"), 
                    'lastUpdate': racf, 
                    'updateDate': datetime.strftime(date.today(), "%Y-%m-%d"), 
                    'timeZone': timeZone,
                    'addCalendar': addCalendar,
                    'approval': approvers,
                    'details': allDetails
                    }]
        
        # History List
        id = getStaffRecord(racf)["_id"]
        if len(getStaffRecord(racf)["leave_record"]) != 0:
            updateRecord = [{"field" : "leave_record", "value" : getStaffRecord(racf)["leave_record"] + updateLst}]
        else:
            updateRecord = [{"field" : "leave_record", "value" : updateLst}]
        
        update = updateDB2(id, updateRecord)

        # Update addCalendar if there is change
        if addCalendar != getStaffRecord(racf)['staff']['addCalendar']:
            update = updateDB2(id, [{"field" : "staff.addCalendar", "value" : addCalendar}])


        if update['pass']:
            sendEmail(getStaffRecord(racf), ref_no, otherRefNo, df['gcActionApply'][0], df['gcActionApply'][0], 1, 1)
            

        return update


    # All pass not submission
    return ({"pass": True, "error_message" : "VALIDATION MODE.  Data pass validation.  Database NOT updated !", "result": [{"workday": workday, "calendarDay": oooday}], "Status_code": 200, "Warnings": warnings})


def listLeave (psInput):

    getLeaveTypes()
    psRacf = psInput.get("racf", "")
    psYear = psInput.get("year", 0)

    if len(psRacf) == 0 or psYear == 0:
        return ({"pass": False, "error_message" : "Incorrect parameters", "result": None, "Status_code": 505})
    displayLeaveHistoryDtl = [ ]
    displayLeaveHistoryHdr = [ ]
    
    staffRecord = getStaffRecord(psRacf)
    if not isinstance(staffRecord, dict):
        return ({"pass": False, "error_message" : "Staff Record Not Exist", "result": None, "Status_code": 504}) 

    for lveType in leaveTypeLst:
        leaveTypeHdr = {
            "leaveType": lveType.get("leave_type"),
            "leaveTypeId": lveType.get("leave_type_id"),
            "taken" : countLeave(psYear, lveType.get("leave_type_id"), df['gcStatusApproved'][0], staffRecord),
            "pending": countLeave(psYear, lveType.get("leave_type_id"), df['gcStatusPending'][0], staffRecord),
            "balance": checkBalance(psYear, lveType, staffRecord, [])
        }
        displayLeaveHistoryHdr.append(leaveTypeHdr)    
        
    leaveHistoryLst = getLeaveHistory(psYear, psYear, staffRecord)
    currRefNo = 0
    currStartDate = ""
    currStartTime = ""
    for lve in leaveHistoryLst:
        if (currRefNo == lve["ref_no"] and currStartDate != lve["startDate"] and currStartTime != lve["startTime"]) or (currRefNo != lve["ref_no"]):
            displayLeaveRecord = {
                    "submitDate":  getMMDDYYYY(lve["submitDate"]),
                    "refNo": getDisplayRefNo(lve["ref_no"], lve["office"], lve["racf"]) if lve["other_leave"] is False else lve["ref_no"],
                    "office": lve["office"],
                    "staffname": lve["staffname"],
                    "empID": lve["empID"],
                    "dept": lve["dept"],
                    "position": lve["position"],
                    "type_id": lve["type"],
                    "sharePointId": lve["sharePointId"],
                    "type" : list(filter(lambda r: (r["leave_type_id"].upper() == lve["type"]), leaveTypeLst))[0].get("leave_type"),
                    "year": getDisplayLeaveYear(lve["year"]) if lve["other_leave"] is False else "",
                    "leaveFrom": getMMDDYYYY(lve["startDate"]),
                    "startPeriod": lve["startTime"],
                    "leaveTo": getMMDDYYYY(lve["endDate"]),
                    "endPeriod": lve["endTime"],
                    "workday": lve["workDay"],
                    "calendarDay": lve["calendarDay"],
                    "applicationStatus": lve["applicationStatus"],
                    "approver1": lve["approver1"],
                    "approver2": lve["approver2"],
                    "approver3": lve["approver3"],
                    "approvalStatus" : lve["approvalStatus"],
                    "lastUpdate": lve["lastUpdate"],
                    "updateDate": getMMDDYYYY(lve["updateDate"])
            }
            displayLeaveHistoryDtl.append(displayLeaveRecord)
            currRefNo = lve["ref_no"]
            currStartDate = lve["startDate"]
            currStartTime = lve["endDate"]

    fullRecord = {
        "header":  displayLeaveHistoryHdr,
        "details": displayLeaveHistoryDtl
    }
    leaveRecordLst = [ ]
    leaveRecordLst.append (fullRecord) 

    if len(leaveRecordLst) == 0:
        return ({"pass": True, "error_message" : None, "result": [], "Status_code": 200}) 
    else:
        return ({"pass": True, "error_message" : None, "result": leaveRecordLst, "Status_code": 200}) 


def listApprove(psInput):
  

    getLeaveTypes()
    psApprover = psInput.get("racf", "")
    #print ('listApprove',  psApprover)

    try:
        if len(psApprover) == 0:
            return ({"pass": False, "error_message" : "Incorrect parameters", "result": None, "Status_code": 505})
    
    except:
        return ({"pass": False, "error_message" : "Incorrect parameters or waiting response ... ", "result": None, "Status_code": 505})

    approvalRecordLst = [ ]

    i = 1
    while i <= 6:
        if i == 1:
            approver = "staff.approver1"
            pendingStatus = df['gcStatusPending1'][0]
        elif i == 2:
            approver = "staff.approver2"
            pendingStatus = df['gcStatusPending2'][0]
        elif i == 3:
            approver = "staff.approver3"
            pendingStatus = df['gcStatusPending3'][0]
        elif i == 4:
            approver = "staff.approver1"
            pendingStatus = df['gcStatusPendingCancel1'][0]
        elif i == 5:
            approver = "staff.approver2"
            pendingStatus = df['gcStatusPendingCancel2'][0]
        else:
            approver = "staff.approver3"
            pendingStatus = df['gcStatusPendingCancel3'][0]
        #tmpApproverLst = [ ]
        staffRecord = list(eleaveDtl.find ( {approver : { '$regex' : psApprover, '$options' : "i"} , "staff.status": { '$regex': "ACTIVE", '$options': "i"} } ) )
        for rec in staffRecord:
            staff = rec["staff"]["name"]
            racf = rec["staff"]["racf"]
            office = rec["staff"]["hr_office"]
            pendingLst = list(filter(lambda r: (r["approvalStatus"] == pendingStatus), rec["leave_record"]))
            for record in pendingLst:
                leaveDetailsLst = [ ]
                for details in record["details"]:
                    # get rows for each leave application
                    leaveDetails = {
                            "startDate": details["start_date"],
                            "startTime": details["start_time"],
                            "endDate": details["end_date"],
                            "endTime": details["end_time"],
                            "workday": details["no_of_workday"],
                            "calendarDay": details["no_of_calendarday"]
                    }
                    leaveDetailsLst.append(leaveDetails)
                # put a single leave application , by ref_no, into a dict

                # Make all approved and pending list each staff record
                balance = "Nil"
                applicationList = list(filter(lambda r: r["approvalStatus"] == "APPROVED" and r["year"] == record["year"] and r["type"] == record["type"], rec["leave_record"]))
                numberOfTaken = sum(detail.get('no_of_workday', 0) for leave_application in applicationList for detail in leave_application.get('details', []))
                try:
                    if record["type"] == "LVE01":
                        balance = (getYearEntitlement(record["year"], getStaffRecord(racf), record["type"]) + getYearCarryForward(record["year"], getStaffRecord(racf), record["type"])) - numberOfTaken
                    else:
                        balance = (getYearEntitlement(record["year"], getStaffRecord(racf), record["type"])) - numberOfTaken
                except:
                    balance = "Nil"

                # Get Other ref number if it exists:
                try:
                    otherRefNo = record["otherRefNo"]
                except:
                    otherRefNo = ""

                leaveRecord = {
                    "staff": staff,
                    "racf": racf,
                    "ref_no": getDisplayRefNo(record["ref_no"], office, racf),
                    "otherRefNo": otherRefNo,
                    "sharePointId": record["sharePointId"],
                    "type_id": record["type"],
                    "type": list(filter(lambda r: (r["leave_type_id"].upper() == record["type"]), leaveTypeLst))[0].get("leave_type"),
                    "balance": balance,
                    "approvalStatus": pendingStatus,
                    "details": leaveDetailsLst
                }
                # add that leave application into the whole list (temporary list)
                approvalRecordLst.append(leaveRecord)
        i += 1

        approvalRecordLst = sorted(approvalRecordLst, key=lambda d: (d["approvalStatus"], d["staff"], d["racf"], d["ref_no"]))
    return ({"pass": True, "error_message" : None, "result": approvalRecordLst, "Status_code": 200}) 

def changeStatus(psInput):

    psRefNo = psInput.get("refNo", 0)
    psRacf = psInput.get("racf", "")
    psAction = psInput.get("action","")

    approval_index = 1
    approver = ""
        ## Super User get session data
    try:
        psSuperUser = session["superUser"]
    except:
        psSuperUser = psInput.get("superUser", False)

    if psRefNo == 0 or len(psRacf) == 0 or len(psAction) == 0:
        return ({"pass": False, "error_message" : "Incorrect parameters", "result": None, "Status_code": 505})        
    
    refNo = getActualRefNo(psRefNo)
    # applicantRacf = "NF" + psRefNo[-4:]
    applicantRacf = "NF1" + psRefNo[-3:]
    staffRecord = getStaffRecord(applicantRacf)

    if not isinstance(staffRecord, dict):
        return ({"pass": False, "error_message" : "Staff Record Not Exist", "result": None, "Status_code": 504}) 

    max_approver = 1
    ## Find final approver
    if len(staffRecord["staff"]["approver2"]) != 0: max_approver = 2
    if len(staffRecord["staff"]["approver3"]) != 0: max_approver = 3

    leaveRecord = [(idx, record) for idx, record in enumerate(staffRecord["leave_record"]) if record["ref_no"] == refNo]

    currApplicationStatus = (leaveRecord[0][1]["applicationStatus"])
    currApprovalStatus = (leaveRecord[0][1]["approvalStatus"])

    ## Get work days
    no_of_workdays = (leaveRecord[0][1]["details"][0]["no_of_workday"])

    ## Get other ref no
    otherRefNo = leaveRecord[0][1]["otherRefNo"]

    index = leaveRecord[0][0]
    newApplicationStatus = currApplicationStatus
    if currApplicationStatus == df['gcStatusReject'][0] or currApplicationStatus == df['gcStatusCancel'][0]:
        return ({"pass": False, "error_message" : "Current Status cannot be changed.", "result": None, "Status_code": 603})

    # Action - Cancel
    if psAction == df['gcActionCancel'][0]:
        if psRacf != applicantRacf and not psSuperUser:
           return ({"pass": False, "error_message" : "Only applicant himself / herself can cancel leave.", "result": None, "Status_code": 604})
        if currApprovalStatus == df['gcStatusPendingCancel1'][0] or currApprovalStatus == df['gcStatusPendingCancel2'][0] or currApprovalStatus == df['gcStatusPendingCancel3'][0]:
            return ({"pass": False, "error_message" : "Leave cancel already submitted and waiting for approval.", "result": None, "Status_code": 605})
        firstLeaveDate = str2Date ("9999-12-31")
        for row in leaveRecord[0][1]["details"]:
            if str2Date(row["start_date"]) < firstLeaveDate:
                firstLeaveDate = str2Date(row["start_date"])
        if datetime.today() > firstLeaveDate and not psSuperUser:
            return ({"pass": False, "error_message" : "Cannot cancel leave in the past period.", "result": None, "Status_code": 606})
        newApplicationStatus = df['gcStatusPending'][0]
        newApprovalStatus = df['gcStatusPendingCancel1'][0]
        updateBy = psRacf
        approver = staffRecord["staff"]["approver1"]
        emailRequest = df['gcActionCancel'][0]

    # Action - Approve     
    elif psAction == df['gcActionApprove'][0]:
        updateBy = psRacf
        if currApprovalStatus == df['gcStatusPending1'][0]:
            approval_index = 1
            approver = psRacf
            if psRacf != staffRecord["staff"]["approver1"] and not psSuperUser:             
               return ({"pass": False, "error_message" : "User is not the first approver.", "result": None, "Status_code": 607}) 
            emailRequest = df['gcActionApply'][0]
            #***
            if len(staffRecord["staff"]["approver2"]) == 0:
                newApplicationStatus = df['gcStatusApproved'][0]
                newApprovalStatus = df['gcStatusApproved'][0]
            else:
                newApprovalStatus = df['gcStatusPending2'][0]
                                
        elif currApprovalStatus == df['gcStatusPending2'][0]:
            approval_index = 2
            approver = psRacf
            if psRacf != staffRecord["staff"]["approver2"] and not psSuperUser:
               return ({"pass": False, "error_message" : "User is not the second approver.", "result": None, "Status_code": 608}) 
            emailRequest = df['gcActionApply'][0]
            if len(staffRecord["staff"]["approver3"]) == 0:
                newApplicationStatus = df['gcStatusApproved'][0]
                newApprovalStatus = df['gcStatusApproved'][0]
            else:
                newApprovalStatus = df['gcStatusPending3'][0]
        elif currApprovalStatus == df['gcStatusPending3'][0]:
            approval_index = 3
            approver = psRacf
            if psRacf != staffRecord["staff"]["approver3"] and not psSuperUser:
               return ({"pass": False, "error_message" : "User is not the third approver.", "result": None, "Status_code": 609}) 
            emailRequest = df['gcActionApply'][0]
            newApplicationStatus = df['gcStatusApproved'][0]
            newApprovalStatus = df['gcStatusApproved'][0]
        elif currApprovalStatus == df['gcStatusPendingCancel1'][0]:
            approval_index = 1
            approver = psRacf
            if psRacf != staffRecord["staff"]["approver1"] and not psSuperUser:
               return ({"pass": False, "error_message" : "User is not the first approver.", "result": None, "Status_code": 607})
            emailRequest = df['gcActionCancel'][0]        
            if len(staffRecord["staff"]["approver2"]) == 0:
                newApplicationStatus = df['gcStatusCancel'][0]
                newApprovalStatus = df['gcStatusCancel'][0]
            else:
                newApprovalStatus = df['gcStatusPendingCancel2'][0]
        elif currApprovalStatus == df['gcStatusPendingCancel2'][0]:
            approval_index = 2
            approver = psRacf
            if psRacf != staffRecord["staff"]["approver2"] and not psSuperUser:
               return ({"pass": False, "error_message" : "User is not the second approver.", "result": None, "Status_code": 608})   
            emailRequest = df['gcActionCancel'][0]
            if len(staffRecord["staff"]["approver3"]) == 0:
                newApplicationStatus = df['gcStatusCancel'][0]
                newApprovalStatus = df['gcStatusCancel'][0]
            else:
                newApprovalStatus = df['gcStatusPendingCancel3'][0]
        elif currApprovalStatus == df['gcStatusPendingCancel3'][0]:
            approval_index = 3
            approver = psRacf
            if psRacf != staffRecord["staff"]["approver3"] and not psSuperUser:
               return ({"pass": False, "error_message" : "User is not the third approver.", "result": None, "Status_code": 609}) 
            emailRequest = df['gcActionCancel'][0]
            newApplicationStatus = df['gcStatusCancel'][0]
            newApprovalStatus = df['gcStatusCancel'][0]
        elif currApprovalStatus == df['gcStatusApproved'][0]:
            return ({"pass": False, "error_message" : "Leave already approved.", "result": None, "Status_code": 610}) 
        else:
            return ({"pass": False, "error_message" : "Incorrect action", "result": None, "Status_code": 611}) 
    #Action = Reject
    else:
        updateBy = psRacf 
        if currApprovalStatus == df['gcStatusPending1'][0]:
            approval_index = 1
            approver = psRacf
            if psRacf != staffRecord["staff"]["approver1"] and not psSuperUser:
                return ({"pass": False, "error_message" : "User is not the first approver.", "result": None, "Status_code": 607})
            emailRequest = df['gcActionApply'][0]
            newApplicationStatus = df['gcStatusReject'][0]
            newApprovalStatus = df['gcStatusReject'][0]
        elif currApprovalStatus == df['gcStatusPending2'][0]:
            approval_index = 2
            approver = psRacf
            if psRacf != staffRecord["staff"]["approver2"] and not psSuperUser:
                return ({"pass": False, "error_message" : "User is not the second approver.", "result": None, "Status_code": 608})
            emailRequest = df['gcActionApply'][0]
            newApplicationStatus = df['gcStatusReject'][0]
            newApprovalStatus = df['gcStatusReject'][0]
        elif currApprovalStatus == df['gcStatusPending3'][0]:
            approval_index = 3
            approver = psRacf
            if psRacf != staffRecord["staff"]["approver3"] and not psSuperUser:
                return ({"pass": False, "error_message" : "User is not the third approver.", "result": None, "Status_code": 609})
            emailRequest = df['gcActionApply'][0]
            newApplicationStatus = df['gcStatusReject'][0]
            newApprovalStatus = df['gcStatusReject'][0]      
        elif currApprovalStatus == df['gcStatusPendingCancel1'][0]:
            approval_index = 1
            approver = psRacf
            if psRacf != staffRecord["staff"]["approver1"] and not psSuperUser:
                return ({"pass": False, "error_message" : "User is not the first approver.", "result": None, "Status_code": 607})
            emailRequest = df['gcActionCancel'][0]
            newApplicationStatus = df['gcStatusApproved'][0]
            newApprovalStatus = df['gcStatusApproved'][0]
        elif currApprovalStatus == df['gcStatusPendingCancel2'][0]:
            approval_index = 2
            approver = psRacf
            if psRacf != staffRecord["staff"]["approver2"] and not psSuperUser:
                return ({"pass": False, "error_message" : "User is not the second approver.", "result": None, "Status_code": 608})
            emailRequest = df['gcActionCancel'][0]
            newApplicationStatus = df['gcStatusApproved'][0]
            newApprovalStatus = df['gcStatusApproved'][0]
        elif currApprovalStatus == df['gcStatusPendingCancel3'][0]:
            approval_index = 3
            approver = psRacf
            if psRacf != staffRecord["staff"]["approver3"] and not psSuperUser:
                return ({"pass": False, "error_message" : "User is not the third approver.", "result": None, "Status_code": 609})
            emailRequest = df['gcActionCancel'][0]
            newApplicationStatus = df['gcStatusApproved'][0]
            newApprovalStatus = df['gcStatusApproved'][0]

    id = staffRecord["_id"]
    updateStatusLst = [ ]
    updateStatus = {
        "field": "leave_record." + str(index) + ".applicationStatus",
        "value": newApplicationStatus
    }
    updateStatusLst.append(updateStatus)     
    updateStatus = {
        "field": "leave_record." + str(index) + ".approvalStatus",
        "value": newApprovalStatus
    }
    updateStatusLst.append(updateStatus)
    updateStatus = {
        "field": "leave_record." + str(index) + ".lastUpdate",
        "value": updateBy
    }
    updateStatusLst.append(updateStatus)
    updateStatus = {
        "field": "leave_record." + str(index) + ".updateDate",
        "value": date2Str(date.today())
    }
    updateStatusLst.append(updateStatus)
    # Update leave record data in MongoDB
    result = updateDB2(id, updateStatusLst)

    updateApproval = {
        "field": "leave_record." + str(index) + ".approval" + ".approver" + str(approval_index),
        "value": approver
    }

    updateApprovalLst = [ ]
    updateApprovalLst.append(updateApproval)

    # Get local time from browser and convert to MongoDB Datetime format YYYY-mm-dd
    date_input = datetime.strptime(psInput['localTime'], '%a %b %d %Y')
    new_date = date_input.strftime('%Y-%m-%d')

    # Action Cancel to clean up all approval date
    #if psAction != df['gcActionCancel'][0]:
    updateApproval = {
        "field": "leave_record." + str(index) + ".approval" + ".approval_date" + str(approval_index),
        "value":  new_date
            }
    updateApprovalLst.append(updateApproval)
    #elif psAction == df['gcActionCancel'][0]:
    #    for m in range(1 , 4):
    #        updateApproval = {
    #            "field": "leave_record." + str(index) + ".approval" + ".approval_date" + str(m),
    #            "value":  ""
    #            }
    #        updateApprovalLst.append(updateApproval)
    #        updateApproval = {
    #            "field": "leave_record." + str(index) + ".approval" + ".approver1",
    #            "value": staffRecord["staff"]["approver1"]
    #            }
    #        updateApproval = {
    #            "field": "leave_record." + str(index) + ".approval" + ".approver2",
    #            "value": staffRecord["staff"]["approver2"]
    #            }
    #        updateApproval = {
    #            "field": "leave_record." + str(index) + ".approval" + ".approver3",
    #            "value": staffRecord["staff"]["approver3"]
    #            }

    # Update approver record data in MongoDB
    approval_result = updateDB2(id, updateApprovalLst)

    
    if result.get("pass") and approval_result.get("pass"):
        staffRecord = getStaffRecord(applicantRacf)
        sendEmail (staffRecord, refNo, otherRefNo, psAction, emailRequest, max_approver, approval_index)
    return (result)  

 
def listApprovedLeaveByYear(psInput):
    
    getLeaveTypes()
    psUser = psInput.get("racf", "")

    # Get Super User 
    try:
        superUser = session["superUser"]
    except:
        superUser = False
    
    if superUser:
        superUser = psInput.get("superUser")
    else:
        superUser = False

    # Get Year
    try:
        years = (json.loads(current_app.config['YEARS'])).get('year')
    except:
        years_str = os.environ['YEARS']      
        years = eval(years_str)
        years = pd.DataFrame(data=years)
        years = years['year'].tolist()

    if len(psUser) == 0:
        return ({"pass": False, "error_message" : "Incorrect parameters", "result": None, "Status_code": 505})

    approvalRecordLst = [ ]


    #tmpApproverLst = [ ]
    staffRecord = list(eleaveDtl.find ( {"staff.racf" : { '$regex' : psUser, '$options' : "i"} , "staff.status": { '$regex': "ACTIVE", '$options': "i"} } ) )
    for rec in staffRecord:
        staff = rec["staff"]["name"]
        racf = rec["staff"]["racf"]
        office = rec["staff"]["hr_office"]
        leaveappliedLst = list(filter(lambda r: (r["approvalStatus"] == df['gcStatusApproved'][0]), rec["leave_record"]))
        for record in leaveappliedLst:
            #print (record)
            leaveDetailsLst = [ ]
            ##if record['year'] == psYear:
            ## Added by Vincent to allow multiple years 
            if record['year'] in years:
                showLeave = True
                for details in record["details"]:
                    #print (details)
                    # get rows for each leave application
                    leaveDetails = {
                            "startDate": details["start_date"],
                            "startTime": details["start_time"],
                            "endDate": details["end_date"],
                            "endTime": details["end_time"],
                            "workday": details["no_of_workday"],
                            "calendarDay": details["no_of_calendarday"]
                                   }
                    #Only show the the leave date after today date if not super user
                    # Get local time from browser and convert to MongoDB Datetime format YYYY-mm-dd
                    date_input = datetime.strptime(psInput['localTime'], '%a %b %d %Y')

                    if (((datetime.strptime(details["start_date"],'%Y-%m-%d').date() - date_input.date()).days) < 1):
                        showLeave = False

                    if (((datetime.strptime(details["start_date"],'%Y-%m-%d').date() - date_input.date()).days) >= 1 and not superUser) or (superUser):
                        leaveDetailsLst.append(leaveDetails)
                    # put a single leave application , by ref_no, into a dict
                leaveRecord = {
                        "staff": staff,
                        "racf": racf,
                        "ref_no": getDisplayRefNo(record["ref_no"], office, racf),
                        "otherRefNo": record["otherRefNo"] if record["otherRefNo"] != "" else "",
                        "type_id": record["type"],
                        "type": list(filter(lambda r: (r["leave_type_id"].upper() == record["type"]), leaveTypeLst))[0].get("leave_type"),
                        "approvalStatus": df['gcStatusApproved'][0],
                        "details": leaveDetailsLst
                              }
                    # add that leave application into the whole list (temporary list)
                if (((datetime.strptime(details["start_date"],'%Y-%m-%d').date() - date_input.date()).days) >= 1 and not superUser and showLeave) or (superUser):
                    approvalRecordLst.append(leaveRecord)
        approvalRecordLst = sorted(approvalRecordLst, key=lambda d: (d["approvalStatus"], d["staff"], d["racf"], d["ref_no"]))

    return ({"pass": True, "error_message" : None, "result": approvalRecordLst, "Status_code": 200}) 

#@app.route("/api/listleave", methods=['POST'])
#@app.route("/")
def apiListLeave():
    psInput = request.get_json()
    result = listLeave(psInput)
    try: 
        return jsonify(result), result['Status_code'] # APP
    except:
        return jsonify(result) # postman


# Status_code 200: passed
# Status_code 501: Fail to generate Leave Summary

@eleave.route("/api/printsummary", methods=['POST'])
@checkLogged.check_logged
def apiPrintSummary():
 
    para = json.loads(request.headers['parameters'])                        
    psInput =  {'year': para['year'], 'racf': para['racf']}    

    result = listLeave(psInput)
 
    rpt = reportMap.find_one ( { "report": "Leave Summary"} )
    
    if (result.get("pass")): 
        hdr = result.get("result")[0]["header"]
        alTaken = 0
        alPending = 0
        alBalance = 0
        clTaken = 0
        clPending = 0
        clBalance = 0
        slTaken = 0
        slPending = 0
        slBalance = 0
        for lve in leaveTypeLst:
            hdrData =  list(filter(lambda r: (r["leaveTypeId"].upper() == lve["leave_type_id"]), hdr))[0]
            if lve["leave_type_id"] == "LVE01":
                alTaken = hdrData["taken"]
                alPending = hdrData["pending"]
                alBalance = hdrData["balance"]
            elif lve["leave_type_id"] == "LVE02":
                clTaken = hdrData["taken"]
                clPending = hdrData["pending"]
                clBalance = hdrData["balance"]
            elif lve["leave_type_id"] == "LVE04":
                slTaken = hdrData["taken"] + slTaken
                slPending = hdrData["pending"] + slPending
                slBalance = hdrData["balance"] + slBalance
            elif lve["leave_type_id"] == "LVE05":
                slTaken = hdrData["taken"] + slTaken
                slPending = hdrData["pending"] + slPending
                slBalance = hdrData["balance"] + slBalance
        rptDtlLst = [ ]
        for record in result.get("result"):
            for dtl in record["details"]:
                rptDtl = {
                    "submitDate": dtl.get("submitDate"),
                    "refNo": dtl.get("refNo"),
                    "office": dtl.get("office"),
                    "staffname": dtl.get("staffname"),
                    "dept": dtl.get("dept"),
                    "type": dtl.get("type"),
                    "year": dtl.get("year"),
                    "leaveFrom": dtl.get("leaveFrom"),
                    "startPeriod": dtl.get("startPeriod"),
                    "leaveTo": dtl.get("leaveTo"),
                    "endPeriod": dtl.get("endPeriod"),
                    "workday": dtl.get("workday"),
                    "calendarDay": dtl.get("calendarDay"),
                    "applicationStatus": dtl.get("applicationStatus"),
                    "lastUpdate": dtl.get("lastUpdate"),
                    "updateDate": dtl.get("updateDate")        
                }
                rptDtlLst.append(rptDtl)

            # Sort by leave start requested by Sandy 1/26/2023
            rptDtlLst.sort(key=lambda x: x.get('leaveFrom'))
            # added by Vincent Cheng on 11/23 
            try:
                if record["details"][0].get("year"):
                    pass
            except:            
                return jsonify({"error_message" : "Sorry, we failed to generate Leave Summary.  Perhaps no data for the year"}), 501   
    
            report = {
                "hdrCalendarYear": getDisplayLeaveYear(psInput["year"]),
                "hdrUser": record["details"][0].get("staffname") + "\nLeave Application Summary",
                "hdrALTaken": alTaken,
                "hdrALPending": alPending,
                "hdrALBalance": alBalance,
                "hdrSLTaken": slTaken,
                "hdrSLPending": slPending,
                "hdrCLTaken": clTaken,
                "hdrCLPending": clPending,
                "hdrCLBalance": clBalance,
                "dtl": rptDtlLst
                
            }

        #filename when using in Heroku:
        fs = gridfs.GridFS(db)
        wb = load_workbook(filename=BytesIO(fs.get(ObjectId(rpt["file"]["fileObj"])).read()))

        # filename in development:
        #wb = load_workbook(filename=rpt["file"]["fileName"])
        ws = wb[rpt["file"]["wsName"]]
        
        result = genReport(ws, report, rpt)
        if result[1] == 0:
            out = BytesIO()
            wb.save(out)
            out.seek(0)        
 
            #wb.save(filename="F:\mmgapp\dev\eleave\output\LeaveSummary.xlsx")
            wb.close()            
            print('sending file...')
            return send_file(out,  download_name='Summary_Report.xls', mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', as_attachment=True)      
        else:            
            return jsonify({"error_message" : "Sorry, we failed to generate Leave Summary.  Perhaps no data for the year"}), 501    
        

@eleave.route("/api/printapply", methods=['POST'])
@checkLogged.check_logged
def apiPrintApply():

    # Get RACF and Ref_no from frontend
    para = json.loads(request.headers['parameters'])
    ref = para['ref']
    racf = para['racf']

    # Get Staff record for output
    StaffRecord = getStaffRecord(racf)

    # Find leave list by racf and ref
    ref_no = ref.replace(StaffRecord['staff']['hr_office'],"") # remove hr office in reference no
    ref_no = ref_no.replace(StaffRecord['staff']['racf'][-3:],"") # remove staff racf in reference no

    # Get leave balance 
    getLeaveTypes()

    approvalRecordLst = [ ]

    for rec in StaffRecord['leave_record']:

        #Select exact application by reference number
        if rec['ref_no'] == int(ref_no):
            leaveDetailsLst = [ ]
            for details in rec["details"]:
                # get rows for each leave application
                leaveDetails = {
                    "startDate": details["start_date"],
                    "startTime": details["start_time"],
                    "endDate": details["end_date"],
                    "endTime": details["end_time"],
                    "workday": details["no_of_workday"],
                    "calendarDay": details["no_of_calendarday"]
                    }
                leaveDetailsLst.append(leaveDetails)

                # Check the balance from Thomas function
                displayLeaveHistoryHdr = [ ]
                for lveType in leaveTypeLst:
                    leaveTypeHdr = {
                                    "leaveType": lveType.get("leave_type"),
                                    "leaveTypeId": lveType.get("leave_type_id"),
                                    "taken" : countLeave(rec['year'], lveType.get("leave_type_id"), df['gcStatusApproved'][0], StaffRecord),
                                    "pending": countLeave(rec['year'], lveType.get("leave_type_id"), df['gcStatusPending'][0], StaffRecord),
                                    "balance": checkBalance(rec['year'], lveType, StaffRecord, [])
                                    }
                    displayLeaveHistoryHdr.append(leaveTypeHdr)
        

            # Summarize the number of balance
            # print (displayLeaveHistoryHdr)
            if rec['type'] == 'LVE01':
                DaysOfApproved = displayLeaveHistoryHdr[0]['taken']
                DaysOfPending = displayLeaveHistoryHdr[0]['pending']
                DaysOfleft = displayLeaveHistoryHdr[0]['balance']
                DaysOfCarryForward = getYearCarryForward(rec['year'], StaffRecord, rec['type'])
                DaysOfEntitlement = str(DaysOfCarryForward) + " (" + str(int(rec['year']-1)) + ") " + "+ " + str(getYearEntitlement(rec['year'], StaffRecord, rec['type'])) + " (" + str(int(rec['year'])) + ") "
            elif rec['type'] == 'LVE02':
                DaysOfApproved = displayLeaveHistoryHdr[1]['taken']
                DaysOfPending = displayLeaveHistoryHdr[1]['pending']
                DaysOfleft = displayLeaveHistoryHdr[1]['balance']
                DaysOfCarryForward = 0
                DaysOfEntitlement = str(getYearEntitlement(rec['year'], StaffRecord, rec['type'])) + " (" + str(int(rec['year'])) + ") "
            elif rec['type'] == 'LVE03':
                DaysOfApproved = displayLeaveHistoryHdr[4]['taken']
                DaysOfPending = displayLeaveHistoryHdr[4]['pending']
                DaysOfleft = "N/A"
                DaysOfCarryForward = 0
                DaysOfEntitlement = "N/A"
            elif rec['type'] == 'LVE04':
                DaysOfApproved = displayLeaveHistoryHdr[3]['taken']
                DaysOfPending = displayLeaveHistoryHdr[3]['pending']
                DaysOfleft = "N/A"
                DaysOfCarryForward = 0
                DaysOfEntitlement = "N/A"
            elif rec['type'] == 'LVE05':
                DaysOfApproved = displayLeaveHistoryHdr[2]['taken']
                DaysOfPending = displayLeaveHistoryHdr[2]['pending']
                DaysOfleft = "N/A"
                DaysOfEntitlement = "N/A"
            else:
                DaysOfApproved = displayLeaveHistoryHdr[2]['taken']
                DaysOfPending = displayLeaveHistoryHdr[2]['pending']
                DaysOfleft = "N/A"
                DaysOfEntitlement = "N/A"


            get_approver1 = ""
            get_pos_approver1 = ""
            get_approver2 = ""
            get_pos_approver2 = ""
            get_approver3 = ""
            get_pos_approver3 = ""

            if len(str(rec['approval']['approver1'])) > 0:
                get_approver1 = getStaffRecord(rec['approval']['approver1'])['staff']['name']
                get_pos_approver1 = getStaffRecord(rec['approval']['approver1'])['staff']['position']

            if len(str(rec['approval']['approver2'])) > 0:
                get_approver2 = getStaffRecord(rec['approval']['approver2'])['staff']['name']
                get_pos_approver2 = getStaffRecord(rec['approval']['approver2'])['staff']['position']

            if len(str(rec['approval']['approver3'])) > 0:
                get_approver3 = getStaffRecord(rec['approval']['approver3'])['staff']['name']
                get_pos_approver3 = getStaffRecord(rec['approval']['approver3'])['staff']['position']

            if rec['sharePointId'] == "":
                DissharePointid = ""
            else:
                DissharePointid = "(" + str(rec['sharePointId']) + ")"

            try:
                TakenApproved = float(DaysOfApproved + DaysOfPending)
            except:
                TakenApproved = "NA"

            # Go back to build the structure for excel output file
            # Array item label must be the same as MongoDB cell field in fileDrectory
            leaveRecord = {
                "staff": StaffRecord['staff']['name'],
                "racf": racf,
                "position": StaffRecord['staff']['position'],
                "dept": StaffRecord['staff']['dept'],
                "date_joined": StaffRecord['staff']['date_join'],
                "ref_no": ref if rec['otherRefNo'] == "" else rec['otherRefNo'],
                "sharePointid": DissharePointid,
                "approver1": get_approver1,
                "approver_pos1": get_pos_approver1,
                "approval_date1": rec['approval']['approval_date1'],
                "approver2": get_approver2,
                "approver_pos2": get_pos_approver2,
                "approval_date2": rec['approval']['approval_date2'],
                "approver3": get_approver3,
                "approver_pos3": get_pos_approver3,
                "approval_date3": rec['approval']['approval_date3'],
                "NoDaysEntitlement": DaysOfEntitlement ,
                "NoDaysTakenApproved": str(TakenApproved) + " (" + str(float(DaysOfApproved)) + " + "+ str(float(DaysOfPending)) + ") ",
                "NoDaysLeft": DaysOfleft,
                "type_id": rec["type"],
                "leaveTypeBalance": list(filter(lambda r: (r["leave_type_id"].upper() == rec["type"]), leaveTypeLst))[0].get("leave_type") + " BALANCE",
                "type": list(filter(lambda r: (r["leave_type_id"].upper() == rec["type"]), leaveTypeLst))[0].get("leave_type"),
                "calendarYear": getDisplayLeaveYear(rec["year"]),
                "submit_date": rec['submit_date'],
                "details": leaveDetailsLst
                }
            
            #Output to array to excel file
            approvalRecordLst.append(leaveRecord)

    # Get mapping from MongoDB
    rpt = reportMap.find_one ( { "report": "Application Form"} )

    #filename when using in Heroku:
    fs = gridfs.GridFS(db)
    wb = load_workbook(filename=BytesIO(fs.get(ObjectId(rpt["file"]["fileObj"])).read()))
    ws = wb[rpt["file"]["wsName"]]

    try:
        genApplyForm(ws, approvalRecordLst, rpt)
    except Exception as e:
        print (e)
        return jsonify({"error_message" : "Sorry, we failed to generate Application form"}), 501    

    # Output 
    out = BytesIO()
    wb.save(out)
    out.seek(0)

    wb.close()            
    print('sending file...')

    return send_file(
        out,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,          # Required to trigger download/"Save As"
        download_name='a_file.xlsx'  # Use .xlsx extension for modern Excel files
    )


def listPartnersLeave(psInput):

    getLeaveTypes()

    psUser = psInput.get("racf", "")

    # Get Year
    try:
        years = (json.loads(current_app.config['YEARS'])).get('year')
    except:
        years_str = os.environ['YEARS']      
        years = eval(years_str)
        years = pd.DataFrame(data=years)
        years = years['year'].tolist()

    # Get date
    date_input = datetime.strptime(psInput['localTime'], '%a %b %d %Y')

    staffRecord = getStaffRecord(psUser)
    if not isinstance(staffRecord, dict):
        return ({"pass": False, "error_message" : "Staff Record Not Exist", "result": None, "Status_code": 504}) 

    if len(psUser) == 0:
        return ({"pass": False, "error_message" : "Incorrect parameters", "result": None, "Status_code": 505})

    # Precompute constants
    date_input_date = date_input.date()
    date_input_str = date_input.strftime('%Y-%m-%d')
    end_date = date_input + timedelta(days=13)
    end_date_str = end_date.strftime('%Y-%m-%d')

    approved = df['gcStatusApproved'][0]
    pending = df['gcStatusPending'][0]
    leaveTypeDict = {r["leave_type_id"].upper(): r["leave_type"] for r in leaveTypeLst}

    partnersLeave = []
    staffRecord = list(eleaveDtl.find({"staff.racf": {'$regex': psUser, '$options': "i"}, "staff.status": {'$regex': "ACTIVE", '$options': "i"}}))
    partners_set = set()
    for rec in staffRecord:
        if len(rec["staff"]["partners"]) > 0:
            partnerslist = str(rec["staff"]["partners"]).replace(" ", "").split(";")
            partners_set.update(p for p in partnerslist if p)

    partnersLeaveList = []
    if partners_set:
        or_clauses = [{"staff.racf": {'$regex': p, '$options': "i"}} for p in partners_set]
        partnersRecords = list(eleaveDtl.find({"$or": or_clauses, "staff.status": {'$regex': "ACTIVE", '$options': "i"}}))

        for partnersrec in partnersRecords:
            staff = partnersrec["staff"]["name"]
            racf = partnersrec["staff"]["racf"]
            office = partnersrec["staff"]["hr_office"]
            dept = partnersrec["staff"]["dept"]
            for record in partnersrec["leave_record"]:
                if record["applicationStatus"] != approved and record["applicationStatus"] != pending:
                    continue
                leaveDetailsLst = []
                for details in record["details"]:
                    for period in details["period"]:
                        ldate_date = datetime.strptime(period["ldate"], '%Y-%m-%d').date()
                        days_diff = (ldate_date - date_input_date).days
                        if 0 <= days_diff <= 13:
                            leaveDetails = {
                                "startDate": period["ldate"],
                                "startTime": period["ltime"],
                                "endDate": period["ldate"],
                                "endTime": period["ltime"],
                                "workday": details["no_of_workday"],
                                "calendarDay": details["no_of_calendarday"]
                            }
                            leaveDetailsLst.append(leaveDetails)
                if leaveDetailsLst:
                    leaveRecord = {
                        "staff": staff,
                        "racf": racf,
                        "office": str(partnersrec["staff"]["office"]),
                        "dept": dept,
                        "ref_no": getDisplayRefNo(record["ref_no"], office, racf),
                        "type_id": record["type"],
                        "type": leaveTypeDict.get(record["type"], ""),
                        "approvalStatus": record["applicationStatus"],
                        "details": leaveDetailsLst
                    }
                    startDate = leaveDetailsLst[0]['startDate']
                    days_diff = (datetime.strptime(startDate, '%Y-%m-%d').date() - date_input_date).days
                    if 0 <= days_diff <= 13:
                        partnersLeaveList.append(leaveRecord)

    partnersLeaveList = sorted(partnersLeaveList, key=lambda d: (d["approvalStatus"], d["staff"], d["racf"], d["ref_no"]))
    partnersLeave.append(partnersLeaveList)

    # Re-structure
    final_result = []
    for index, rec in enumerate(partnersLeave[0]):
        final_result.append(rec)

    ### start preparing frontend presentation 

    # Check whether there are any partners to show.  If empty, return nothing     
    if final_result == []:
       return ({"pass": True, "error_message" : None, "result": [], "status_code": 200}) 

    ##print('final_result',  final_result)        
        
    weekday= ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']

    # Build a distinct staff list of staff, racf and office for partners with their respective office.
    unique_staff = {}
    for rec in final_result:
        key = (rec['staff'], rec['racf'], rec['office'], rec['dept'])
        if key not in unique_staff:
            unique_staff[key] = {'staff': rec['staff'], 'racf': rec['racf'], 'office': rec['office'], 'dept': rec['dept']}
    staffList = list(unique_staff.values())

    # Loop the staff (partners) list for each staff and put each staff's presentation into fortnight_all
    # the presentation includes 4 rows - Day of Month, Day of Week, AM and PM for 14 days 

    fortnight_all = []        

    # build holidays based for all office  
    query = { "_id":0, "Office":1, "Date":1, "Time":1, "Remark":1}           
    query_filter = {'$and':  [  {'Date' : { '$gte' : date_input_str  }},     
                    {'Date' : { '$lte' : end_date_str + "PM" } }
                     ] }                    
    holiday_all = list(holidays.find(query_filter, query))    ## holidays     

    # Group holidays by office
    from collections import defaultdict
    office_to_holidays = defaultdict(list)
    for h in holiday_all:
        office_to_holidays[h['Office']].append(h)

    for staff in staffList:       

        individual_result = [x for x in final_result if x["racf"] == staff['racf']]
            
        # step 1 - build the initial fortnight template 
        fortnight = []
        fortnight_dict = {}
        for x in range(14):
            day_date = date_input + timedelta(days=x)
            day_str = day_date.strftime('%Y-%m-%d')
            for time_slot in ['AM', 'PM']:
                dt = day_str + time_slot
                entry = {"datetime": dt}
                fortnight.append(entry)
                fortnight_dict[dt] = entry
        
        # step 2 - build holidays based on the staff's office          
        holidayList = office_to_holidays[staff['office']]
        
        # step 3 - fill the fortnight with holidays and weekend first. 
        for rec in fortnight:
            parse_date = parser.parse(rec["datetime"][0:10])
            weekOfDay = parse_date.weekday() + 1
            rec['status'] = ''
            rec['remark'] = ''
            if weekOfDay in [6, 7]:
                rec['leaveType'] = 'weekend'
            else:
                rec['leaveType'] = ''

        for rec in holidayList:                            
            remark = rec.get('Remark', "")                    
            dt = rec['Date'] + rec['Time']
            if dt in fortnight_dict and fortnight_dict[dt]['leaveType'] == '':
                fortnight_dict[dt]['leaveType'] = 'holiday'
                fortnight_dict[dt]['remark'] = remark
                
    
        # step 4 - fill each of the 28 slot (AM + PM for 14 days) with applied leaves if any from the staff 
        for rec in individual_result:                         
            for rec2 in rec['details']:
                date_str = rec2['startDate']
                start_time = rec2['startTime']
                status = rec['approvalStatus']
                type_id = rec['type_id']
                if start_time == "Full Day":
                    for time in ["AM", "PM"]:
                        dt = date_str + time
                        if dt in fortnight_dict and fortnight_dict[dt]["leaveType"] == '':
                            fortnight_dict[dt]['status'] = status
                            fortnight_dict[dt]['leaveType'] = type_id
                else:
                    dt = date_str + start_time
                    if dt in fortnight_dict and fortnight_dict[dt]["leaveType"] == '':
                        fortnight_dict[dt]['status'] = status
                        fortnight_dict[dt]['leaveType'] = type_id                   

        ##print('fortnight ', fortnight)                                  
                        
        # step 5 - build the presentation includes 4 rows - Day of Month, Day of Week, AM and PM for 14 days  
        partnersDOM = []        
        partnersDOW = []        
        partnersAM = []        
        partnersAMstatus = []        
        partnersPM = []        
        partnersPMstatus = []        
        partnersAMholiday= []
        partnersPMholiday= []

        for rec in fortnight:  
            if rec['datetime'][10:12] == "AM":
                   partnersDOM.append(rec['datetime'][8:10])       ## Day of Month 
                   partnersDOW.append(weekday[parser.parse(rec["datetime"][0:10]).weekday()])      # Day of Week
                   partnersAM.append(rec["leaveType"])   # AM Leave Type
                   partnersAMstatus.append(rec["status"])   # Approval Status
                   partnersAMholiday.append(rec["remark"])   # Holiday remark
            elif rec['datetime'][10:12] == "PM":
                   partnersPM.append(rec["leaveType"])   # PM Leave Type
                   partnersPMstatus.append(rec["status"])   # Approval Status
                   partnersPMholiday.append(rec["remark"])   # Holiday remark        

        partners = { 
            'partnersDOM' : partnersDOM,
            'partnersDOW' : partnersDOW,  
            'partnersAM'  : partnersAM,
            'partnersAMstatus' : partnersAMstatus, 
            'partnersPM': partnersPM,         
            'partnersPMstatus': partnersPMstatus,
            'partnersAMholiday' : partnersAMholiday,
            'partnersPMholiday' : partnersPMholiday
        }             

        # step 6 - package and send the list of staff and their presentation to React client 

        fortnight_all.append({
            'staff' : staff['staff'],
            'dept'  : staff['dept'],
            'partners' : partners 
        })                                               

    return ({"pass": True, "error_message" : None, "result": fortnight_all, "status_code": 200})

# Status_code 200: passed
# Status_code 504: failed, Staff Record Not Exist
# Status_code 505: failed, Incorrect parameters

@eleave.route("/api/partnersleave", methods=['POST'])
@checkLogged.check_logged
def apiPartnersLeave():
    psInput = request.get_json()
    result = listPartnersLeave(psInput)
    try: 
        return result, result['status_code'] # APP
    except:
        return result # postman

# Status_code 200: passed
# Status_code 501: failed, Not enough days left for the leave
# Status_code 502: failed, Overlapped leave period
# Status_code 503: failed, Leave Type Not Found
# Status_code 504: failed, Staff Record Not Exist
# Status_code 505: failed, Incorrect parameters
# Status_code 506: failed, Over 14 days

@eleave.route("/api/applyleave", methods=['POST'])
@checkLogged.check_logged
#@app.route("/api/applyleave", methods=['POST'])
def apiApplyLeave():    
    psInput = json.loads(request.form.get('entireLeaveRequest'))
    
    result = applyLeave(psInput)
    
    try: 
        return jsonify(result), result['Status_code'] 
    except:
        return jsonify(result)

@eleave.route("/api/listapprove", methods=['POST'])
@checkLogged.check_logged
#@app.route("/api/listapprove", methods=['POST'])
def apiListApprove():    
    psInput = request.get_json()
    result = listApprove(psInput)
    try: 
        return jsonify(result), result['Status_code'] # APP
    except:
        return jsonify(result) # postman

@eleave.route("/api/listleave", methods=['POST'])
@checkLogged.check_logged
#@app.route("/api/listleave", methods=['POST'])
#@app.route("/")
def apiListLeave():
    psInput = request.get_json()
    result = listLeave(psInput)
    try: 
        return jsonify(result), result['Status_code'] # APP
    except:
        return jsonify(result) # postman

@eleave.route("/api/ListApprovedByYear", methods=['POST'])
@checkLogged.check_logged
def apiListApprovedByYear():    
    psInput = request.get_json()
    result = listApprovedLeaveByYear(psInput)
    try: 
        return jsonify(result), result['Status_code'] # APP
    except:
        return jsonify(result) # postman

# Status_code 200: passed
# Status_code 504: failed, Staff Record Not Exist
# Status_code 505: failed, Incorrect parameters
# Status_code 603: failed, Current Status cannot be changed
# Status_code 604: failed, Only applicant himself / herself can cancel leave
# Status_code 605: failed, Leave cancel already submitted and waiting for approval
# Status_code 606: failed, Cannot cancel leave in the past period
# Status_code 607: failed, User is not the first approver
# Status_code 608: failed, User is not the second approver
# Status_code 609: failed, User is not the third approver
# Status_code 610: failed, Leave already approved
# Status_code 611: failed, Incorrect action

@eleave.route("/api/changestatus", methods=['POST'])
@checkLogged.check_logged
#@app.route("/api/changestatus", methods=['POST'])
def apiChangeStatus():
    psInput = request.get_json()
    result = changeStatus(psInput)
    try: 
        return jsonify(result), result['Status_code'] # APP
    except:
        return jsonify(result) # postman


# For maintenance 1 time only
# psInput = {"ref_no": "2023001", "racf": "NF1BHC", "sumbitdate": "2023-01-17", "approver1": "NF1KWY", "approval_date1": "2023-01-18" , "approver2": "NF1WWT", "approval_date2": "2023-01-20" , "approver3": "NF1VCC", "approval_date3": "2023-01-25" }

def apiChangeLeaveRecordDate(psInput):

    racf = psInput.get("racf")
    ref_no = psInput.get("ref_no")

    if len(racf) < 1 or len(ref_no) < 1:
        return ({"pass": False, "error_message" : "Incorrect parameters", "result": [], "status_code": 505}) 

    staffRecord = list(eleaveDtl.find ( {"staff.racf" : { '$regex' : racf, '$options' : "i"} } ) )

    for k in range(len(staffRecord[0]["leave_record"])):
        if str(staffRecord[0]["leave_record"][k]["ref_no"]) == ref_no:
            try:
                eleaveDtl.update_one ({"_id":staffRecord[0]["_id"] },
                                      { "$set" : { "leave_record." + str(k) + ".submit_date" : str(psInput.get("submitdate", "")) } } ) 
            except:
                return ({"pass": False, "error_message" : "Error on submit date", "result": [], "status_code": 505}) 
            try:
                eleaveDtl.update_one ({"_id":staffRecord[0]["_id"] },
                                      { "$set" : { "leave_record." + str(k) + ".approval.approver1" : psInput.get("approver1", "") } } )   
            except:
                return ({"pass": False, "error_message" : "Error on approver 1", "result": [], "status_code": 505}) 
            try:
                eleaveDtl.update_one ({"_id":staffRecord[0]["_id"] },
                                      { "$set" : { "leave_record." + str(k) + ".approval.approval_date1" : psInput.get("approval_date1", "") } } )   
            except:
                return ({"pass": False, "error_message" : "Error on approval date 1", "result": [], "status_code": 505}) 
            try:
                eleaveDtl.update_one ({"_id":staffRecord[0]["_id"] },
                                      { "$set" : { "leave_record." + str(k) + ".approval.approver2" : psInput.get("approver2", "") } } )   
            except:
                return ({"pass": False, "error_message" : "Error on approver 2", "result": [], "status_code": 505}) 
            try:
                eleaveDtl.update_one ({"_id":staffRecord[0]["_id"] },
                                      { "$set" : { "leave_record." + str(k) + ".approval.approval_date2" : psInput.get("approval_date2", "") } } )   
            except:
                return ({"pass": False, "error_message" : "Error on approval date 2", "result": [], "status_code": 505}) 
            try:
                eleaveDtl.update_one ({"_id":staffRecord[0]["_id"] },
                                      { "$set" : { "leave_record." + str(k) + ".approval.approver3" : psInput.get("approver3", "") } } )   
            except:
                return ({"pass": False, "error_message" : "Error on approver3", "result": [], "status_code": 505}) 
            try:
                eleaveDtl.update_one ({"_id":staffRecord[0]["_id"] },
                                      { "$set" : { "leave_record." + str(k) + ".approval.approval_date3" : psInput.get("approval_date3", "") } } )  
            except:
                return ({"pass": False, "error_message" : "Error on approval date 3", "result": [], "status_code": 505}) 
    return ({"pass": True, "error_message" : None, "result": [], "status_code": 200}) 


def submitRequest(psInput):

    # Regional = Hong Kong Office (for validation checking)
    ofc_for_checking = "HKG" if psInput['office'] in ('REG', 'HKG') else psInput['office']

    # Validation

    # duplicated approval#
    history_h = list(otherLeaves.find({"ref_no": psInput['ref_no']}))
    if len(history_h) > 0:
        # print (history_h)
        return ({"pass": False, "error_message" : "Error on duplicated approval #", "result": [], "status_code": 901}) 
    
    if psInput['office'] == "":
        return ({"pass": False, "error_message" : "Office must be selected", "result": [], "status_code": 902}) 
    
    if psInput['racf'] == "":
        return ({"pass": False, "error_message" : "RACF cannot be blank", "result": [], "status_code": 903}) 

    if psInput['leave_type'] == "":
        return ({"pass": False, "error_message" : "Leave Type must be selected", "result": [], "status_code": 904}) 
    
    if float(psInput['entitled_days']) > 0 and psInput['entitled_days'] == "":
        return ({"pass": False, "error_message" : "Entitled Days cannot be 0 or blank", "result": [], "status_code": 905}) 
    
    if psInput['period_start'] == "":
        return ({"pass": False, "error_message" : "Period Start cannot be blank", "result": [], "status_code": 907}) 
    
    if psInput['period_end'] == "":
        return ({"pass": False, "error_message" : "Period End cannot be blank", "result": [], "status_code": 908}) 
    
    # Entitled day should less than the days between start and end
    duration = (datetime.strptime(psInput['period_end'], "%Y-%m-%d") - datetime.strptime(psInput['period_start'], "%Y-%m-%d")) + timedelta(days=1)
    
    # Count Public Holidays if exclusive
    start_date = datetime.strptime(psInput['period_start'], "%Y-%m-%d").date()
    end_date = datetime.strptime(psInput['period_end'], "%Y-%m-%d").date()

    am_count = 0
    pm_count = 0

    holidays = getPublicHolidays(
        office=ofc_for_checking, 
        start_date=datetime.strptime(psInput['period_start'], "%Y-%m-%d"),
        end_date=datetime.strptime(psInput['period_end'], "%Y-%m-%d")
    )

    for h in holidays:
        h_date = datetime.strptime(h['Date'], "%Y-%m-%d").date()
        
        if start_date <= h_date <= end_date:
            if h['Time'] == 'AM':
                am_count += 1
            elif h['Time'] == 'PM':
                pm_count += 1

    if eval(psInput['excluded_holidays']):
        total_half_days = am_count + pm_count
        total_days = total_half_days * 0.5
    else:
        total_days = 0


    print (total_days)
    if duration < timedelta(days=float(psInput['entitled_days']) + total_days):
        return ({"pass": False, "error_message" : "Entitled days should be less than the duration of period start and end", "result": [], "status_code": 912}) 
    
    if psInput['accumulated'] == "":
        return ({"pass": False, "error_message" : "Contiguous option must be selected", "result": [], "status_code": 909}) 
    
    if psInput['excluded_holidays'] == "":
        return ({"pass": False, "error_message" : "Holiday/Weekend option must be selected", "result": [], "status_code": 910}) 

    # End date is eariler than period start
    if psInput['period_start'] > psInput['period_end']:
        return ({"pass": False, "error_message" : "Your end date of application is earlier than your start date input!", "result": [], "status_code": 911}) 
    
    # Match staff office
    if getStaffRecord(psInput['racf'])['staff']['hr_office'] != psInput['office']:
        return ({"pass": False, "error_message" : "Office code does not match with staff records", "result": [], "status_code": 912}) 
    
    # Check same type and same period start and end
    history_u = list(otherLeaves.find({"racf": psInput['racf'], "leave_type": psInput['leave_type']}))

    input_start = datetime.strptime(psInput['period_start'], '%Y-%m-%d').date()
    input_end = datetime.strptime(psInput['period_end'], '%Y-%m-%d').date()

    for record in history_u:
        if record.get('status') == 'Canceled' or record.get('status') == "Rejected":
            continue

        history_start = datetime.strptime(record['period_start'], '%Y-%m-%d').date()
        history_end = datetime.strptime(record['period_end'], '%Y-%m-%d').date()

        # Check if the two date ranges overlap
        if input_start <= history_end and input_end >= history_start:
            return {
                "pass": False, 
                "error_message": f"Requested period overlaps with an existing requisition ({record['ref_no']}).", 
                "result": [],
                "status_code": 913
            }

    # Create a document to be inserted
    request = {
        'office': psInput['office'],
        'year': int(psInput['year']),
        'racf': psInput['racf'],
        'ref_no': psInput['ref_no'],
        'leave_type': psInput['leave_type'],
        'entitled_days': float(psInput['entitled_days']),
        'ref_date': psInput['ref_date'],
        'period_start': psInput['period_start'],
        'period_end': psInput['period_end'],
        'accumulated': eval(psInput['accumulated']),
        'excluded_holidays': eval(psInput['excluded_holidays']),
        'status': "Pending"
    }

    # Insert the document into the collection
    otherLeaves.insert_one(request)

    # Email Session
    staff_fullname = getStaffRecord(psInput['racf'])['staff']['name']
    staff_dept = getStaffRecord(psInput['racf'])['staff']['dept']
    leave_name = str(list(leaveTypes.find({'leave_type_id': psInput['leave_type']}))[0]['leave_type']).title()
    reference_no_inEmail = psInput['ref_no']


    # Date presentation in email
    period_start = psInput['period_start'].replace('-', '/').split('/')
    period_start = f"{period_start[1]}/{period_start[2]}/{period_start[0]}"

    period_end = psInput['period_end'].replace('-', '/').split('/')
    period_end = f"{period_end[1]}/{period_end[2]}/{period_end[0]}"

    if psInput['ref_date']:
        ref_date = psInput['ref_date']
    else:
        ref_date = "NA"

    title = f"<E-LEAVE> {staff_fullname} ({staff_dept}) - Requisition for {leave_name} #PENDING"

    # Attachment
    attachments = psInput.get('attachments', None)
    attachment_names = None
        
    if attachments is not None:
        attachment_names = [file.filename for file in attachments]

    message = (
        f"Dear HR Approver, \n\n"
        f"A leave requisition is pending your approval.\n\n"
        f"APPLICANT: {staff_fullname} ({psInput['office']} / {staff_dept})\n"
        f"REFERENCE NO.: {reference_no_inEmail}\n"
        f"REQUISITION DATE: {ref_date}\n"
        f"ENTITLED DAYS: {float(psInput['entitled_days'])} DAY(S)\n"
        f"ALLOWED PERIOD: {period_start} to {period_end}\n\n"
        f"Please log into the system to review the details and take appropriate action.\n\n"
        f"Should you have any questions, please contact the requester.\n\n"
        f"Thank you,\n"
        f"e-Leave"
    )

    # message = (
    #     f"Dear Applicant, \n\n"
    #     f"Your requisition for {leave_name} is approved and details are as below.\n\n"
    #     f"REFERENCE NO.: {reference_no_inEmail}\n"
    #     f"REQUISITION DATE: {ref_date}\n"
    #     f"ENTITLED DAYS: {float(psInput['entitled_days'])} DAY(S)\n"
    #     f"ALLOWED PERIOD: {period_start} to {period_end}\n\n"
    #     f"When submitting your application, please choose the 'Others' Leave Type and Reference # indicated above from the system. "
    #     f"The remaining procedures will be the same as those for the prior leave request.\n\n"
    #     f"Should you have any queries, please contact local HR.\n\n"
    #     f"Thanks,\n"
    #     f"e-Leave"
    # )

    sendTo = getStaffRecord(psInput['racf'])['staff']['email']
    sendCc = "billy.chan@macys.com"

    try:
        postmarker(message, title, sendTo, sendCc, attachments, attachment_names)
    except:
        # If you're writing this to a file or sending it over a network, ensure you encode it properly
        message = message.encode('utf-8')
        localSend(message, title, sendTo, sendCc)


    return ({"pass": True, "error_message" : None, "result": [], "status_code": 200}) 



def cancelRequest(psInput):

    ref_no = psInput['ref_no']
    racf = psInput['racf']
    year = list(otherLeaves.find({"ref_no": psInput['ref_no']}))[0]['year']
    type = list(otherLeaves.find({"ref_no": psInput['ref_no']}))[0]['leave_type']

    # Validation
    workday = float(len(getAllLeave(racf, year, [type], False, ref_no))) * 0.5

    if workday > 0:
        return ({"pass": False, "error_message" : "Reference has not been completely canceled.", "result": [], "status_code": 810}) 
    
    # pass validation
    update = otherLeaves.update_one({"ref_no": ref_no}, {"$set": {"status": "Canceled"}})

    if update.matched_count > 0:

        # Successfully
        refDetails = getRefDetails(ref_no)[0]

        staff_fullname = getStaffRecord(refDetails['racf'])['staff']['name']
        staff_dept = getStaffRecord(refDetails['racf'])['staff']['dept']
        leave_name = str(list(leaveTypes.find({'leave_type_id': refDetails['leave_type']}))[0]['leave_type']).title()
        reference_no_inEmail = psInput['ref_no']

        title = f"<E-LEAVE> {staff_fullname} ({staff_dept}) - Requisition for {leave_name} #CANCELED"

        message = (
            f"Dear Applicant, \n\n"
            f"Your requisition for {leave_name} has been canceled.\n\n"
            f"REFERENCE NO.: {reference_no_inEmail}\n\n"
            f"This requisition is no longer available for applying leave.\n\n"
            f"Should you have any queries, please contact local HR.\n\n"
            f"Thanks,\n"
            f"e-Leave"
        )

        sendTo = getStaffRecord(refDetails['racf'])['staff']['email']
        sendCc = "billy.chan@macys.com"

        try:
            postmarker(message, title, sendTo, sendCc, None, None)
        except:
            # If you're writing this to a file or sending it over a network, ensure you encode it properly
            message = message.encode('utf-8')
            localSend(message, title, sendTo, sendCc)




        return ({"pass": True, "error_message" : None, "result": [], "status_code": 200}) 
    else:
        return ({"pass": False, "error_message" : "Cancel failed, please check the Approval #", "result": [], "status_code": 810}) 
    

def approveRequest(psInput):

    ref_no = psInput['ref_no']
    
    # pass validation
    update = otherLeaves.update_one({"ref_no": ref_no}, {"$set": {"status": "Approved"}})

    if update.matched_count > 0:

        # Successfully
        refDetails = getRefDetails(ref_no)[0]

        staff_fullname = getStaffRecord(refDetails['racf'])['staff']['name']
        staff_dept = getStaffRecord(refDetails['racf'])['staff']['dept']
        leave_name = str(list(leaveTypes.find({'leave_type_id': refDetails['leave_type']}))[0]['leave_type']).title()
        reference_no_inEmail = psInput['ref_no']

        title = f"<E-LEAVE> {staff_fullname} ({staff_dept}) - Requisition for {leave_name} #APPROVED"

        message = (
            f"Dear Applicant, \n\n"
            f"Your requisition for {leave_name} is approved and details are as below.\n\n"
            f"REFERENCE NO.: {reference_no_inEmail}\n"
            f"REQUISITION DATE: {refDetails['ref_date']}\n"
            f"ENTITLED DAYS: {float(refDetails['entitled_days'])} DAY(S)\n"
            f"ALLOWED PERIOD: {refDetails['period_start']} to {refDetails['period_end']}\n\n"
            f"When submitting your application, please choose the 'Others' Leave Type and Reference # indicated above from the system. "
            f"The remaining procedures will be the same as those for the prior leave request.\n\n"
            f"Should you have any queries, please contact local HR.\n\n"
            f"Thanks,\n"
            f"e-Leave"
        )

        sendTo = getStaffRecord(refDetails['racf'])['staff']['email']
        sendCc = "billy.chan@macys.com"

        try:
            postmarker(message, title, sendTo, sendCc, None, None)
        except:
            # If you're writing this to a file or sending it over a network, ensure you encode it properly
            message = message.encode('utf-8')
            localSend(message, title, sendTo, sendCc)

        return ({"pass": True, "error_message" : None, "result": [], "status_code": 200}) 
    else:
        return ({"pass": False, "error_message" : "Approve failed, please check the Approval #", "result": [], "status_code": 810}) 


def rejectRequest(psInput):

    ref_no = psInput['ref_no']
    
    # pass validation
    update = otherLeaves.update_one({"ref_no": ref_no}, {"$set": {"status": "Rejected"}})

    if update.matched_count > 0:

        # Successfully
        refDetails = getRefDetails(ref_no)[0]

        staff_fullname = getStaffRecord(refDetails['racf'])['staff']['name']
        staff_dept = getStaffRecord(refDetails['racf'])['staff']['dept']
        leave_name = str(list(leaveTypes.find({'leave_type_id': refDetails['leave_type']}))[0]['leave_type']).title()
        reference_no_inEmail = psInput['ref_no']

        title = f"<E-LEAVE> {staff_fullname} ({staff_dept}) - Requisition for {leave_name} #REJECTED"

        message = (
            f"Dear Applicant, \n\n"
            f"Your requisition for {leave_name} has been rejected.\n\n"
            f"REFERENCE NO.: {reference_no_inEmail}\n\n"
            f"This requisition is no longer available for applying leave.\n\n"
            f"Should you have any queries, please contact local HR.\n\n"
            f"Thanks,\n"
            f"e-Leave"
        )

        sendTo = getStaffRecord(refDetails['racf'])['staff']['email']
        sendCc = "billy.chan@macys.com"

        try:
            postmarker(message, title, sendTo, sendCc, None, None)
        except:
            # If you're writing this to a file or sending it over a network, ensure you encode it properly
            message = message.encode('utf-8')
            localSend(message, title, sendTo, sendCc)

        return ({"pass": True, "error_message" : None, "result": [], "status_code": 200}) 
    else:
        return ({"pass": False, "error_message" : "Approve failed, please check the Approval #", "result": [], "status_code": 810}) 


def getRefDetails(ref_no):
    result = []
    
    # Get the special leave record from other_leaves
    record = list(otherLeaves.find({"ref_no": ref_no}))

    for rec in record:
        # 1. Fetch all leave applications associated with this specific Reference Number
        # We query eleave_dtl for records where otherRefNo matches our ref_no
        # and ignore REJECTED or CANCELLED applications.
        staff_record = eleaveDtl.find_one({"staff.racf": rec['racf']})
        
        approved_days = 0.0
        pending_days = 0.0
        
        if staff_record and "leave_record" in staff_record:
            for leave in staff_record["leave_record"]:
                # Check if this leave application is linked to the special leave ref_no
                if leave.get("otherRefNo") == ref_no:
                    status_upper = leave.get("applicationStatus", "").upper()
                    
                    # Calculate total days in this specific application
                    leave_days = 0.0
                    for detail in leave.get("details", []):
                        leave_days += float(detail.get("no_of_workday", 0))
                    
                    # Sort into Approved or Pending
                    if status_upper == "APPROVED":
                        approved_days += leave_days
                    elif status_upper in ["PENDING", "SUBMITTED"]: 
                        pending_days += leave_days

        # 2. Calculate Balance
        entitled = float(rec['entitled_days'])
        days_left = entitled - approved_days - pending_days

        # 3. Append to result
        result.append({
            'ref_no': rec['ref_no'],
            'office': rec['office'],
            'year': rec['year'],
            'racf': rec['racf'],
            'leave_type': rec['leave_type'],
            'leave_name': str(list(leaveTypes.find({'leave_type_id': rec['leave_type']}))[0]['leave_type']).title(),
            'entitled_days': entitled,
            'approved_days': approved_days, # Added
            'pending_days': pending_days,   # Added
            'days_left': days_left,         # Added
            'ref_date': rec['ref_date'],
            'period_start': rec['period_start'],
            'period_end': rec['period_end'],
            'accumulated': rec['accumulated'],
            'excluded_holidays': rec['excluded_holidays']
        })

    return result

def checkOtherLeave(psInput):

    return ({"pass": True, "error_message" : "VALIDATION MODE.  Data pass validation.  Database NOT updated !", "result": [], "Status_code": 200})

def getAllSpecialRefPerUser(racf, super, current_time):

    special_leaves = getAllSpecialLeave()
    leave_type_map = {item['leave_type_id']: item['leave_type'] for item in special_leaves}
    
    query = {
        "racf": racf,
        "status": "Approved",
    }
    
    if not super:
        try:
            parsed_time = datetime.strptime(current_time, '%a %b %d %Y %H:%M:%S')
            formatted_time = parsed_time.strftime('%Y-%m-%d')
        except ValueError:
            formatted_time = current_time
            
        query["period_end"] = {"$gte": formatted_time}
        
    records = list(otherLeaves.find(query))
    result = []

    for rec in records:
        ref_no = rec['ref_no']
        detail = getRefDetails(ref_no)[0]

        if detail['days_left'] < 1:
            continue

        if '_id' in rec:
            rec['_id'] = str(rec['_id']) 
        
        type_id = rec.get('leave_type')
        rec['leave_type_name'] = leave_type_map.get(type_id, "Unknown Leave Type")

        result.append(rec)

    return result

@eleave.route("/api/getoffice")
@checkLogged.check_logged
def apiGetOffice():
    result = getAllOffice()
    return jsonify(result) 
    
@eleave.route("/api/getSpecialLeave")
@checkLogged.check_logged
def apiGetSpecialLeave():

    result = getAllSpecialLeave()
    return jsonify(result) 


@eleave.route("/api/getAllSpecialRef", methods=['GET'])
@checkLogged.check_logged
def apiGetAllSpecialRef():
    try:
        result = getAllSpecialRef()
        return jsonify(result)
    
    except Exception as e:
        print("Error in getAllSpecialRef:", e)
        return jsonify({"error": str(e)}), 500


@eleave.route("/api/getAllSpecialRefPerUser", methods=['POST'])
@checkLogged.check_logged
def apiGetAllSpecialRefPerUser():
    try:

        psInput = request.get_json()

        result = getAllSpecialRefPerUser(psInput['racf'], psInput['super_mode'], psInput['localTime'])
        return jsonify(result)
    
    except Exception as e:
        print("Error in getAllSpecialRefPerUser:", e)
        return jsonify({"error": str(e)}), 500



@eleave.route("/api/getRefDetails", methods=['POST'])
@checkLogged.check_logged
def apiGetRefDeatils():
    psInput = request.get_json()
    ref_no = psInput['ref_no']

    result = getRefDetails(ref_no)

    return jsonify(result) 

@eleave.route("/api/specialLeaveRefNo", methods=['POST'])
@checkLogged.check_logged
def apiSpecialLeaveRefNo():
    psInput = request.get_json()
    # Get Office code
    ofc = psInput['Office']
    year = psInput['Year']

    if ofc != '' and year != '':
        result = specialLeaveRefNo(ofc, year)
    else:
        result = ""
    try: 
        return str(result) # APP
    except:
        return str(result) # postman

@eleave.route("/api/submitRequest", methods=['POST'])
@checkLogged.check_logged
def apiSubmitRequest():    
    psInput = request.form.to_dict()
    
    result = submitRequest(psInput)
    
    try: 
        return jsonify(result), result['status_code']
    except:
        return jsonify(result)

@eleave.route("/api/cancelRequest", methods=['POST'])
@checkLogged.check_logged
def apiCancelRequest():    
    psInput = request.get_json()
    result = cancelRequest(psInput)
    try: 
        return jsonify(result), result['status_code'] # APP
    except:
        return jsonify(result) # postman


@eleave.route("/api/approveRequest", methods=['POST'])
@checkLogged.check_logged
def apiApproveRequest():    
    psInput = request.get_json()
    result = approveRequest(psInput)
    try: 
        return jsonify(result), result['status_code'] # APP
    except:
        return jsonify(result) # postman


@eleave.route("/api/rejectRequest", methods=['POST'])
@checkLogged.check_logged
def apiRejectRequest():    
    psInput = request.get_json()
    result = rejectRequest(psInput)
    try: 
        return jsonify(result), result['status_code'] # APP
    except:
        return jsonify(result) # postman

# @eleave.route('/api/downloadAzureDoc', methods=['POST'])
# @checkLogged.check_logged
# def download_azure_doc():
#     try:
#         data = request.json
#         doc_id = str(data.get('docId'))
        
#         if not doc_id:
#             return jsonify({"error_message": "Missing document ID"}), 400

#         azure_conn = os.environ['AZURE_CONNECTION_STRING']
#         azure_storage_name = os.environ['AZURE_CONTAINER_NAME']
        
#         service_client = BlobServiceClient.from_connection_string(azure_conn)
#         container_client = service_client.get_container_client(azure_storage_name)
        
#         target_blobs = []
#         ref_no = None
#         racf = None
        
#         blobs = container_client.list_blobs(include=['metadata'])
#         for blob in blobs:
#             if blob.metadata and blob.metadata.get('sharePointId') == doc_id:
#                 target_blobs.append(blob.name)
                
#                 if not ref_no:
#                     ref_no = blob.metadata.get('ref_no')
#                 if not racf:
#                     racf = blob.metadata.get('racf')
                
#         if not target_blobs:
#             return jsonify({"error_message": "Document not found in Azure Storage"}), 404
            
#         if len(target_blobs) == 1:
#             blob_client = container_client.get_blob_client(target_blobs[0])
#             download_stream = blob_client.download_blob()
#             file_bytes = download_stream.readall()
            
#             return send_file(
#                 io.BytesIO(file_bytes),
#                 download_name=target_blobs[0],
#                 as_attachment=True
#             )
#         else:
#             memory_file = io.BytesIO()
#             with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
#                 for blob_name in target_blobs:
#                     blob_client = container_client.get_blob_client(blob_name)
#                     download_stream = blob_client.download_blob()
#                     file_bytes = download_stream.readall()
                    
#                     zf.writestr(blob_name, file_bytes)
            
#             memory_file.seek(0)
            
#             if ref_no and racf:
#                 zip_filename = f"{ref_no} ({racf}).zip"
#             else:
#                 zip_filename = f"documents_{doc_id}.zip"
            
#             return send_file(
#                 memory_file,
#                 download_name=zip_filename,
#                 as_attachment=True
#             )

#     except Exception as e:
#         return jsonify({"error_message": f"Failed to download document: {str(e)}"}), 500




#########################################################################################################################
# To be deleted
#########################################################################################################################


# def Mailer_to_Go(message, title, sendTo, sendCC, attachment = "", attachmentname = ""):

#     # sender
#     try: #Heroku
#         sender_user = 'noreply'
#         sender_email = "@".join([sender_user, current_app.mailertogo_domain])
#         sender_name = 'noreply@mmgoverseas.app'
#     except: #Local
#         sender_email = 'noreply' + "@" + os.environ["MAILERTOGO_DOMAIN"]

#     # recipient
#     # By Vincent Cheng temp on 11/23/22
#     #recipient_email = "ken.yip@macys.com;vincent.cheng@macys.com"
#     #recipient_email = "brown.michael.v@gmail.com"
#     #recipient_email = "ken.yip@macysinc.onmicrosoft.com;vincent.cheng@macysinc.onmicrosoft.com"    

#     #Get recipient domain name
#     try: #Heroku
#         recipient_domain = current_app.recipient_domain
#         local_recipient_domain = current_app.macys_domain
#         recipient_email = sendTo.replace(local_recipient_domain, recipient_domain)
#     except: #Local
#         recipient_domain = current_app.config['recipient_domain']
#         recipient_email = sendTo

#     #Get recipient cc domain name
#     try: #Heroku
#         recipient_domain = current_app.recipient_domain
#         local_recipient_domain = current_app.macys_domain
#         recipient_cc_email = sendCC.replace(local_recipient_domain, recipient_domain)
#     except: #Local
#         recipient_domain = current_app.config['recipient_domain']
#         recipient_cc_email = sendCC

#     # subject
#     subject = title

#     # text body
#     body_plain = message

#     # html body
#     line_break = '\n' #used to replace line breaks with html breaks
#     body_html = f'''<html>
#         <head></head>
#         <body>
#         {'<br/>'.join(body_plain.split(line_break))}
#         </body>
#         </html>'''

#     # create message container
#     message = MIMEMultipart('alternative')
#     message['Subject'] = subject
#     message['From'] = sender_email
#     message['To'] = recipient_email
#     message['Cc'] = recipient_cc_email

#     print (message['From'])
#     print (message['To'])
#     print (message['Cc'])

#     # prepare plain and html message parts
#     part1 = MIMEText(body_plain, 'plain')
#     part2 = MIMEText(body_html, 'html')

#     # attach parts to message

#     message.attach(part1)
#     message.attach(part2)

#     # transform recipient to list
#     recipient_email = list(recipient_email.split(";"))
#     recipient_cc_email = list(recipient_cc_email.split(";"))

#     # Attachment Part
#     if len(attachment) > 0:
#         for index, bytesIOfile in enumerate(attachment):
#             try:
#                 part3 = MIMEApplication(bytesIOfile.getvalue())
#                 application_type = mimetypes.guess_type("a.xlsx")[0] or 'application/octet-stream' + " ;charset=UTF-8"
#                 part3.add_header('Content-Disposition', 'attachment', filename=attachmentname[index])
#                 part3.add_header('Content-Type', application_type)
#                 message.attach(part3)
#             except:
#                 pass
#     else:
#         pass

#     try:
#         host = current_app.mailertogo_host
#         port = current_app.mailertogo_port
#     except:
#         host = os.environ["MAILERTOGO_SMTP_HOST"]
#         port = os.environ["MAILERTOGO_SMTP_PORT"]
#         sc = checkSSL(host, int(port))
#         if sc == 404:
#             print ("Error Code 404, SMTP connection timeout.")
#             quit()
#     else:
#         pass

#     # send the message.
#     try:
#         server = smtplib.SMTP(host, port)
#         server.ehlo()
#         server.starttls()
#         server.ehlo()
#         server.login(current_app.mailertogo_user, current_app.mailertogo_password)
#         server.sendmail(sender_email, (recipient_email + recipient_cc_email), message.as_string())
#         server.close()
#     except Exception as e:
#         server = smtplib.SMTP(host, port)
#         server.ehlo()
#         server.sendmail(sender_email, (recipient_email + recipient_cc_email), message.as_string())
#         server.close()
#         print ("Local SMTP Email Sent!")
#     else:
#         print ("Email sent!")

# def sendGrid(message, title, sendTo, sendCC, attachment=None, attachmentname=None):

#     # Create the SendGrid message

#     # line break
#     html_message = '<br>'.join(message.split('\n'))
#     sg_message = Mail(
#         from_email='noreply@mmgoverseas.app',
#         to_emails=sendTo,
#         subject=title,
#         html_content=html_message
#     )        

#     # prevent empty cc list program error
#     if sendCC:
#         sendCC_list = list(sendCC.split(";"))
#         sendCC_list = list(set([value for value in sendCC_list if len(value) >= 8]))
#         cc_mail = [] 


#         for email in sendCC_list:
#             cc_mail.append(Cc(email, email)) 

#         n = len(sendCC_list)

#         cc_mail_first_n = cc_mail[:n]

        
#         sg_message.add_cc (cc_mail_first_n)

#     # BCc
#     sg_message.add_bcc(os.environ.get('Send_BCc'))
        
#     # Add attachments if specified
#     if attachment and attachmentname:
#         for index, file in enumerate(attachment):

#             encoded_file = base64.b64encode(file.getvalue()).decode()

#             attached_file = Attachment(
#                 FileContent(encoded_file),
#                 FileName(attachmentname[index]),
#                 FileType('application/octet-stream'),
#                 Disposition('attachment')
#             )
#             sg_message.attachment = attached_file

#     try:
#         # Send the message using the SendGrid API
#         sg = SendGridAPIClient(os.environ.get('SENDGRID_API_KEY'))
#         response = sg.send(sg_message)
#         print(response.status_code)
#     except Exception as e:
#         print(e.message)


# def checkSSL(host,port,timeout=1):
#     sock = socket.socket(socket.AF_INET,socket.SOCK_STREAM) #presumably 
#     sock.settimeout(timeout)
#     try:
#        sock.connect((host,port))
#     except:
#        return 404
#     else:
#        sock.close()
#        return 250


# Combine 2 time slot list and then sort by date and time.
# parameters:
# psLst1 : in the format : [{"ldate": datetime, "ltime": "AM" / "PM"}]
# psLst2 : in the format : [{"ldate": datetime, "ltime": "AM" / "PM"}]
# return : combined list in the format [{"ldate": datetime, "ltime": "AM" / "PM"}]
# def combineTime(psLst1, psLst2):
#     combinedLst = []

#     for s in psLst1:
#         slot = {
#             "ldate" : s["ldate"],
#             "ltime" : s["ltime"],
#             "type" : s["type"]
#         }
#         combinedLst.append(slot)
    
#     for s in psLst2:
#         slot = {
#             "ldate" : s["ldate"],
#             "ltime" : s["ltime"],
#             "type": s["type"]
#             }
#         combinedLst.append(slot)

#     combinedLst = sorted(combinedLst, key=lambda d: (d['ldate'], d["ltime"]))
#     return (combinedLst)



# function to apply leave
# parameters:
# psOffice : Office o f that staff for calculating holidays.
# psYear : Annual Leave Year
# psRacf : RACF of the applicant
# psLeaveType : Leave type applying. either :
#           "Annual Leave"
#           "Casual Leave"
#           "Sick Leave - No Medical Cert."
#           "Sick Leave - With Medical Cert."
#           "Work From Home"
# psLeaveLst : Leave period list, coverted into AM, PM.  format as [{"startDate": "2021-07-20", "startTime": "AM", "endDate": "2021-07-20", "endTime": "PM"}]
# psLeaveScreenLst : Leave period list, as at screen showing.  format as [{"startDate": "2021-07-20", "startTime": "Full Day, "endDate": "2021-07-21", "endTime": "AM"}]
# psSuperUser : whether the action is in super use mode.  True / False
# return:
# reject = 0 : leave application, pass = true, no error message.  Leave details will insert into database
# reject = 1 : leave application failed, period overlap found.  pass = false, error message : Leave applied are overlapping each other.
# reject = 2 : leave application failed, consecutive days not pass for Annual Leave and Casual Leave.  pass = false, error message : Leave applied is over 2 weeks.
# reject = 3 : leave application failed, consecutive days not pass for Sick Leave with no cert.  pass = false, error message : Medical certificate is required if sick leave application is more than 1 Day.
# reject = 4 : leave application failed, leave applying > leave balance. pass = false, error message : Not enough days left for the leave.
# reject = 5 : leave application failed, cannot update database.  pass = false, error message : Fail to update database

# def applyLeaveOld (psInput):
# #def applyLeave (psInput):

#     getLeaveTypes()
#     getLeaveGroups()
#     psYear = psInput.get("year", 0)
#     psRacf = psInput.get("racf", "")
#     psLeaveType = psInput.get("type", "")
#     psLeaveLst = psInput.get("applying", "")
#     psLeaveScreenLst = psInput.get("applyingScreen", "")
#     psUpdateDB = psInput.get("updateDB", True)
#     SharePointID = psInput.get("sharePointId")
#     timeZone = psInput.get("timeZone", "")

#     # For development
#     newres = applyLeaveNew(psInput)

#     # Get Super User 
#     try:
#         psSuperUser = session["superUser"]
#     except:
#         psSuperUser = False
    
#     if psSuperUser:
#         psSuperUser = psInput.get("superUser")
#     else:
#         psSuperUser = False

#     if psYear == 0 or len(psRacf) == 0 or len(psLeaveType) == 0 or len(psLeaveLst) == 0 or len(psLeaveScreenLst) == 0:
#         return ({"pass": False, "error_message" : "Incorrect parameters", "result": None, "Status_code": 505})
#     staffRecord = getStaffRecord(psRacf)
#     leaveHistoryLst = getLeaveHistory(psYear, psYear, staffRecord)
#     leaveTypeAttr = (list(filter(lambda r: (r["leave_type_id"].upper() == psLeaveType), leaveTypeLst))[0])
#     if not isinstance(staffRecord, dict):
#         oldres = {"pass": False, "error_message" : "Staff Record Not Exist", "result": None, "Status_code": 504}
#         checkResult(oldres, newres, psInput, psRacf)
#         return (oldres)
#         #return ({"pass": False, "error_message" : "Staff Record Not Exist", "result": None, "Status_code": 504}) 

#     if len (leaveTypeAttr) == 0:
#         oldres = {"pass": False, "error_message" : "Leave Type Not Found", "result": None, "Status_code": 503}
#         checkResult(oldres, newres, psInput, psRacf)
#         return (oldres)
#         #return ({"pass": False, "error_message" : "Leave Type Not Found", "result": None, "Status_code": 503}) 
#     office = staffRecord["staff"]["office"]
#     # applyingSlotLst : list which combine leave application in all rows
#     # applyingSlotLstByRow : list which keep leave application row by row.
#     applyingSlotLst = [ ]
#     applyingSlotLstByRow = [ ]
#     overlap = False
#     # Loop through each applying leave period
#     for rec in psLeaveLst:  
        
#         withinYr = chkPeriod (rec["startDate"], rec["endDate"], psYear)
#         if not withinYr.get('pass'):
#             oldres = {"pass": False, "error_message": withinYr.get('error_message'), "result": None, "Status_code": withinYr.get('Status_code')}
#             checkResult(oldres, newres, psInput, psRacf)
#             return (oldres)
#             #return({"pass": False, "error_message": withinYr.get('error_message'), "result": None, "Status_code": withinYr.get('Status_code')})
#         tmpApplyingSlotLst = checkOverlap(rec["startDate"], rec["startTime"], rec["endDate"], rec["endTime"],  psYear, office, staffRecord, applyingSlotLst, psLeaveType)
#         # If no overlap, will get the expanded date slot for leave applying, else leaveSlotLst is empty    
#         # put the expanded date slot into applyingSlotLstByRow for saving to DB into separate document.
#         #if len(tmpApplyingSlotLst) > 0:
#         if isinstance(tmpApplyingSlotLst, list):
#             applyingSlotLst = combineTime(applyingSlotLst, tmpApplyingSlotLst)
#             applyingSlotLstByRow.append(tmpApplyingSlotLst)
#         else:
#             overlap = True
#             errormsg = tmpApplyingSlotLst
#             break    
#     if overlap:
#         oldres = {"pass": False, "error_message" : errormsg, "result": None, "Status_code": 502}
#         checkResult(oldres, newres, psInput, psRacf)
#         return (oldres)
#         #return ({"pass": False, "error_message" : errormsg, "result": None, "Status_code": 502})
#     else:
#         # Leave type = Annual Leave or Casual Leave :
#         # 1. No overlap with
#         #  the period already applied
#         # 2. Consecutive leave days cannot more than the limits (include annual leave, casual leave, public holidays and weekends), unless leave is applied under superuser mode 
#         # 3. Leave applied cannot more than leave entitle + carry forward.
#         #if leaveTypeAttr.get("max_consecutive_days",0) > 0:
#         result = checkConsecutiveDays(psYear, office, staffRecord, applyingSlotLst, leaveTypeAttr) 
#         if not result.get("pass") and not psSuperUser:
#             oldres = result
#             checkResult(oldres, newres, psInput, psRacf)
#             return (result)

#         if leaveTypeAttr.get("entitlement_field", "") != "":
#             if checkBalance(psYear, leaveTypeAttr, staffRecord, applyingSlotLst) < 0:
#                 oldres = {"pass": False, "error_message" : "Not enough days left for the leave", "result": None, "Status_code": 501}
#                 checkResult(oldres, newres, psInput, psRacf)
#                 return (oldres)
#                 #return ({"pass": False, "error_message" : "Not enough days left for the leave.", "result": None, "Status_code": 501})
    
#         newRefNo = getNewRefNo(psYear, psRacf)
#         rowNo = 0

#         total_sl = 0
#         warning_sl_message = ""
#         rowDtlLst = [ ]
#         for row in applyingSlotLstByRow:
#             #noOfCalendarDay = getCalendarDay(psYear, office, staffRecord, row)
#             noOfCalendarDay = getCalendarDay(psYear, office, staffRecord, row, leaveTypeAttr)
#             noOfWorkDay = getWorkDay (row)
#             timeSlotLst = [ ]
#             for s in row:
#                 timeslot = {
#                     "ldate" : date2Str(s["ldate"]),
#                     "ltime": s["ltime"]
#                 }
#                 timeSlotLst.append(timeslot)
        

#             if (row[0]['type']) not in ["LVE04","LVE05"]:
                
#                 rowDtl = {
#                     "start_date": psLeaveScreenLst[rowNo]["startDate"],
#                     "start_time": psLeaveScreenLst[rowNo]["startTime"],
#                     "end_date": psLeaveScreenLst[rowNo]["endDate"],
#                     "end_time": psLeaveScreenLst[rowNo]["endTime"],
#                     "no_of_workday": noOfWorkDay,
#                     "no_of_calendarday": noOfCalendarDay,             
#                     "period" : timeSlotLst
#                         }
#             elif (row[0]['type']) in ["LVE04","LVE05"]:
#                 rowDtl = {
#                     "start_date": psLeaveScreenLst[rowNo]["startDate"],
#                     "start_time": psLeaveScreenLst[rowNo]["startTime"],
#                     "end_date": psLeaveScreenLst[rowNo]["endDate"],
#                     "end_time": psLeaveScreenLst[rowNo]["endTime"],
#                     "no_of_workday": noOfWorkDay,
#                     "no_of_calendarday": noOfCalendarDay,             
#                     "no_of_consective": countConsecutiveDaysByType(psYear, staffRecord["staff"]["office"], leaveHistoryLst, applyingSlotLst, ["LVE04","LVE05"]),
#                     "period" : timeSlotLst
#                         }
#                 # check total sick leave should be less than 7, else it will appear warnings
#                 total_sl = noOfWorkDay
#                 total_sl += countLeave(psYear, str("LVE04"), df['gcStatusApproved'][0], staffRecord) + countLeave(psYear, str("LVE05"), df['gcStatusApproved'][0], staffRecord)
#                 total_sl += countLeave(psYear, str("LVE04"), df['gcStatusPending'][0], staffRecord) + countLeave(psYear, str("LVE05"), df['gcStatusPending'][0], staffRecord)
#                 if total_sl > 7:
#                     warning_sl_message = "Reminder:  Total Full Paid Sick Leave taken has already reached 7 days which is the maximum cap of current leave calendar year (included below leave application)"
#             rowDtlLst.append(rowDtl)
#             rowNo += 1
#             approvallist = {
#                 "approver1": staffRecord['staff']['approver1'],
#                 "approval_date1": "",
#                 "approver2": staffRecord['staff']['approver2'],
#                 "approval_date2": "",                
#                 "approver3": staffRecord['staff']['approver3'],
#                 "approval_date3": ""
#             }

#         newLeaveRecord = {
#             "ref_no" : newRefNo,
#             "sharePointId" : SharePointID,
#             "year" : psYear,
#             "type" : psLeaveType,
#             "applicationStatus" : df['gcStatusPending'][0],
#             "approvalStatus": df['gcStatusPending1'][0],
#             "submit_date": date2Str(date.today()),
#             "lastUpdate": psRacf,
#             "updateDate": date2Str(date.today()),
#             "timeZone": timeZone,
#             "approval": approvallist,
#             "details": rowDtlLst
#         }
#         id = staffRecord["_id"]
#         leaveRecord = staffRecord["leave_record"]
#         leaveRecord.append(newLeaveRecord)
#         updateRecordLst = [ ]
#         updateRecord = {
#             "field" : "leave_record",
#             "value" : leaveRecord,
#         }
#         updateRecordLst.append(updateRecord)
#         if psUpdateDB:
#             result = updateDB2(id, updateRecordLst)
#         else:
#             oldres = {"pass": True, "error_message" : "VALIDATION MODE.  Data pass validation.  Database NOT updated !", "result": [{"workday": noOfWorkDay, "calendarDay": noOfCalendarDay}], "Status_code": 200, "Warnings": warning_sl_message}
#             checkResult(oldres, newres, psInput, psRacf)
#             return oldres
#             #return ({"pass": True, "error_message" : "VALIDATION MODE.  Data pass validation.  Database NOT updated !", "result": [{"workday": noOfWorkDay, "calendarDay": noOfCalendarDay}], "Status_code": 200, "Warnings": warning_sl_message})
#         if result.get("pass") and psUpdateDB:
#             sendEmail(staffRecord, newRefNo, df['gcStatusPending1'][0], df['gcActionApply'][0], df['gcActionApply'][0], 1, 1)

#         return (result)                

# # Temp to check the difference of result between new apply leave and old apply leave
# def checkResult(old_result, new_result, input, user):

#     boolean = new_result != old_result

#     if boolean:
#         message = f"{user} is entering the leave input {input} \n\n ** But the result does not match **: \n\n Old : {old_result} \n New: {new_result} \n\n"
#         title = "Please check the old and new result for applyleave"
#         sendTo = "billy.chan@macys.com"
#         sendCc = ""
#     else:
#         message = f"{user} is entering the leave input {input} \n\n ** The result is correct: **\n\n Old : {old_result} \n New: {new_result} \n\n"
#         title = "No need to check the result for applyleave"
#         sendTo = "billy.chan@macys.com"
#         sendCc = ""

#     try:
#         mailer_to_mongoDB(message, title, sendTo, sendCc, "", "")
#         #sendGrid(message, title, sendTo, sendCc, "", "")
#     except:
#         pass

# parameters - leaveHistoryLst, Type
# leaveHistoryLst Example: {'ref_no': 2022001, 'office': 'REG', 'racf': 'NF1BHC', 'staffname': 'BILLY CHAN', 'empID': '00013', 'dept': 'PBT', 'position': 'Regional Analyst Programmer', 'year': 2022, 'type': 'LVE05', 'sharePointId': '', 'startDate': '2022-12-05', 'startTime': 'Full Day', 'endDate': '2022-12-05', 'endTime': 'Full Day', 'applicationStatus': 'PENDING' ...
# Type Example (String): LVE01, LVE02 
# def countConsecutiveDaysByType( psYear, psOffice, leaveHistoryLst, ApplyLeaveLst, Type):
#     consecutiveSlot = 0

#     applicationleave = False

#     LeaveHistoryLst = list(filter(lambda r: (r["applicationStatus"].upper() != df['gcStatusCancel'][0] and r["applicationStatus"].upper() != df['gcStatusReject'][0] and r["type"] in Type), leaveHistoryLst))
#     LeaveHistoryLst.sort(key=lambda x: x.get('ldate'))
#     combinedTimeSlot = combineTime(LeaveHistoryLst, ApplyLeaveLst)
#     weekendHolidaysLst = getWeekendHolidays((psYear - 1), (psYear + 1), psOffice)
#     combinedTimeSlot = combineTime (combinedTimeSlot, weekendHolidaysLst)


#     currDate = combinedTimeSlot[0]["ldate"]
#     currTime = combinedTimeSlot[0]["ltime"]

#     maxvalue = 0
#     canQuit = False

#     for t in combinedTimeSlot:
#         if currDate == t["ldate"] and t["type"]:
#             if t["type"] == "weekend" or t["type"] == "holiday":
#                 consecutiveSlot = consecutiveSlot
#             else:
#                 consecutiveSlot += 1
#             currTime = "PM"
#         elif (currDate == t["ldate"] + timedelta(-1)) and (currTime == "PM") and (t["ltime"] == "AM"):
#             if t["type"] == "weekend" or t["type"] == "holiday":
#                 consecutiveSlot = consecutiveSlot
#                 currDate = t["ldate"]
#                 currTime = t["ltime"]
#             else:
#                 consecutiveSlot += 1
#                 currDate = t["ldate"]
#                 currTime = t["ltime"]
#         else:
#             if t["type"] in Type:
#                 consecutiveSlot = 1
#             else:
#                 consecutiveSlot = 0
#             currDate = t["ldate"]
#             currTime = t["ltime"]
#         #print ("currDate : " + str(currDate) + " / " + "currTime : " + str(currTime)) #for my checking
#         #print ("tDate : " + str(t["ldate"]) + " / " + "tTime : " + str(t["ltime"]) + " / " + "tType : " + str(t["type"])) #for my checking
#         #print ("consecutiveSlot: " + str(consecutiveSlot)) #for my checking

#         #if str((currDate).year) == "2023":

#         #    print ("currDate : " + str(currDate) + " / " + "currTime : " + str(currTime)) #for my checking
#         #    print ("tDate : " + str(t["ldate"]) + " / " + "tTime : " + str(t["ltime"]) + " / " + "tType : " + str(t["type"])) #for my checking
#         #    print ("consecutiveSlot: " + str(consecutiveSlot)) #for my checking
#         #    print ("maxvalue :" + str(maxvalue))

#         # check apply leave is in checking period
#         if (ApplyLeaveLst[-1]['ldate'] == currDate) and currTime == ApplyLeaveLst[-1]["ltime"]:
#             canQuit = True
        
#         if consecutiveSlot > 0 :
#             maxvalue = consecutiveSlot
#         if consecutiveSlot == 0 and maxvalue > 0 and canQuit:
#             consecutiveSlot = maxvalue / 2
#             return consecutiveSlot
    
    
#     workDay = getWorkDay(ApplyLeaveLst)
#     if workDay > consecutiveSlot:
#         consecutiveSlot = workDay


#     return consecutiveSlot



# def checkConsecutiveSickLeave (psCombinedSickLeave, psMaxSlNoCert, psApplyingSlotLst):
#     slNoCertConsecutiveSlot = 0

#     currDate = psCombinedSickLeave[0]["ldate"]
#     currTime = psCombinedSickLeave[0]["ltime"]

#     currConsecutiveDay = False
    
#     for t in psCombinedSickLeave:
#         if currDate == t["ldate"]:

#             if t["type"] == "LVE05" or t["type"] == "LVE04":
#                 slNoCertConsecutiveSlot += 1
#                 #if t["ldate"] ==  psApplyingSlotLst[0]["ldate"] and currTime == psApplyingSlotLst[0]["ltime"]:
#                 if t["ldate"] ==  psApplyingSlotLst[0]["ldate"] and t["ltime"] == psApplyingSlotLst[0]["ltime"]:
#                     currConsecutiveDay = True
                
#             currTime = "PM"
#         elif (currDate == t["ldate"] + timedelta(-1)) and (currTime == "PM") and (t["ltime"] == "AM"):
#             if t["type"] == "LVE05" or t["type"] == "LVE04":
#                 slNoCertConsecutiveSlot += 1
#                 #if t["ldate"] ==  psApplyingSlotLst[0]["ldate"] and currTime == psApplyingSlotLst[0]["ltime"]:
#                 if t["ldate"] ==  psApplyingSlotLst[0]["ldate"] and t["ltime"] == psApplyingSlotLst[0]["ltime"]:
#                     currConsecutiveDay = True
#             currDate = t["ldate"]
#             currTime = t["ltime"]
#         else:
#             #currConsecutiveDay = True
#             if t["type"] == "LVE05" or t["type"] == "LVE04":
#                 slNoCertConsecutiveSlot = 1
#                 if t["ldate"] ==  psApplyingSlotLst[0]["ldate"] and t["ltime"] == psApplyingSlotLst[0]["ltime"]:
#                     currConsecutiveDay = True
#             else:
#                 currConsecutiveDay = False
#                 slNoCertConsecutiveSlot = 0
#             currDate = t["ldate"]
#             currTime = t["ltime"]

#         #print ("currDate : " + str(currDate) + " / " + "currTime : " + str(currTime))

#         #print ("tDate : " + str(t["ldate"]) + " / " + "tTime : " + str(t["ltime"]) + " / " + "tType : " + str(t["type"]))

#         #print ("consecutiveSlot: " + str(slNoCertConsecutiveSlot) + " currConsectiveDay : " + str(currConsecutiveDay))


#         if slNoCertConsecutiveSlot > (psMaxSlNoCert * 2) and currConsecutiveDay and psApplyingSlotLst[0]['type'] == 'LVE05':
#             return ({"pass": False, "error_message" : "Reminder: For any sick leave periods that exceed 2 contiguous days, sick leave certificate is required", "result": None, "Status_code": 506})

#     return({"pass": True, "error_message": "", "result": None, "daycount": 1, "Status_code": 200}) 



# expand leave application days to half day timeslot
# check if the leave slot fall into holidays, skip that leave slot if it is 
# check if the leave slot overlapped with those applied before, retrun error if it is
# return whole leave slot otherwise
#parameter : 
#psStartDate : Start Date of leave applying, in string (yyyy-mm-dd) format
#psStartTime : Start Time of leave applying, in string, either "AM" or "PM"
#psEndDate : End Date of leave applying in string (yyyy-mm-dd) format
#psEndTime : End Time of leave applyingm in string, either "AM" or "PM"
#psHolidayLst : List of time slot with holidays and weekends
#psLeaveHistoryLst : List of time slot with leave history in the format [{"ref_no" int, "year" int, "type" string, "status" string, "ldate": datetime, "ltime": "AM" / "PM"}]
#return:
#if no overlap, return list of leave slot for the leave applying in the format :[{"ldate": datetime, "ltime": "AM" / "PM"}]
#if overlap, return empty list
# def checkOverlap(psStartDate, psStartTime, psEndDate, psEndTime, psYear, psOffice, psRecord, psApplyingSlotLst, psLeaveType):
#     currDate = str2Date(psStartDate)
#     currTime = psStartTime
#     leaveDtl = [ ]
#     weekendHolidaysLst = getWeekendHolidays((psYear - 1), (psYear + 1), psOffice)
#     leaveHistoryLst = getLeaveHistory((psYear - 1), (psYear + 1), psRecord)
#     #exclude leave that is canceled or rejected
#     leaveHistoryLst = list(filter(lambda r: (r["applicationStatus"].upper() != df['gcStatusCancel'][0] and r["applicationStatus"].upper() != df['gcStatusReject'][0]), leaveHistoryLst))
#     # loop through the holidays range applied

#     while currDate <= str2Date(psEndDate):
#         found = [ ]
#         # check if the leave slot overlapped with the leave already applied
#         overlap = filter(lambda o: (o["ldate"] == currDate) and (o["ltime"] == currTime), leaveHistoryLst)
#         if (len(list(overlap))) > 0:
#             leaveDtl = [ ]
#             leaveDtl = "Leave applying is overlapping"
#             return leaveDtl
#         # check if the leave slot overlapped with the leave applying in different rows
#         overlap = filter(lambda o: (o["ldate"] == currDate) and (o["ltime"] == currTime), psApplyingSlotLst)
#         if (len(list(overlap))) > 0:
#             leaveDtl = [ ]
#             leaveDtl = "Leave applying is overlapping"
#             return leaveDtl        
#         # check if the leave slot is in holiday and weekend, if it is not in the holiday list, "found" will be empty and proceed to record in leave detail
#         # if the leave slot is in holiday, "found" will not empty and will skip to record in leave detail
#         found = list(filter(lambda d: (d["ldate"] == currDate) and (d["ltime"] == currTime), weekendHolidaysLst))
#         if len(found) == 1 and (len(leaveDtl)) == 0:
#             leaveDtl = [ ]
#             leaveDtl = "Leave applying start in Weekends / Holidays"
#             return leaveDtl
#         if len(found) == 0:
#             isHoliday = False
#             leaveSlot = { "ldate": currDate, "ltime": currTime, "type": psLeaveType}
#             leaveDtl.append(dict(leaveSlot))
#         else:
#             isHoliday = True
#         if currTime.upper() == "PM":
#             currDate = currDate + timedelta(1)
#             currTime = "AM"
#         elif (currDate == str2Date(psEndDate)) and (psEndTime == "AM"):
#             break
#         else:
#             currTime = "PM" 
#     if (isHoliday):
#         leaveDtl = [ ]
#         leaveDtl = "Leave applying end in Weekends / Holidays"
#         return leaveDtl   
#     return leaveDtl


# check consecutive days of the leave applying and see if it exceeds the consecutive days allowed.
# parameters: 
# psCombinedTimeSlot : list of applying leave, leave already applied, holidays and weekend in the format [{"ldate": datetime, "ltime": "AM" / "PM"}]
# psLimit : max. consecutive days allowed 
# return :
# total consecutive days or the max. consecutive days allowed + 0.5 (when the consective days exceeds the days allowed, it will stop checking and return.)
# def checkConsecutiveDays (psYear, psOffice, psRecord, psApplyingSlotLst, psLeaveTypeAttr):
#     consecutiveSlot = -3
#     groupAttrLst = list(filter(lambda r: (r["groupID"] == psLeaveTypeAttr.get("leave_group")), leaveGroupLst))[0]
#     if groupAttrLst.get("max_consecutive_days", "")  != "":
#         relatedLveLst = []
#         for lve in leaveTypeLst:
#             #if lve["leave_group"] == groupAttrLst.get("groupID"):
#             if lve["consecutive_days_group"] == psLeaveTypeAttr.get("consecutive_days_group"):
#                 relatedLveLst.append(lve["leave_type_id"])
    
#         leaveHistoryLst = getLeaveHistory((psYear - 1), (psYear + 1), psRecord)
#         # if leave is sick leave with cert or sick leave with no cert, check the Consecutive Sick Leave with no cert days
#         if groupAttrLst.get("sick_leave", False):
#             relatedSlLst = []
#             for sl in leaveTypeLst:
#                 slGrpAttrLst = list(filter(lambda r: (r["groupID"] == sl["leave_group"]), leaveGroupLst))[0]
#                 if slGrpAttrLst.get("sick_leave", False):
#                     relatedSlLst.append(sl["leave_type_id"])
#                     if slGrpAttrLst.get("max_consecutive_days", "") != "":
#                         maxConsecutiveSlNoCert = slGrpAttrLst.get("max_consecutive_days")
#             slLeaveHistoryLst = list(filter(lambda r: (r["applicationStatus"].upper() != df['gcStatusCancel'][0] and r["applicationStatus"].upper() != df['gcStatusReject'][0] and r["type"] in relatedSlLst), leaveHistoryLst))
#             combinedTimeSlot = combineTime(slLeaveHistoryLst, psApplyingSlotLst)

#             # Added to check weekend and holiday if sick leave
#             slweekendHolidaysLst = getWeekendHolidays((psYear - 1), (psYear + 1), psOffice)
#             combinedTimeSlot = combineTime (combinedTimeSlot, slweekendHolidaysLst)

#             result = checkConsecutiveSickLeave(combinedTimeSlot, maxConsecutiveSlNoCert, psApplyingSlotLst)
#             if not result.get("pass"):
#                 return (result)
#         leaveHistoryLst = list(filter(lambda r: (r["applicationStatus"].upper() != df['gcStatusCancel'][0] and r["applicationStatus"].upper() != df['gcStatusReject'][0] and r["type"] in relatedLveLst), leaveHistoryLst))
#         combinedTimeSlot = combineTime(leaveHistoryLst, psApplyingSlotLst)
#         if groupAttrLst.get("consecutive_include_holidays", False):
#             weekendHolidaysLst = getWeekendHolidays((psYear - 1), (psYear + 1), psOffice)
#             combinedTimeSlot = combineTime (combinedTimeSlot, weekendHolidaysLst)
    
#         currDate = combinedTimeSlot[0]["ldate"]
#         currTime = combinedTimeSlot[0]["ltime"]

#         currConsecutiveDay = False

#         for t in combinedTimeSlot:
#             #print (str(currDate) + " / " + str(t["ldate"]) + " consecutiveSlot: " + str(consecutiveSlot))
#             if currDate == t["ldate"]:
#                 consecutiveSlot += 1
#                 if t["ldate"] ==  psApplyingSlotLst[0]["ldate"] and currTime == psApplyingSlotLst[0]["ltime"]:
#                     currConsecutiveDay = True
#                 currTime = "PM"
#             elif (currDate == t["ldate"] + timedelta(-1)) and (currTime == "PM") and (t["ltime"] == "AM"):
#                 consecutiveSlot += 1
#                 if t["ldate"] ==  psApplyingSlotLst[0]["ldate"] and currTime == psApplyingSlotLst[0]["ltime"]:
#                     currConsecutiveDay = True
#                 currDate = t["ldate"]
#                 currTime = t["ltime"]                
#             else:
#                 if t["ldate"].strftime('%A') in ["Saturday","Sunday"]:
#                     consecutiveSlot = -3
#                 else:
#                     consecutiveSlot = 0
#                 #consecutiveSlot = -3
#                 #if t["ldate"] ==  psApplyingSlotLst[0]["ldate"]:
#                 #    consecutiveSlot = 0
#                 currDate = t["ldate"]
#                 currTime = t["ltime"]
#                 currConsecutiveDay = False

#             if consecutiveSlot > (groupAttrLst.get("max_consecutive_days", 0) * 2) and currConsecutiveDay:
#                 #if groupAttrLst.get("groupID") != 1:
#                     #return ({"pass": False, "error_message" : "Reminder: For any sick leave periods that exceed 2 contiguous days, sick leave certificate is required", "result": None,  "Status_code": 506})
#                 if groupAttrLst.get("groupID") == 1:
#                     return ({"pass": False, "error_message" : "Reminder: Maximum vacation taken at any one time is 2 WEEKS including Public Holidays, Saturdays and Sundays", "result": None,  "Status_code": 506})
   
#     elif psApplyingSlotLst[0]['type'] == 'LVE04':

#         relatedSlLst = ['LVE04', 'LVE05']

#         leaveHistoryLst = getLeaveHistory((psYear - 1), (psYear + 1), psRecord)

#         slLeaveHistoryLst = list(filter(lambda r: (r["applicationStatus"].upper() != df['gcStatusCancel'][0] and r["applicationStatus"].upper() != df['gcStatusReject'][0] and r["type"] in relatedSlLst), leaveHistoryLst))
#         combinedTimeSlot = combineTime(slLeaveHistoryLst, psApplyingSlotLst)
#         # Added to check weekend and holiday if sick leave
#         slweekendHolidaysLst = getWeekendHolidays((psYear - 1), (psYear + 1), psOffice)
#         combinedTimeSlot = combineTime (combinedTimeSlot, slweekendHolidaysLst)

#         result = checkConsecutiveSickLeave(combinedTimeSlot, 1, psApplyingSlotLst)

        
#         if not result.get("pass"):
#             return (result)

#     else:

#         return({"pass": True, "error_message": "", "result": None, "Status_code": 200}) 
     
#     return ({"pass": True, "error_message": "", "result": None,  "Status_code": 200})

# count the total calendar date.
# input : list of leave applied - psPeriod, list of leave applied + holidays +weekend - psCombinedSlotLst
# total calendar date = 
#   consecutive calendar days (holidays + weekend) before the leave period
#   consecutive calendar days for the leave period +
#   consecutive calendar days (holidays + weekend) after the leave period
# parameter:
# psPeriod - leave slots that need to check for calendar day in the format [{"ldate" : datetime, "ltime": "AM"/ "PM"} ]
# psCombinedSlotLst - leave slots that need to check + all leave applied before + holidays + weekend in the format  [{"ldate" : datetime, "ltime": "AM"/ "PM"} ]
# return :
# No. of calendarDay in int.
# def getCalendarDay(psYear, psOffice, psRecord, psPeriod, psLeaveTypeAttr):
#     weekendHolidays = getWeekendHolidays((psYear - 1), (psYear + 1), psOffice)
#     leaveHistoryLst = getLeaveHistory((psYear - 1), (psYear + 1), psRecord)
#     relatedLveLst = []
#     for lve in leaveTypeLst:
#         if lve["calendar_days_group"] == psLeaveTypeAttr.get("calendar_days_group"):
#             relatedLveLst.append(lve["leave_type_id"])
#     #leaveHistoryLst = list(filter(lambda r: (r["applicationStatus"].upper() != df['gcStatusCancel'][0] and r["applicationStatus"].upper() != df['gcStatusReject'][0]), leaveHistoryLst))
#     leaveHistoryLst = list(filter(lambda r: (r["applicationStatus"].upper() != df['gcStatusCancel'][0] and r["applicationStatus"].upper() != df['gcStatusReject'][0] and r["type"] in relatedLveLst), leaveHistoryLst))
#     combinedSlotLst = combineTime(weekendHolidays, leaveHistoryLst)
#     combinedSlotLst = combineTime(combinedSlotLst, psPeriod)
#     # count the consecutive calendar days (holidays + weekend) before the leave period
#     currPosition = next((index for (index, d) in enumerate(combinedSlotLst) if (d["ldate"] == psPeriod[0]["ldate"]) and (d["ltime"] == psPeriod[0]["ltime"])), None)

#     calendarDayBefore = 0
#     currTime = psPeriod[0]["ltime"]
#     currDate = psPeriod[0]["ldate"]
#     while currPosition != 0:
#         if currDate == combinedSlotLst[currPosition]["ldate"] and currTime == combinedSlotLst[currPosition]["ltime"]:
#             currDate = combinedSlotLst[currPosition]["ldate"]
#             currTime = combinedSlotLst[currPosition]["ltime"]
#             currPosition -= 1
#         elif currDate == combinedSlotLst[currPosition]["ldate"] and currTime == "PM" and combinedSlotLst[currPosition]["ltime"] == "AM":
#             calendarDayBefore += 0.5
#             currDate = combinedSlotLst[currPosition]["ldate"]
#             currTime = combinedSlotLst[currPosition]["ltime"]
#             currPosition -= 1
#         elif currDate == combinedSlotLst[currPosition]["ldate"] + timedelta(days = 1) and currTime == "AM" and combinedSlotLst[currPosition]["ltime"] == "PM":
#             calendarDayBefore += 0.5
#             currDate = combinedSlotLst[currPosition]["ldate"]
#             currTime = combinedSlotLst[currPosition]["ltime"]
#             currPosition -= 1
#         else:
#             break

#     # count the consecutive calendar days (holidays + weekend) after the leave period
#     currTime = psPeriod[-1]["ltime"]
#     currDate = psPeriod[-1]["ldate"]
#     currPosition = next((index for (index, d) in enumerate(combinedSlotLst) if (d["ldate"] == psPeriod[-1]["ldate"]) and (d["ltime"] == psPeriod[-1]["ltime"])), None)
#     calendarDayAfter = 0
#     while currPosition != (len(combinedSlotLst) - 1):
#         if currDate == combinedSlotLst[currPosition]["ldate"] and currTime == combinedSlotLst[currPosition]["ltime"]:
#             currDate = combinedSlotLst[currPosition]["ldate"]
#             currTime = combinedSlotLst[currPosition]["ltime"]
#             currPosition += 1      
#         elif currDate == combinedSlotLst[currPosition]["ldate"] and currTime == "AM" and combinedSlotLst[currPosition]["ltime"] == "PM":
#             calendarDayAfter += 0.5
#             currDate = combinedSlotLst[currPosition]["ldate"]
#             currTime = combinedSlotLst[currPosition]["ltime"]
#             currPosition += 1
#         elif currDate == combinedSlotLst[currPosition]["ldate"] - timedelta(days = 1) and currTime == "PM" and combinedSlotLst[currPosition]["ltime"] == "AM":
#             calendarDayAfter += 0.5
#             currDate = combinedSlotLst[currPosition]["ldate"]
#             currTime = combinedSlotLst[currPosition]["ltime"]
#             currPosition += 1
#         else:
#             break
        
#     # count the consecutive calendar days for the leave period
#     firstPosition = next((index for (index, d) in enumerate(combinedSlotLst) if (d["ldate"] == psPeriod[0]["ldate"]) and (d["ltime"] == psPeriod[0]["ltime"])), None)
#     lastPosition = next((index for (index, d) in enumerate(combinedSlotLst) if (d["ldate"] == psPeriod[-1]["ldate"]) and (d["ltime"] == psPeriod[-1]["ltime"])), None)

#     calendarDayLeave = ((lastPosition - firstPosition) / 2) + 0.5
#     calendarDayTotal = calendarDayBefore + calendarDayLeave + calendarDayAfter


#     return (calendarDayTotal)

# get all date slot for weekend + holidays within the year period
# parameters:
# psYearStart: Beginning year of the weekend and holidays required
# psYearEnd : Ending year of the weekend and holidays required
# psOffice : Office of the holidays required
# return:
# list of date slot for weekends + holidays within the year period, format : [{"ldate": datetime, "ltime": "AM"/ "PM"}]
# def getWeekendHolidays(psYearStart, psYearEnd, psOffice):
#     yr = psYearStart
#     weekendSlotLst = [ ]
#     holidaySlotLst = [ ]
#     while yr <= psYearEnd:
#         weekendSlotLst = combineTime(weekendSlotLst, getAllWeekend(yr))
#         holidaySlotLst = combineTime(holidaySlotLst, getHolidays(yr, psOffice))
#         yr += 1
#     return (combineTime(weekendSlotLst, holidaySlotLst))

# def keep_mail_session(message, title, sendTo, sendCC, attachments = "", attachmentname = ""):

#     # Convert attachments to bytes if provided
#     if attachments:
#         attachment_bytes = []
#         for attachment in attachments:
#             if isinstance(attachment, io.BytesIO):
#                 attachment_bytes.append(attachment.getvalue())
#             else:
#                 attachment_bytes.append(attachment)
#     else:
#         attachment_bytes = []

#     # Create a document to be inserted
#     doc = {
#         'message': message,
#         'subject': title,
#         'sendTo': sendTo,
#         'sendCc': sendCC,
#         'attachment': attachment_bytes,
#         'attachmentname': attachmentname,
#         'send_status': False
#     }

#     # Insert the document into the collection
#     mailsession.insert_one(doc)



def geticalFile(organizer, title, content, startDate, startTime, endDate, endTime, timeZone):
    cal = Calendar()
    cal.add('version', '2.0')

    # Timezone to use for our dates - change as needed
    # should get the timezone from local browser client -> console.log(Intl.DateTimeFormat().resolvedOptions().timeZone)
    if len(timeZone) < 1:
        tz = pytz.timezone("Asia/Shanghai") 
    else:
        tz = pytz.timezone(timeZone) 

    event = Event()
    #event.add('attendee', attendee)
    event.add('organizer', organizer)
    event.add('status', "confirmed")
    event.add('CATEGORIES', vText('Red category'))
    event.add('summary', title)
    event.add('description', content)
    event.add('location', "Online")
    event.add('X-MICROSOFT-CDO-BUSYSTATUS', "FREE")

    if startTime == "AM" and endTime == "AM": 
        hour1 = 8
        mintues1 = 00
        hour2 = 13
        mintues2 = 30
    elif startTime == "AM" and endTime == "PM": 
        hour1 = 8
        mintues1 = 00
        hour2 = 17
        mintues2 = 30
    elif startTime == "PM" and endTime == "PM": 
        hour1 = 13
        mintues1 = 30
        hour2 = 17
        mintues2 = 30
    elif startTime == "PM" and endTime == "AM": 
        hour1 = 13
        mintues1 = 30
        hour2 = 8
        mintues2 = 00


    day1 = datetime.strptime(startDate, '%m/%d/%Y').day
    month1 = datetime.strptime(startDate, '%m/%d/%Y').month
    year1 = datetime.strptime(startDate, '%m/%d/%Y').year

    day2 = datetime.strptime(endDate, '%m/%d/%Y').day
    month2 = datetime.strptime(endDate, '%m/%d/%Y').month
    year2 = datetime.strptime(endDate, '%m/%d/%Y').year   

    if startTime != "AM" and endTime != "PM":
        start = tz.localize(datetime(year1,month1,day1,hour1,mintues1,0))
        end = tz.localize(datetime(year2,month2,day2,hour2,mintues2,0))
        event.add('dtstart', start)
        event.add('dtend', end)
    else:
        start = datetime.strptime(startDate, '%m/%d/%Y')
        end = datetime.strptime(endDate, '%m/%d/%Y')
        end += timedelta(days=1)
        start_date = start.astimezone(tz).strftime('%Y%m%d')
        end_date = end.astimezone(tz).strftime('%Y%m%d')
        event.add('dtstart;VALUE=DATE', start_date)
        event.add('dtend;VALUE=DATE', end_date)

    
    event.add('dtstamp', tz.localize(datetime.now()))
    event.add('created', tz.localize(datetime.now()))

    # Adding events to calendar
    cal.add_component(event)

    #directory = str(Path(__file__).parent.parent) + "/"
    #directory = "./"
    #print("ics file will be generated at ", directory)
    #f = open(os.path.join(directory, 'example.ics'), 'wb')
    #f.write(cal.to_ical())
    #f.close()

    # Output 
    out = BytesIO()
    out.write(cal.to_ical())
    #wb.save(out)
    out.seek(0)

    #out.close()            
    print('sending file...')

    return out

## New add Azure upload at 07/10/26 because of sharepoint issue
# def azureUpload(attachments, metadata):
#     azure_conn = os.environ['AZURE_CONNECTION_STRING']
#     azure_storage_name = os.environ['AZURE_CONTAINER_NAME']
    
#     now = datetime.now()
#     date_str = f"{now.strftime('%Y')}{now.strftime('%m')}{now.strftime('%d')}"

#     for attachment in attachments:
#         try:
            
#             if hasattr(attachment, 'filename'):
#                 base_name = attachment.filename
#                 file_bytes = attachment.read()
#             else:
                
#                 base_name = attachment.get('name', 'unknown_file')
#                 base64_data = attachment.get('base64', '')
#                 if ',' in base64_data:
#                     base64_data = base64_data.split(',')[1]
#                 file_bytes = base64.b64decode(base64_data)
            
#             name_part, ext_part = os.path.splitext(base_name)
#             service_client = BlobServiceClient.from_connection_string(azure_conn)
            
#             new_base_name = f"{name_part}_{metadata['racf']}({date_str})"
#             blob_name = f"{new_base_name}{ext_part}"
            
#             counter = 1
#             blob_client = service_client.get_blob_client(container=azure_storage_name, blob=blob_name)

#             while True:
#                 try:
#                     blob_client.upload_blob(
#                         file_bytes, 
#                         overwrite=False, 
#                         metadata=metadata
#                     )
#                     break
#                 except ResourceExistsError:
#                     blob_name = f"{new_base_name}_{counter}{ext_part}"
#                     blob_client = service_client.get_blob_client(container=azure_storage_name, blob=blob_name)
#                     counter += 1

#         except Exception as e:
#             return ({"result": f"UPLOAD FAILED : {e}", "Status_code": 409})

#     return ({"result": "PASSED", "Status_code": 200})

# def getAzurefiles(sharepoint_id):
#     file_bytes_list = []
#     file_names_list = []
    
#     if not sharepoint_id:
#         return file_bytes_list, file_names_list
        
#     try:
#         azure_conn = os.environ['AZURE_CONNECTION_STRING']
#         azure_storage_name = os.environ['AZURE_CONTAINER_NAME']
        
#         service_client = BlobServiceClient.from_connection_string(azure_conn)
#         container_client = service_client.get_container_client(azure_storage_name)
        
#         blobs = container_client.list_blobs(include=['metadata'])
#         for blob in blobs:
#             if blob.metadata and blob.metadata.get('sharePointId') == str(sharepoint_id):
#                 blob_client = container_client.get_blob_client(blob.name)
#                 download_stream = blob_client.download_blob()
#                 file_bytes = download_stream.readall()
                
#                 file_bytes_list.append(io.BytesIO(file_bytes))
#                 file_names_list.append(blob.name)
                
#     except Exception as e:
#         print(f"Error fetching Azure files: {str(e)}")
#         pass
        
#     return file_bytes_list, file_names_list